"""
Excel处理模块
将策划案内容转换为Excel格式
"""

import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def parse_prd_to_excel_data(prd_content: str) -> list:
    """
    解析策划案文本，转换为Excel数据格式
    按标题层级分配到不同列：
    - 一级标题（如 1、xxx）在第1列
    - 二级标题（如 1.1、xxx）在第2列
    - 三级标题（如 1.1.1、xxx）在第3列
    - 普通内容在最近标题的下一列
    
    Returns:
        list: [(row_data, level), ...] 每行数据和其层级
    """
    lines = prd_content.strip().split('\n')
    excel_data = []
    current_level = 0
    
    # 匹配各级标题的正则表达式
    # 一级标题: 1、 或 1. 或 1  开头（纯数字）
    level1_pattern = re.compile(r'^(\d+)[、\.．]\s*(.+)$')
    # 二级标题: 1.1、 或 1.1. 或 1.1 开头
    level2_pattern = re.compile(r'^(\d+\.\d+)[、\.．]?\s*(.+)$')
    # 三级标题: 1.1.1、 或 1.1.1. 或 1.1.1 开头
    level3_pattern = re.compile(r'^(\d+\.\d+\.\d+)[、\.．]?\s*(.+)$')
    # 四级标题: 1.1.1.1 开头
    level4_pattern = re.compile(r'^(\d+\.\d+\.\d+\.\d+)[、\.．]?\s*(.+)$')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # 检查是否是标题行，从高级别往低级别检查
        level4_match = level4_pattern.match(line)
        level3_match = level3_pattern.match(line)
        level2_match = level2_pattern.match(line)
        level1_match = level1_pattern.match(line)
        
        if level4_match:
            # 四级标题 -> 第4列
            current_level = 4
            excel_data.append((line, 4))
        elif level3_match:
            # 三级标题 -> 第3列
            current_level = 3
            excel_data.append((line, 3))
        elif level2_match:
            # 二级标题 -> 第2列
            current_level = 2
            excel_data.append((line, 2))
        elif level1_match:
            # 一级标题 -> 第1列
            current_level = 1
            excel_data.append((line, 1))
        else:
            # 普通内容 -> 当前标题的下一列，至少在第2列
            content_level = max(current_level + 1, 2) if current_level > 0 else 1
            excel_data.append((line, content_level))
    
    return excel_data


def create_excel_file(prd_content: str, check_result: str = "") -> bytes:
    """
    创建Excel文件
    
    Args:
        prd_content: 策划案内容
        check_result: AI复检结果（可选）
    
    Returns:
        bytes: Excel文件的二进制数据
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "策划案"
    
    # 定义样式
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    level1_font = Font(bold=True, size=12, color="1F4E79")
    level2_font = Font(bold=True, size=11, color="2E75B6")
    level3_font = Font(bold=False, size=10, color="5B9BD5")
    normal_font = Font(size=10)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # 设置列宽
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50
    
    # 添加表头
    headers = ["一级标题", "二级标题/内容", "三级标题/详情", "四级标题/说明", "详细内容"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 解析并填充策划案内容
    excel_data = parse_prd_to_excel_data(prd_content)
    
    row_num = 2
    for content, level in excel_data:
        # 将内容放到对应层级的列
        cell = ws.cell(row=row_num, column=level, value=content)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        
        # 根据层级设置字体样式
        if level == 1:
            cell.font = level1_font
        elif level == 2:
            cell.font = level2_font
        elif level == 3:
            cell.font = level3_font
        else:
            cell.font = normal_font
        
        # 为该行的所有列添加边框
        for col in range(1, 6):
            if col != level:
                empty_cell = ws.cell(row=row_num, column=col, value="")
                empty_cell.border = thin_border
        
        row_num += 1
    
    # 如果有复检结果，添加到新的sheet
    if check_result:
        ws_check = wb.create_sheet(title="AI复检结果")
        ws_check.column_dimensions['A'].width = 100
        
        # 添加标题
        title_cell = ws_check.cell(row=1, column=1, value="AI复检清单检查结果")
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal='center')
        
        # 解析复检结果
        check_lines = check_result.strip().split('\n')
        for idx, line in enumerate(check_lines, 2):
            cell = ws_check.cell(row=idx, column=1, value=line)
            cell.alignment = wrap_alignment
            
            # 根据内容设置样式
            if '✅' in line:
                cell.font = Font(color="228B22")  # 绿色
            elif '⚠️' in line:
                cell.font = Font(color="FF8C00")  # 橙色
            elif '❌' in line:
                cell.font = Font(color="DC143C")  # 红色
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
