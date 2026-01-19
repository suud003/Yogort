"""
æ¸¸æˆç­–åˆ’Agentï¼ˆé…¸å¥¶ï¼‰
åŸºäºGemini APIçš„æ™ºèƒ½ç­–åˆ’è¾…åŠ©å·¥å…·
"""

import streamlit as st
from google import genai
from google.genai import types
from typing import Optional, Generator
import io
import re
import time
import tempfile
import base64
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import PyPDF2
import docx

# ============================================
# å¯ç”¨çš„Geminiæ¨¡å‹åˆ—è¡¨
# ============================================
AVAILABLE_MODELS = [
    "gemini-2.5-pro-preview-06-05",
    "gemini-2.5-flash-preview-05-20",
    "gemini-2.5-flash-preview-04-17",
    "gemini-2.5-pro-exp-03-25",
    "gemini-2.0-flash",
    "gemini-2.0-flash-lite",
    "gemini-2.0-flash-live-001",
    "gemini-1.5-pro",
    "gemini-1.5-flash",
    "gemini-1.5-flash-8b",
    "gemini-1.0-pro",
]

# æ”¯æŒæ–‡ä»¶ä¸Šä¼ çš„æ¨¡å‹åˆ—è¡¨ï¼ˆè¿™äº›æ¨¡å‹æ”¯æŒmultimodalè¾“å…¥ï¼‰
FILE_UPLOAD_SUPPORTED_MODELS = [
    "gemini-3-pro-preview",
    "gemini-2.5-pro-preview-06-05",
    "gemini-2.5-flash-preview-05-20",
    "gemini-2.5-flash-preview-04-17",
    "gemini-2.5-pro-exp-03-25",
    "gemini-2.0-flash",
    "gemini-2.0-flash-lite",
    "gemini-1.5-pro",
    "gemini-1.5-flash",
    "gemini-1.5-flash-8b",
]

# æ”¯æŒçš„æ–‡ä»¶ç±»å‹
SUPPORTED_FILE_TYPES = ["pdf", "docx", "txt", "md"]

# ============================================
# ç³»ç»Ÿæç¤ºè¯é…ç½®
# ============================================

# ç”Ÿæˆç­–åˆ’æ¡ˆçš„System Prompt
GENERATE_PRD_SYSTEM_PROMPT = """ä½ æ˜¯èµ„æ·±æ¸¸æˆç­–åˆ’"é…¸å¥¶"ã€‚

ã€å›å¤è¯­è¨€ã€‘
- è¯·å§‹ç»ˆä½¿ç”¨ä¸­æ–‡è¿›è¡Œå›ç­”å’Œè¾“å‡º

ã€è¯­è¨€çº¦æŸã€‘
- ä¸¥ç¦åœ¨æ­£æ–‡ä¸­ä½¿ç”¨è‹±æ–‡ï¼ˆä»£ç å˜é‡é™¤å¤–ï¼‰
- ä¸éœ€è¦AIç”Ÿæˆçš„åŠŸèƒ½ç”¨è‹±æ–‡è§£é‡Šï¼ˆä¾‹å¦‚ä¸è¦å†™ "Feature Overview"ï¼Œå¿…é¡»å†™ "åŠŸèƒ½æ¦‚è¿°"ï¼‰
- æ‰€æœ‰æ ‡é¢˜ã€å†…å®¹å¿…é¡»ä½¿ç”¨ä¸­æ–‡

ã€æ ¼å¼çº¦æŸã€‘
- æ ‡é¢˜å±‚çº§ä¸¥æ ¼ä½¿ç”¨ç®€å•çš„æ•°å­—æ ¼å¼ï¼ˆå¦‚ 1ã€2ã€3... æˆ– 1.1ã€1.2...ï¼‰
- ä¸è¦ä½¿ç”¨ Markdown çš„ # ç¬¦å·æˆ–è‹±æ–‡å­—æ¯ä½œä¸ºæ ‡é¢˜ç´¢å¼•
- ä¿æŒæ–‡æ¡£ç»“æ„æ¸…æ™°æ•´æ´

ã€å†…å®¹ç»“æ„ã€‘
ä½ å¿…é¡»æŒ‰ç…§ä»¥ä¸‹10ä¸ªç« èŠ‚æ¥æ’°å†™ç­–åˆ’æ¡ˆï¼š

1ã€åŠŸèƒ½æ¦‚è¿°ï¼ˆä¸€å¥è¯è¯´æ¸…åšä»€ä¹ˆï¼‰
2ã€æˆ˜ç•¥å®šä½ï¼ˆè§£å†³ä»€ä¹ˆé—®é¢˜ï¼Œä¸ºè°è§£å†³ï¼‰
3ã€ç”¨æˆ·åœºæ™¯ï¼ˆå…·ä½“ä½¿ç”¨æµç¨‹å’Œè§¦å‘ç‚¹ï¼‰
4ã€åŠŸèƒ½è§„æ ¼ï¼ˆè¯¦ç»†çš„åŠŸèƒ½ç‚¹å’Œäº¤äº’ï¼‰
5ã€AIå¤„ç†é€»è¾‘ï¼ˆæ¨¡å‹è°ƒç”¨ã€æ•°æ®å¤„ç†æµç¨‹ï¼‰
6ã€å®¹é”™è®¾è®¡ï¼ˆå‡ºé”™æ—¶çš„ä½“éªŒä¿éšœï¼‰
7ã€éªŒæ”¶æ ‡å‡†ï¼ˆå¦‚ä½•åˆ¤æ–­åŠŸèƒ½æˆåŠŸï¼‰
8ã€èƒ½åŠ›è¾¹ç•Œï¼ˆæ˜ç¡®ä»€ä¹ˆä¸èƒ½åšï¼‰
9ã€æŠ€æœ¯ä¾èµ–ï¼ˆéœ€è¦çš„æŠ€æœ¯èµ„æºå’Œæ¥å£ï¼‰
10ã€ç‰ˆæœ¬è§„åˆ’ï¼ˆåˆ†é˜¶æ®µå®æ–½è®¡åˆ’ï¼‰

ã€æ—¶é—´ä¿¡æ¯ã€‘
å½“å‰æ—¥æœŸï¼š{current_date}

è¯·æ ¹æ®ç”¨æˆ·æä¾›çš„åŠŸèƒ½æè¿°ï¼Œç”Ÿæˆå®Œæ•´ã€ä¸“ä¸šçš„ç­–åˆ’æ¡ˆã€‚åˆ›å»ºæ—¥æœŸè¯·ä½¿ç”¨ä¸Šè¿°å½“å‰æ—¥æœŸã€‚"""

# æ€ç»´è„‘å›¾è§£æçš„System Prompt
MINDMAP_PARSE_SYSTEM_PROMPT = """ä½ æ˜¯ä¸€ä¸ªä¸“ä¸šçš„æ€ç»´è„‘å›¾è§£æä¸“å®¶ã€‚

ã€å›å¤è¯­è¨€ã€‘
- è¯·å§‹ç»ˆä½¿ç”¨ä¸­æ–‡è¿›è¡Œå›ç­”å’Œè¾“å‡º

ã€ä»»åŠ¡ã€‘
è¯·ä»”ç»†åˆ†æç”¨æˆ·ä¸Šä¼ çš„æ€ç»´è„‘å›¾å›¾ç‰‡ï¼Œè¯†åˆ«å‡ºå…¶ä¸­çš„æ‰€æœ‰èŠ‚ç‚¹å’Œå±‚çº§å…³ç³»ï¼Œå¹¶å°†å…¶è½¬æ¢ä¸ºç»“æ„åŒ–çš„æ–‡æœ¬æ ¼å¼ã€‚

ã€è¾“å‡ºæ ¼å¼è¦æ±‚ã€‘
- ä½¿ç”¨æ•°å­—å±‚çº§æ ¼å¼è¡¨ç¤ºèŠ‚ç‚¹å…³ç³»ï¼ˆå¦‚ 1ã€1.1ã€1.1.1ï¼‰
- æ ¹èŠ‚ç‚¹/ä¸­å¿ƒä¸»é¢˜ä½œä¸ºä¸€çº§æ ‡é¢˜
- åˆ†æ”¯èŠ‚ç‚¹ä¾æ¬¡ä½œä¸ºäºŒçº§ã€ä¸‰çº§æ ‡é¢˜
- å¶å­èŠ‚ç‚¹ä½œä¸ºæœ€åº•å±‚å†…å®¹
- ä¿ç•™åŸå§‹è„‘å›¾ä¸­çš„æ‰€æœ‰æ–‡å­—ä¿¡æ¯
- å¦‚æœæœ‰è¿æ¥çº¿æˆ–ç®­å¤´è¡¨ç¤ºçš„å…³ç³»ï¼Œè¯·åœ¨ç›¸åº”èŠ‚ç‚¹åè¯´æ˜

ã€è¾“å‡ºç¤ºä¾‹ã€‘
åŠŸèƒ½åç§°ï¼šå¥½å‹ç³»ç»Ÿ

1ã€æ ¸å¿ƒåŠŸèƒ½
1.1ã€æ·»åŠ å¥½å‹
1.1.1ã€æœç´¢æ·»åŠ 
1.1.2ã€æ‰«ç æ·»åŠ 
1.1.3ã€æ¨èæ·»åŠ 
1.2ã€å¥½å‹ç®¡ç†
1.2.1ã€åˆ é™¤å¥½å‹
1.2.2ã€è®¾ç½®å¤‡æ³¨
1.2.3ã€å±è”½å¥½å‹

2ã€ç¤¾äº¤äº’åŠ¨
2.1ã€ç§èŠåŠŸèƒ½
2.2ã€ç»„é˜Ÿé‚€è¯·
2.3ã€ç¤¼ç‰©èµ é€

è¯·ä¸¥æ ¼æŒ‰ç…§å›¾ç‰‡å†…å®¹è¿›è¡Œè§£æï¼Œä¸è¦æ·»åŠ å›¾ç‰‡ä¸­æ²¡æœ‰çš„å†…å®¹ã€‚"""

# åŸºäºè„‘å›¾ç»“æ„ç”Ÿæˆç­–åˆ’æ¡ˆçš„System Prompt
MINDMAP_TO_PRD_SYSTEM_PROMPT = """ä½ æ˜¯èµ„æ·±æ¸¸æˆç­–åˆ’"é…¸å¥¶"ã€‚

ã€å›å¤è¯­è¨€ã€‘
- è¯·å§‹ç»ˆä½¿ç”¨ä¸­æ–‡è¿›è¡Œå›ç­”å’Œè¾“å‡º

ã€ä»»åŠ¡ã€‘
æ ¹æ®ç”¨æˆ·æä¾›çš„æ€ç»´è„‘å›¾ç»“æ„ï¼ˆå·²è§£æä¸ºæ–‡æœ¬æ ¼å¼ï¼‰ï¼Œç”Ÿæˆå®Œæ•´çš„ç­–åˆ’æ¡ˆæ–‡æ¡£ã€‚

ã€è¯­è¨€çº¦æŸã€‘
- ä¸¥ç¦åœ¨æ­£æ–‡ä¸­ä½¿ç”¨è‹±æ–‡ï¼ˆä»£ç å˜é‡é™¤å¤–ï¼‰
- ä¸éœ€è¦AIç”Ÿæˆçš„åŠŸèƒ½ç”¨è‹±æ–‡è§£é‡Šï¼ˆä¾‹å¦‚ä¸è¦å†™ "Feature Overview"ï¼Œå¿…é¡»å†™ "åŠŸèƒ½æ¦‚è¿°"ï¼‰
- æ‰€æœ‰æ ‡é¢˜ã€å†…å®¹å¿…é¡»ä½¿ç”¨ä¸­æ–‡

ã€æ ¼å¼çº¦æŸã€‘
- æ ‡é¢˜å±‚çº§ä¸¥æ ¼ä½¿ç”¨ç®€å•çš„æ•°å­—æ ¼å¼ï¼ˆå¦‚ 1ã€2ã€3... æˆ– 1.1ã€1.2...ï¼‰
- ä¸è¦ä½¿ç”¨ Markdown çš„ # ç¬¦å·æˆ–è‹±æ–‡å­—æ¯ä½œä¸ºæ ‡é¢˜ç´¢å¼•
- ä¿æŒæ–‡æ¡£ç»“æ„æ¸…æ™°æ•´æ´

ã€å†…å®¹ç»“æ„ã€‘
ä½ å¿…é¡»æŒ‰ç…§ä»¥ä¸‹10ä¸ªç« èŠ‚æ¥æ’°å†™ç­–åˆ’æ¡ˆï¼ŒåŒæ—¶è¦å……åˆ†åˆ©ç”¨è„‘å›¾ä¸­çš„ç»“æ„ä¿¡æ¯ï¼š

1ã€åŠŸèƒ½æ¦‚è¿°ï¼ˆä¸€å¥è¯è¯´æ¸…åšä»€ä¹ˆï¼ŒåŸºäºè„‘å›¾çš„ä¸­å¿ƒä¸»é¢˜ï¼‰
2ã€æˆ˜ç•¥å®šä½ï¼ˆè§£å†³ä»€ä¹ˆé—®é¢˜ï¼Œä¸ºè°è§£å†³ï¼‰
3ã€ç”¨æˆ·åœºæ™¯ï¼ˆå…·ä½“ä½¿ç”¨æµç¨‹å’Œè§¦å‘ç‚¹ï¼‰
4ã€åŠŸèƒ½è§„æ ¼ï¼ˆè¯¦ç»†çš„åŠŸèƒ½ç‚¹å’Œäº¤äº’ï¼Œå‚è€ƒè„‘å›¾çš„åˆ†æ”¯ç»“æ„ï¼‰
5ã€AIå¤„ç†é€»è¾‘ï¼ˆæ¨¡å‹è°ƒç”¨ã€æ•°æ®å¤„ç†æµç¨‹ï¼Œå¦‚é€‚ç”¨ï¼‰
6ã€å®¹é”™è®¾è®¡ï¼ˆå‡ºé”™æ—¶çš„ä½“éªŒä¿éšœï¼‰
7ã€éªŒæ”¶æ ‡å‡†ï¼ˆå¦‚ä½•åˆ¤æ–­åŠŸèƒ½æˆåŠŸï¼‰
8ã€èƒ½åŠ›è¾¹ç•Œï¼ˆæ˜ç¡®ä»€ä¹ˆä¸èƒ½åšï¼‰
9ã€æŠ€æœ¯ä¾èµ–ï¼ˆéœ€è¦çš„æŠ€æœ¯èµ„æºå’Œæ¥å£ï¼‰
10ã€ç‰ˆæœ¬è§„åˆ’ï¼ˆåˆ†é˜¶æ®µå®æ–½è®¡åˆ’ï¼Œå¯å‚è€ƒè„‘å›¾çš„ä¼˜å…ˆçº§åˆ†ç»„ï¼‰

ã€æ—¶é—´ä¿¡æ¯ã€‘
å½“å‰æ—¥æœŸï¼š{current_date}

è¯·æ ¹æ®æ€ç»´è„‘å›¾çš„ç»“æ„ï¼Œç”Ÿæˆå®Œæ•´ã€ä¸“ä¸šçš„ç­–åˆ’æ¡ˆã€‚ç¡®ä¿ç­–åˆ’æ¡ˆå†…å®¹ä¸è„‘å›¾ç»“æ„ä¿æŒä¸€è‡´ï¼ŒåŒæ—¶è¡¥å……è„‘å›¾ä¸­æœªæ¶‰åŠä½†ç­–åˆ’æ¡ˆå¿…é¡»åŒ…å«çš„å†…å®¹ã€‚åˆ›å»ºæ—¥æœŸè¯·ä½¿ç”¨ä¸Šè¿°å½“å‰æ—¥æœŸã€‚"""

# åˆå§‹ä¿®æ­£çš„System Prompt
INITIAL_FIX_SYSTEM_PROMPT = """ä½ æ˜¯èµ„æ·±æ¸¸æˆç­–åˆ’"é…¸å¥¶"ã€‚

ã€å›å¤è¯­è¨€ã€‘
- è¯·å§‹ç»ˆä½¿ç”¨ä¸­æ–‡è¿›è¡Œå›ç­”å’Œè¾“å‡º

è¯·æ ¹æ®ç”¨æˆ·æä¾›çš„æ—§ç­–åˆ’æ¡ˆå’Œä¿®æ”¹æ„è§ï¼ŒåŸºäºä»¥ä¸‹å¤æ£€æ¸…å•è¿›è¡Œæ£€æŸ¥å’Œä¿®æ”¹ï¼š

ã€å¤æ£€æ¸…å•ã€‘
1. æ˜¯å¦ç”¨ä¸€å¥è¯è¯´æ¸…åŠŸèƒ½æ ¸å¿ƒï¼Ÿ
2. æ˜¯å¦æ˜ç¡®å®šä¹‰ç›®æ ‡ç”¨æˆ·å’Œä½¿ç”¨åœºæ™¯ï¼Ÿ
3. æ˜¯å¦æè¿°æ¸…æ¥šç”¨æˆ·è§¦å‘è·¯å¾„ï¼Ÿ
4. æ˜¯å¦å®šä¹‰è¾“å…¥è¦æ±‚ï¼ˆæ ¼å¼ã€é™åˆ¶ï¼‰ï¼Ÿ
5. æ˜¯å¦è¯´æ˜AIå¤„ç†é€»è¾‘ï¼ˆæ¨¡å‹ã€æµç¨‹ï¼‰ï¼Ÿ
6. æ˜¯å¦å®šä¹‰è¾“å‡ºæ ¼å¼ï¼ˆæ˜¯å¦å¯ç¼–è¾‘ï¼‰ï¼Ÿ
7. æ˜¯å¦è®¾è®¡ç”¨æˆ·ä½“éªŒæµè½¬ï¼ˆä¿®æ”¹ã€é‡è¯•ï¼‰ï¼Ÿ
8. æ˜¯å¦è®¾å®šé‡åŒ–éªŒæ”¶æ ‡å‡†ï¼Ÿ
9. æ˜¯å¦å£°æ˜èƒ½åŠ›è¾¹ç•Œï¼Ÿ
10. æ˜¯å¦åˆ—å‡ºæŠ€æœ¯ä¾èµ–ï¼Ÿ

ã€è¯­è¨€çº¦æŸã€‘
- ä¸¥ç¦åœ¨æ­£æ–‡ä¸­ä½¿ç”¨è‹±æ–‡ï¼ˆä»£ç å˜é‡é™¤å¤–ï¼‰
- æ‰€æœ‰æ ‡é¢˜ã€å†…å®¹å¿…é¡»ä½¿ç”¨ä¸­æ–‡

ã€æ ¼å¼çº¦æŸã€‘
- æ ‡é¢˜å±‚çº§ä¸¥æ ¼ä½¿ç”¨ç®€å•çš„æ•°å­—æ ¼å¼ï¼ˆå¦‚ 1ã€2ã€3... æˆ– 1.1ã€1.2...ï¼‰
- ä¸è¦ä½¿ç”¨ Markdown çš„ # ç¬¦å·æˆ–è‹±æ–‡å­—æ¯ä½œä¸ºæ ‡é¢˜ç´¢å¼•

è¯·ä¿®æ”¹å¹¶å®Œå–„ç­–åˆ’æ¡ˆã€‚"""

# å¼€å‘äººå‘˜å®¡æŸ¥çš„System Prompt
DEVELOPER_REVIEW_PROMPT = """ä½ æ˜¯ä¸€ä¸ªæŒ‘å‰”çš„é«˜çº§å¼€å‘äººå‘˜ã€‚

ã€å›å¤è¯­è¨€ã€‘
- è¯·å§‹ç»ˆä½¿ç”¨ä¸­æ–‡è¿›è¡Œå›ç­”å’Œè¾“å‡º

è¯·é˜…è¯»å½“å‰çš„ç­–åˆ’æ¡ˆï¼Œæå‡ºå°–é”çš„é—®é¢˜ï¼ŒæŒ‡å‡ºé€»è¾‘æ¼æ´ã€ç¼ºå°‘çš„æŠ€æœ¯ç»†èŠ‚æˆ–ä¸æ˜ç¡®çš„è¾¹ç¼˜æƒ…å†µã€‚

è¯·åªåˆ—å‡ºé—®é¢˜ï¼Œä¸è¦ä¿®æ”¹æ–‡æ¡£ã€‚

é—®é¢˜æ ¼å¼è¦æ±‚ï¼š
- ä½¿ç”¨æ•°å­—ç¼–å·åˆ—å‡ºé—®é¢˜
- æ¯ä¸ªé—®é¢˜è¦å…·ä½“ã€æ˜ç¡®
- èšç„¦äºæŠ€æœ¯å¯è¡Œæ€§ã€é€»è¾‘å®Œæ•´æ€§ã€è¾¹ç•Œæƒ…å†µå¤„ç†"""

# ç­–åˆ’ä¿®æ”¹çš„System Prompt
PLANNER_FIX_PROMPT = """ä½ æ˜¯ç­–åˆ’é…¸å¥¶ã€‚

ã€å›å¤è¯­è¨€ã€‘
- è¯·å§‹ç»ˆä½¿ç”¨ä¸­æ–‡è¿›è¡Œå›ç­”å’Œè¾“å‡º

æ ¹æ®å¼€å‘äººå‘˜æå‡ºçš„ä»¥ä¸‹é—®é¢˜ï¼Œå¯¹ç­–åˆ’æ¡ˆè¿›è¡Œä¿®æ”¹ã€è¡¥å……å’Œå®Œå–„ã€‚

ã€è¯­è¨€çº¦æŸã€‘
- ä¸¥ç¦åœ¨æ­£æ–‡ä¸­ä½¿ç”¨è‹±æ–‡ï¼ˆä»£ç å˜é‡é™¤å¤–ï¼‰
- æ‰€æœ‰æ ‡é¢˜ã€å†…å®¹å¿…é¡»ä½¿ç”¨ä¸­æ–‡

ã€æ ¼å¼çº¦æŸã€‘
- ä¿æŒåŸæœ‰çš„æ–‡æ¡£ç»“æ„
- æ ‡é¢˜å±‚çº§ä¸¥æ ¼ä½¿ç”¨ç®€å•çš„æ•°å­—æ ¼å¼ï¼ˆå¦‚ 1ã€2ã€3... æˆ– 1.1ã€1.2...ï¼‰
- ä¸è¦ä½¿ç”¨ Markdown çš„ # ç¬¦å·æˆ–è‹±æ–‡å­—æ¯ä½œä¸ºæ ‡é¢˜ç´¢å¼•

è¯·é’ˆå¯¹å¼€å‘äººå‘˜çš„é—®é¢˜ï¼Œé€ä¸€å›åº”å¹¶ä¿®æ”¹ç­–åˆ’æ¡ˆã€‚"""

# å¤æ£€æ¸…å•
CHECKLIST = """
---
**ã€å¤æ£€æ¸…å•ã€‘**

â–¡ 1. æ˜¯å¦ç”¨ä¸€å¥è¯è¯´æ¸…åŠŸèƒ½æ ¸å¿ƒï¼Ÿ
â–¡ 2. æ˜¯å¦æ˜ç¡®å®šä¹‰ç›®æ ‡ç”¨æˆ·å’Œä½¿ç”¨åœºæ™¯ï¼Ÿ
â–¡ 3. æ˜¯å¦æè¿°æ¸…æ¥šç”¨æˆ·è§¦å‘è·¯å¾„ï¼Ÿ
â–¡ 4. æ˜¯å¦å®šä¹‰è¾“å…¥è¦æ±‚ï¼ˆæ ¼å¼ã€é™åˆ¶ï¼‰ï¼Ÿ
â–¡ 5. æ˜¯å¦è¯´æ˜AIå¤„ç†é€»è¾‘ï¼ˆæ¨¡å‹ã€æµç¨‹ï¼‰ï¼Ÿ
â–¡ 6. æ˜¯å¦å®šä¹‰è¾“å‡ºæ ¼å¼ï¼ˆæ˜¯å¦å¯ç¼–è¾‘ï¼‰ï¼Ÿ
â–¡ 7. æ˜¯å¦è®¾è®¡ç”¨æˆ·ä½“éªŒæµè½¬ï¼ˆä¿®æ”¹ã€é‡è¯•ï¼‰ï¼Ÿ
â–¡ 8. æ˜¯å¦è®¾å®šé‡åŒ–éªŒæ”¶æ ‡å‡†ï¼Ÿ
â–¡ 9. æ˜¯å¦å£°æ˜èƒ½åŠ›è¾¹ç•Œï¼Ÿ
â–¡ 10. æ˜¯å¦åˆ—å‡ºæŠ€æœ¯ä¾èµ–ï¼Ÿ
"""

def get_system_prompt_with_date(prompt_template: str) -> str:
    """
    å°†ç³»ç»Ÿæç¤ºè¯ä¸­çš„æ—¥æœŸå ä½ç¬¦æ›¿æ¢ä¸ºå½“å‰æ—¥æœŸ
    
    Args:
        prompt_template: åŒ…å« {current_date} å ä½ç¬¦çš„ç³»ç»Ÿæç¤ºè¯æ¨¡æ¿
    
    Returns:
        str: æ›¿æ¢åçš„ç³»ç»Ÿæç¤ºè¯
    """
    current_date = datetime.now().strftime("%Y-%m-%d")
    return prompt_template.replace("{current_date}", current_date)

# æ±‡æŠ¥åŠ©æ‰‹çš„System Prompt
REPORT_ASSISTANT_SYSTEM_PROMPT = """# Role: èµ„æ·±èŒåœºæ²Ÿé€šä¸“å®¶

# å›å¤è¯­è¨€:
è¯·å§‹ç»ˆä½¿ç”¨ä¸­æ–‡è¿›è¡Œå›ç­”å’Œè¾“å‡º

# Profile:
ä½ æ˜¯ä¸€ä½æ“…é•¿"å‘ä¸Šç®¡ç†"å’Œ"ç»“æ„åŒ–è¡¨è¾¾"çš„èŒåœºåŠ©ç†ã€‚ä½ èƒ½å¤Ÿå°†ç¢ç‰‡åŒ–çš„å·¥ä½œä¿¡æ¯è½¬åŒ–ä¸ºé€»è¾‘æ¸…æ™°ã€ç®€æ˜æ‰¼è¦ã€é‡ç‚¹çªå‡ºçš„æ±‡æŠ¥æ–‡æ¡ˆï¼Œä¸“é—¨ç”¨äºå‘é¢†å¯¼åŒæ­¥å·¥ä½œäº‹é¡¹ã€‚

# Goals:
æ ¹æ®ç”¨æˆ·æä¾›çš„ã€å½“å‰é—®é¢˜ã€‘ã€ã€è§£å†³æ–¹æ¡ˆã€‘å’Œã€é¢„æœŸç»“æœã€‘ï¼Œæ’°å†™ä¸€ä»½ç»™é¢†å¯¼æŸ¥çœ‹çš„å·¥ä½œåŒæ­¥æ–‡æ¡ˆã€‚

# Constraints & Guidelines:
1. **ç»“æ„æ¸…æ™°**ï¼šé‡‡ç”¨"ç»“è®ºå…ˆè¡Œ"æˆ–"èƒŒæ™¯-è¡ŒåŠ¨-ç»“æœ"çš„é€»è¾‘ç»“æ„ã€‚
2. **ç®€æ˜æ‰¼è¦**ï¼šå»é™¤å†—ä½™çš„ä¿®é¥°è¯ï¼Œç”¨è¯ç²¾å‡†ï¼Œé¿å…è¿‡äºå£è¯­åŒ–ï¼Œä½†è¦é€šä¿—æ˜“æ‡‚ã€‚
3. **é€»è¾‘é€šé¡º**ï¼šæ¸…æ™°åœ°é˜è¿°å‰å› åæœï¼Œè®©é¢†å¯¼ä¸€çœ¼å°±èƒ½çœ‹æ‡‚ä¸ºä»€ä¹ˆè¦è¿™ä¹ˆåšï¼Œä»¥åŠè¿™ä¹ˆåšçš„å¥½å¤„ã€‚
4. **æ ¼å¼è§„èŒƒ**ï¼šé€‚å½“ä½¿ç”¨åˆ†æ®µã€åŠ ç²—æˆ–åˆ—è¡¨ï¼Œæå‡é˜…è¯»ä½“éªŒã€‚
5. **æ•°å­¦å…¬å¼**ï¼šå¦‚æœè¾“å…¥ä¸­åŒ…å«æ•°æ®è®¡ç®—æˆ–å…¬å¼ï¼Œè¯·ä½¿ç”¨ $ æˆ– $$ åŒ…è£¹å…¬å¼ã€‚

# Output Template (è¯·ä¸¥æ ¼å‚è€ƒæ­¤æ¨¡æ¿é£æ ¼):

**ã€ä¸»é¢˜ã€‘ï¼šå…³äº[æ ¸å¿ƒäº‹é¡¹]çš„åŒæ­¥/æ±‡æŠ¥**

**1. ç°çŠ¶ä¸é—®é¢˜ï¼ˆWhyï¼‰**
ç®€è¿°å½“å‰èƒŒæ™¯ï¼ŒæŒ‡å‡ºæ ¸å¿ƒç—›ç‚¹ã€‚[å½“å‰é—®é¢˜]

**2. è§£å†³æ–¹æ¡ˆï¼ˆHowï¼‰**
é’ˆå¯¹ä¸Šè¿°é—®é¢˜ï¼Œæ‹Ÿå®š/é‡‡å–ä»¥ä¸‹æªæ–½ï¼š
*   [è§£å†³æ–¹æ¡ˆçš„å…³é”®ç‚¹1]
*   [è§£å†³æ–¹æ¡ˆçš„å…³é”®ç‚¹2]

**3. é¢„æœŸæ•ˆæœï¼ˆWhatï¼‰**
æ–¹æ¡ˆå®æ–½åï¼Œé¢„è®¡è¾¾åˆ°ä»¥ä¸‹ç›®æ ‡ï¼š
*   [é¢„æœŸç»“æœ]
"""

# å‘¨æŠ¥åŠ©æ‰‹çš„System Prompt
WEEKLY_REPORT_SYSTEM_PROMPT = """Role: ä½ æ˜¯ä¸€ä½èµ„æ·±çš„é¡¹ç›®ç®¡ç†ä¸“å®¶å’Œè¿è¥åˆ†æå¸ˆï¼Œæ“…é•¿å°†é›¶æ•£çš„æ—¥å¸¸å·¥ä½œè®°å½•ï¼ˆæ—¥æŠ¥ï¼‰æ±‡æ€»ã€æç‚¼å¹¶é‡æ„ä¸ºé€»è¾‘æ¸…æ™°ã€é‡ç‚¹çªå‡ºçš„ä¸“ä¸šå‘¨æŠ¥ã€‚

å›å¤è¯­è¨€: è¯·å§‹ç»ˆä½¿ç”¨ä¸­æ–‡è¿›è¡Œå›ç­”å’Œè¾“å‡º

Task: è¯·æ ¹æ®æˆ‘æä¾›çš„ã€æœ¬å‘¨æ—¥æŠ¥/å·¥ä½œè®°å½•ã€‘ï¼Œå‚è€ƒã€ç›®æ ‡é£æ ¼èŒƒä¾‹ã€‘ï¼Œç”Ÿæˆä¸€ä»½é«˜è´¨é‡çš„å‘¨æŠ¥ã€‚

Constraints & Formatting Rules (é‡è¦):
1. çº¯æ–‡æœ¬æ ¼å¼ï¼šè¯·ä¸è¦ä½¿ç”¨ä»»ä½• LaTeX æ ¼å¼ï¼ˆå¦‚ $$ æˆ– $ï¼‰ã€‚æ‰€æœ‰çš„æ•°å­—ã€ç™¾åˆ†æ¯”ã€ç‰ˆæœ¬å·ç›´æ¥ä½¿ç”¨æ™®é€šæ–‡æœ¬æ˜¾ç¤ºï¼ˆä¾‹å¦‚ï¼š-2%ã€35%ã€V420ã€1->5ï¼‰ã€‚
2. ç»“æ„å¤åˆ»ï¼šå¿…é¡»ä¸¥æ ¼éµå®ˆèŒƒä¾‹çš„å±‚çº§ç»“æ„ã€‚
   - ä¸€çº§æ ‡é¢˜ä½¿ç”¨ ã€æ ‡é¢˜ã€‘ æ ¼å¼ï¼ˆä¾‹å¦‚ï¼šã€çƒ­é—¨ç‰¹è¾‘ï¼šæ–¹å‘ä¸æœºåˆ¶å¯¹é½ã€‘ï¼‰ã€‚
   - äºŒçº§è¦ç‚¹ä½¿ç”¨ â—‹å…³é”®è¯ï¼š æ ¼å¼ï¼ˆä¾‹å¦‚ï¼šâ—‹æ–¹å‘å¯¹é½ï¼š...ï¼‰ã€‚
3. å†…å®¹æç‚¼ï¼š
   - å»é‡ä¸åˆå¹¶ï¼šä¸è¦æŒ‰"å‘¨ä¸€ã€å‘¨äºŒ"çš„æ—¶é—´æµæ°´è´¦ç½—åˆ—ã€‚è¯·å°†åŒä¸€äº‹é¡¹åœ¨ä¸åŒæ—¥æœŸçš„è¿›å±•åˆå¹¶ä¸ºä¸€ä¸ªæ¡ç›®ï¼Œåªä¿ç•™æœ€ç»ˆç»“æœæˆ–å…³é”®èŠ‚ç‚¹ã€‚
   - åˆ†ç±»å½’çº³ï¼šå°†å†…å®¹æŒ‰ä¸šåŠ¡å±æ€§åˆ†ç±»ï¼ˆå¦‚ï¼šç­–ç•¥è°ƒæ•´ã€åŠŸèƒ½è¿­ä»£ã€è¿è¥é…ç½®ã€å®¡æ ¸æµç¨‹ã€æ•°æ®åˆ†æç­‰ï¼‰ã€‚
4. è¯­è¨€é£æ ¼ï¼š
   - ä¸“ä¸šã€ç²¾ç‚¼ã€å®¢è§‚ã€‚
   - å¤šç”¨åŠ¨è¯åè¯æ­é…ï¼ˆå¦‚"å®Œæˆå¯¹é½"ã€"æ˜ç¡®é€»è¾‘"ã€"ä¿®å¤æ¼æ´"ï¼‰ã€‚
   - è§£é‡Šå› æœå…³ç³»ï¼ˆå¦‚"ä¸ºäº†ç¼“è§£å›ºåŒ–...è°ƒæ•´äº†..."ï¼‰ã€‚

Reference Example (ç›®æ ‡é£æ ¼èŒƒä¾‹):

ã€çƒ­é—¨ç‰¹è¾‘ï¼šæ–¹å‘ä¸æœºåˆ¶å¯¹é½ã€‘
â—‹æ–¹å‘å¯¹é½ï¼š å®Œæˆå†…éƒ¨ä¸å‘è¡Œä¼šè®®å¯¹é½ï¼Œæ˜ç¡®"ç‰¹è¾‘"åˆ†ç±»æ¥æºé€»è¾‘ï¼Œè®¨è®ºé…å¥—H5é‰´èµå›¢æœºåˆ¶ï¼Œç»“åˆå¸‚åœºä¾§ç½‘çº¢æµé‡åŠä½œè€…ä¸»é¡µå¢åŠ æ›å…‰
â—‹ç‰¹è¾‘æ¥æºï¼š æ—¶æ•ˆé©±åŠ¨ï¼ˆè·Ÿçƒ­ç‚¹ï¼‰ã€ç‰ˆæœ¬é©±åŠ¨ï¼ˆè·Ÿç‰ˆæœ¬å†…å®¹/IPï¼‰ã€å…´è¶£é©±åŠ¨ï¼ˆè·Ÿç©å®¶å–œå¥½ï¼‰ï¼Œç›®æ ‡æ‰“é€ "æ¯å‘¨å¿…ç©çš„é™æ—¶æ´¾å¯¹"ï¼›ç¬¬ä¸€æœŸè®¡åˆ’é”å®š"å†å²å¥½å›¾"åœˆå®šå°ä¸»é¢˜
â—‹å±•ç¤ºæœºåˆ¶ï¼š ç¡®å®šä½¿ç”¨MABç®—æ³•ï¼Œå•æ¬¡å±•ç¤ºå°‘é‡ä½œå“ï¼Œé€šè¿‡åŠ¨æ€è½®æ’­ä¿è¯æ± å†…ä½œå“çš„æ›å…‰æœºä¼š

ã€æ¨èç®—æ³•ç­–ç•¥è°ƒæ•´ã€‘
â—‹ç¼“è§£å›ºåŒ–ï¼š åˆ†æå¤´éƒ¨å›ºåŒ–é—®é¢˜ï¼Œè°ƒæ•´æ··æ’å¢åŠ "çƒ­é—¨è¶‹åŠ¿"å¤šæ ·æ€§ï¼›åˆ†æ"çŒœä½ å–œæ¬¢"çš„é›†ä¸­æ›å…‰é—®é¢˜ï¼Œæ–°çš„åŒå¡”å¬å›è™½è½¬åŒ–ç‡å¾®é™ï¼ˆ-2%ï¼‰ï¼Œä½†å¤´éƒ¨æ•ˆæœæœ‰éå¸¸æ˜æ˜¾çš„æ”¹å–„
â—‹è´¨é‡ç­›é€‰ï¼š æ–°å¢å¹³å‡å¯¹å±€æ—¶é•¿çš„å‡†å…¥ç­›é€‰æ¡ä»¶ï¼Œæé«˜ä½œå“å¢é•¿é€Ÿåº¦çš„æƒé‡ï¼Œç›¸å¯¹æ›´ä¼˜å…ˆæ¨èå¿«é€Ÿå´›èµ·çš„æ–°å†…å®¹

ã€æ ‡ç­¾ä¸å®¡æ ¸æµç¨‹ä¼˜åŒ–ã€‘
â—‹é˜ˆå€¼è°ƒæ•´ï¼š æé«˜äººå®¡ä¸¾æŠ¥é˜ˆå€¼ï¼ˆ1â†’5ï¼‰ï¼Œå‡å°‘è¯¯æŠ¥å¹²æ‰°
â—‹æµç¨‹ä¼˜åŒ–ï¼š ä¿®å¤ä½œå“æ›´æ–°åï¼Œæ²¡æœ‰é‡æ–°è¿›å…¥å®¡æ ¸çš„é—®é¢˜ï¼›å‘ç°éƒ¨åˆ†ä½œè€…åˆ©ç”¨é«˜é¢‘æ›´æ–°ï¼ŒçŸ­æš‚ç»•è¿‡æ ‡ç­¾æµç¨‹ï¼Œå·²æŠ¥å¤‡11æœˆ26æ—¥Patchä¿®å¤è¯¥æ¼æ´
"""

# ç™½çš®ä¹¦åŠ©æ‰‹çš„System Prompt
WHITEPAPER_ASSISTANT_SYSTEM_PROMPT = """# Role: PUBGM WoWæ¨¡å¼ ç‰ˆæœ¬æ–‡æ¡£æ’°å†™åŠ©ç†

# å›å¤è¯­è¨€:
è¯·å§‹ç»ˆä½¿ç”¨ä¸­æ–‡è¿›è¡Œå›ç­”å’Œè¾“å‡º

# Context:
ä½ æ­£åœ¨ååŠ©æ•´ç†PUBGM WoWæ¨¡å¼ï¼ˆUGCç©æ³•ï¼‰çš„ç‰ˆæœ¬ç™½çš®ä¹¦åŠŸèƒ½åˆ—è¡¨ã€‚ç”¨æˆ·ä¼šè¾“å…¥ç®€å•çš„åŠŸèƒ½å…³é”®è¯æˆ–çŸ­è¯­ï¼Œä½ éœ€è¦å°†å…¶æ‰©å†™æˆä¸€å¥æ ‡å‡†ã€ä¸“ä¸šä¸”ä¿¡æ¯é‡å®Œæ•´çš„ç‰ˆæœ¬åŠŸèƒ½é™ˆè¿°ã€‚

# Goal:
å°†ç®€çŸ­çš„å…³é”®è¯æ‰©å†™ä¸ºæ ‡å‡†çš„"åŠŸèƒ½ç‚¹é™ˆè¿°å¥"ã€‚

# Output Rules (Strict):
1.  **å¥å¼ç»“æ„**ï¼šè¯·ä¸¥æ ¼å¥—ç”¨ä»¥ä¸‹å¥å¼è¿›è¡Œæ‰©å†™ï¼š
    `[åºå·]. æ–°å¢[åŠŸèƒ½åç§°]åŠŸèƒ½ï¼Œæ”¯æŒ[å…·ä½“æœºåˆ¶/æ“ä½œæ–¹å¼]ï¼Œç”¨äº[åº”ç”¨åœºæ™¯/å…³è”çš„è®¾å¤‡æˆ–ç³»ç»Ÿ]ã€‚`
2.  **ä¸“ä¸šæ€§**ï¼šä½¿ç”¨PUBGM WoWæ¨¡å¼çš„å¸¸ç”¨æœ¯è¯­ï¼ˆå¦‚ï¼šå¯è§†åŒ–ç¼–ç¨‹ã€è‡ªå®šä¹‰UIã€å…¨å±€å˜é‡ã€äº’åŠ¨ç‰©ä½“ã€æ­¦è£…AIç­‰ï¼‰ã€‚
3.  **ç®€æ´æ€§**ï¼šä¸è¦ä½¿ç”¨æ„Ÿå¹å·ï¼Œä¸è¦å‘è¡¨è¯„è®ºï¼Œä¸è¦ä½¿ç”¨"å¿«æ¥è¯•è¯•"ç­‰è¥é”€è¯æ±‡ã€‚åªé™ˆè¿°äº‹å®ã€‚
4.  **æ•°å­¦å…¬å¼**ï¼šå¦‚æœæ¶‰åŠæ•°å€¼é€»è¾‘ï¼Œè¯·ä½¿ç”¨ LaTeX æ ¼å¼ï¼Œä¾‹å¦‚ $y=x+1$ã€‚

# Input Example:
ç”¨æˆ·è¾“å…¥ï¼šåŠ¨ç”»ç”Ÿæˆ
è¾“å‡ºï¼š1. æ–°å¢åŠ¨ç”»ç”ŸæˆåŠŸèƒ½ï¼Œæ”¯æŒä½œè€…ä¸Šä¼ è§†é¢‘åç”Ÿæˆå¯¹åº”éª¨éª¼åŠ¨ç”»ï¼Œç”¨äºå¯è§†åŒ–ç¼–ç¨‹æ§åˆ¶æ­¦è£…AIå’Œè™šæ‹ŸæŠ•å½±è£…ç½®ã€‚

ç”¨æˆ·è¾“å…¥ï¼šè‡ªå®šä¹‰UI
è¾“å‡ºï¼š1. æ–°å¢è‡ªå®šä¹‰UIç¼–è¾‘å™¨ï¼Œæ”¯æŒåˆ›ä½œè€…è‡ªç”±æ‹–æ‹½æŒ‰é’®ä¸å›¾ç‰‡å¸ƒå±€ï¼Œç”¨äºåˆ¶ä½œä¸ªæ€§åŒ–çš„æ¸¸æˆç•Œé¢ä¸äº¤äº’èœå•ã€‚

# Workflow:
1.  åˆ†æç”¨æˆ·è¾“å…¥çš„å…³é”®è¯ã€‚
2.  è”æƒ³è¯¥åŠŸèƒ½åœ¨PUBGM WoWä¸­çš„å®é™…è¿ä½œé€»è¾‘ï¼ˆæœºåˆ¶ï¼‰å’Œç”¨é€”ï¼ˆåœºæ™¯ï¼‰ã€‚
3.  æŒ‰ç…§è§„å®šå¥å¼è¾“å‡ºã€‚
"""

# ============================================
# ä¼šè¯å†å²ç®¡ç†
# ============================================

# å†å²è®°å½•å­˜å‚¨ç›®å½•
HISTORY_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "user_histories")

def get_user_id() -> str:
    """
    è·å–æˆ–ç”Ÿæˆç”¨æˆ·å”¯ä¸€ID
    
    Returns:
        ç”¨æˆ·å”¯ä¸€IDå­—ç¬¦ä¸²
    """
    import uuid
    if "user_id" not in st.session_state:
        # ç”Ÿæˆä¸€ä¸ªæ–°çš„ç”¨æˆ·ID
        st.session_state.user_id = str(uuid.uuid4())[:8]
    return st.session_state.user_id

def get_user_history_path() -> str:
    """
    è·å–å½“å‰ç”¨æˆ·çš„å†å²è®°å½•æ–‡ä»¶è·¯å¾„
    
    Returns:
        ç”¨æˆ·å†å²è®°å½•æ–‡ä»¶çš„å®Œæ•´è·¯å¾„
    """
    user_id = get_user_id()
    # ç¡®ä¿ç›®å½•å­˜åœ¨
    if not os.path.exists(HISTORY_DIR):
        os.makedirs(HISTORY_DIR)
    return os.path.join(HISTORY_DIR, f"history_{user_id}.json")

def load_history_from_file() -> list:
    """
    ä»æœ¬åœ°æ–‡ä»¶åŠ è½½ä¼šè¯å†å²
    
    Returns:
        å†å²è®°å½•åˆ—è¡¨
    """
    try:
        history_path = get_user_history_path()
        if os.path.exists(history_path):
            with open(history_path, 'r', encoding='utf-8') as f:
                return json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        print(f"åŠ è½½å†å²è®°å½•å¤±è´¥: {e}")
    return []

def save_history_to_file(history: list):
    """
    ä¿å­˜ä¼šè¯å†å²åˆ°æœ¬åœ°æ–‡ä»¶
    
    Args:
        history: å†å²è®°å½•åˆ—è¡¨
    """
    try:
        history_path = get_user_history_path()
        with open(history_path, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    except IOError as e:
        print(f"ä¿å­˜å†å²è®°å½•å¤±è´¥: {e}")

def get_download_data(item: dict) -> bytes:
    """
    è·å–å†å²è®°å½•ä¸­çš„ä¸‹è½½æ•°æ®ï¼Œå¤„ç†base64è§£ç 
    
    Args:
        item: å†å²è®°å½•é¡¹
    
    Returns:
        è§£ç åçš„äºŒè¿›åˆ¶æ•°æ®ï¼Œå¦‚æœæ²¡æœ‰åˆ™è¿”å›None
    """
    download_data = item.get("download_data")
    if download_data:
        # å¦‚æœæ˜¯å­—ç¬¦ä¸²ï¼ˆbase64ç¼–ç ï¼‰ï¼Œåˆ™è§£ç 
        if isinstance(download_data, str):
            try:
                return base64.b64decode(download_data)
            except Exception:
                return download_data.encode('utf-8')
        # å¦‚æœå·²ç»æ˜¯bytesï¼Œç›´æ¥è¿”å›
        return download_data
    return None

def init_session_history():
    """åˆå§‹åŒ–ä¼šè¯å†å²å­˜å‚¨ï¼Œä»æœ¬åœ°æ–‡ä»¶åŠ è½½"""
    if "session_history" not in st.session_state:
        # ä»æœ¬åœ°æ–‡ä»¶åŠ è½½å†å²è®°å½•
        st.session_state.session_history = load_history_from_file()


# ============================================
# å¤šè½®å¯¹è¯ç®¡ç†
# ============================================

def init_chat_history(chat_key: str):
    """
    åˆå§‹åŒ–æŒ‡å®šåŠŸèƒ½çš„å¯¹è¯å†å²
    
    Args:
        chat_key: å¯¹è¯å†å²çš„é”®åï¼ˆå¦‚ 'generate_chat', 'report_chat' ç­‰ï¼‰
    """
    if chat_key not in st.session_state:
        st.session_state[chat_key] = []

def add_chat_message(chat_key: str, role: str, content: str):
    """
    æ·»åŠ æ¶ˆæ¯åˆ°å¯¹è¯å†å²
    
    Args:
        chat_key: å¯¹è¯å†å²çš„é”®å
        role: è§’è‰²ï¼ˆ'user' æˆ– 'assistant'ï¼‰
        content: æ¶ˆæ¯å†…å®¹
    """
    init_chat_history(chat_key)
    st.session_state[chat_key].append({
        "role": role,
        "content": content,
        "timestamp": datetime.now().strftime("%H:%M:%S")
    })

def get_chat_history(chat_key: str) -> list:
    """
    è·å–å¯¹è¯å†å²
    
    Args:
        chat_key: å¯¹è¯å†å²çš„é”®å
    
    Returns:
        å¯¹è¯å†å²åˆ—è¡¨
    """
    init_chat_history(chat_key)
    return st.session_state[chat_key]

def clear_chat_history(chat_key: str):
    """
    æ¸…ç©ºå¯¹è¯å†å²
    
    Args:
        chat_key: å¯¹è¯å†å²çš„é”®å
    """
    st.session_state[chat_key] = []

def clear_module_session(module_name: str):
    """
    æ¸…ç©ºæŒ‡å®šæ¨¡å—çš„æ‰€æœ‰ä¼šè¯æ•°æ®
    
    Args:
        module_name: æ¨¡å—åç§°
    """
    if module_name == "ç”Ÿæˆç­–åˆ’æ¡ˆ":
        st.session_state.generated_prd = ""
        st.session_state.uploaded_file_content = ""
        st.session_state.uploaded_file_name = ""
        st.session_state.show_preview_gen = False
        st.session_state.generated_check_result = ""
        st.session_state.current_stage = "idle"
        st.session_state.generate_saved_to_history = False
        clear_chat_history("generate_prd_chat")
    elif module_name == "è„‘å›¾ç”Ÿæˆç­–åˆ’æ¡ˆ":
        st.session_state.mindmap_parsed_structure = None
        st.session_state.mindmap_generated_prd = None
        st.session_state.mindmap_image_data = None
        st.session_state.mindmap_saved = False
        st.session_state.mindmap_mermaid_code = ""
        st.session_state.mindmap_input_mode = "å›¾ç‰‡ä¸Šä¼ "
        clear_chat_history("mindmap_prd_chat")
    elif module_name == "ä¼˜åŒ–ç­–åˆ’æ¡ˆ":
        st.session_state.optimized_prd = ""
        st.session_state.optimize_saved_to_history = False
        clear_chat_history("optimize_prd_chat")
    elif module_name == "æ±‡æŠ¥åŠ©æ‰‹":
        if "generated_report" in st.session_state:
            st.session_state.generated_report = ""
        if "report_saved_to_history" in st.session_state:
            st.session_state.report_saved_to_history = False
        clear_chat_history("report_chat")
    elif module_name == "å‘¨æŠ¥åŠ©æ‰‹":
        if "generated_weekly_report" in st.session_state:
            st.session_state.generated_weekly_report = ""
        if "weekly_saved_to_history" in st.session_state:
            st.session_state.weekly_saved_to_history = False
        clear_chat_history("weekly_chat")
    elif module_name == "ç™½çš®ä¹¦åŠ©æ‰‹":
        if "generated_feature_desc" in st.session_state:
            st.session_state.generated_feature_desc = ""
        if "whitepaper_saved_to_history" in st.session_state:
            st.session_state.whitepaper_saved_to_history = False
        clear_chat_history("whitepaper_chat")
    elif module_name == "æ¸¸æˆç­–åˆ’(lina)":
        st.session_state.lina_chat_history = []
        st.session_state.lina_is_processing = False
    elif module_name == "è¡¨æ ¼å¤„ç†åŠ©æ‰‹":
        st.session_state.table_dataframes = {}
        st.session_state.table_selected_dfs = []
        st.session_state.table_result_df = None
        st.session_state.table_is_processing = False
        st.session_state.table_uploaded_files_info = {}
    elif module_name == "æ€è·¯å¼•å¯¼åŠ©æ‰‹ (linmo)":
        st.session_state.linmo_chat_history = []
        st.session_state.linmo_is_processing = False
        st.session_state.linmo_input_key_counter = st.session_state.get("linmo_input_key_counter", 0) + 1
    elif module_name == "PUBGM WoW ç©æ³•è¯„å®¡":
        st.session_state.wow_review_result = ""
        st.session_state.wow_is_processing = False
        st.session_state.wow_uploaded_video = None

def build_chat_context(chat_key: str, system_prompt: str, max_history: int = 10) -> str:
    """
    æ„å»ºåŒ…å«å¯¹è¯å†å²çš„ä¸Šä¸‹æ–‡Prompt
    
    Args:
        chat_key: å¯¹è¯å†å²çš„é”®å
        system_prompt: ç³»ç»Ÿæç¤ºè¯
        max_history: æœ€å¤§å†å²æ¶ˆæ¯æ•°é‡
    
    Returns:
        åŒ…å«å†å²ä¸Šä¸‹æ–‡çš„å®Œæ•´Prompt
    """
    history = get_chat_history(chat_key)
    
    if not history:
        return ""
    
    # åªå–æœ€è¿‘çš„Næ¡å†å²
    recent_history = history[-max_history:] if len(history) > max_history else history
    
    # æ„å»ºå¯¹è¯å†å²æ–‡æœ¬
    history_text = "\n\nã€å¯¹è¯å†å²ã€‘\n"
    for msg in recent_history:
        role_label = "ç”¨æˆ·" if msg["role"] == "user" else "åŠ©æ‰‹"
        history_text += f"{role_label}: {msg['content']}\n\n"
    
    return history_text

def render_chat_interface(chat_key: str, system_prompt: str, container, 
                          placeholder: str = "è¯·è¾“å…¥æ‚¨çš„é—®é¢˜æˆ–ä¿®æ”¹è¦æ±‚...",
                          function_context: str = ""):
    """
    æ¸²æŸ“å¤šè½®å¯¹è¯ç•Œé¢
    
    Args:
        chat_key: å¯¹è¯å†å²çš„é”®å
        system_prompt: ç³»ç»Ÿæç¤ºè¯
        container: Streamlitå®¹å™¨
        placeholder: è¾“å…¥æ¡†å ä½æ–‡æœ¬
        function_context: å½“å‰åŠŸèƒ½çš„ä¸Šä¸‹æ–‡ï¼ˆå¦‚å·²ç”Ÿæˆçš„å†…å®¹ï¼‰
    
    Returns:
        æ˜¯å¦æœ‰æ–°çš„å¯¹è¯äº§ç”Ÿ
    """
    init_chat_history(chat_key)
    history = get_chat_history(chat_key)
    
    # æ˜¾ç¤ºå¯¹è¯å†å² - ä½¿ç”¨ ChatGPT é£æ ¼çš„å¯¹è¯æ°”æ³¡
    if history:
        with container:
            st.markdown("#### ğŸ’¬ å¯¹è¯å†å²")
            for i, msg in enumerate(history):
                if msg["role"] == "user":
                    with st.chat_message("user"):
                        st.markdown(msg["content"])
                else:
                    with st.chat_message("assistant", avatar="ğŸ¤–"):
                        st.markdown(msg["content"])
    
    # ç”¨äºæ§åˆ¶å¯¹è¯è¾“å…¥çš„çŠ¶æ€
    chat_input_key = f"{chat_key}_input"
    chat_processing_key = f"{chat_key}_processing"
    
    if chat_processing_key not in st.session_state:
        st.session_state[chat_processing_key] = False
    
    # å¯¹è¯è¾“å…¥åŒºåŸŸ - ä½¿ç”¨ chat_input
    chat_input_value = container.chat_input(
        placeholder=placeholder,
        key=chat_input_key
    )
    
    # æ¸…ç©ºæŒ‰é’®
    if container.button("ğŸ—‘ï¸ æ¸…ç©ºå¯¹è¯å†å²", key=f"{chat_key}_clear", use_container_width=False):
        clear_chat_history(chat_key)
        st.rerun()
    
    # å¤„ç†ç”¨æˆ·è¾“å…¥
    if chat_input_value and chat_input_value.strip() and not st.session_state[chat_processing_key]:
        pass  # å®é™…å¤„ç†é€»è¾‘åœ¨å„æ¨¡å—ä¸­å®ç°

def process_chat_message(chat_key: str, user_message: str, system_prompt: str,
                         function_context: str, output_container):
    """
    å¤„ç†ç”¨æˆ·çš„å¯¹è¯æ¶ˆæ¯å¹¶ç”Ÿæˆå›å¤
    
    Args:
        chat_key: å¯¹è¯å†å²çš„é”®å
        user_message: ç”¨æˆ·æ¶ˆæ¯
        system_prompt: ç³»ç»Ÿæç¤ºè¯
        function_context: å½“å‰åŠŸèƒ½çš„ä¸Šä¸‹æ–‡
        output_container: è¾“å‡ºå®¹å™¨
    
    Returns:
        ç”Ÿæˆçš„å›å¤å†…å®¹
    """
    # æ·»åŠ ç”¨æˆ·æ¶ˆæ¯åˆ°å†å²
    add_chat_message(chat_key, "user", user_message)
    
    # æ„å»ºå®Œæ•´çš„Prompt
    history_context = build_chat_context(chat_key, system_prompt)
    
    full_prompt = f"""{function_context}

{history_context}

ã€å½“å‰ç”¨æˆ·è¾“å…¥ã€‘
{user_message}

è¯·åŸºäºä»¥ä¸Šä¸Šä¸‹æ–‡å’Œå¯¹è¯å†å²ï¼Œå›ç­”ç”¨æˆ·çš„é—®é¢˜æˆ–æŒ‰è¦æ±‚è¿›è¡Œä¿®æ”¹ã€‚"""
    
    # è°ƒç”¨APIç”Ÿæˆå›å¤
    full_response = ""
    was_stopped = False
    has_error = False
    error_message = ""
    
    for chunk in call_gemini_stream(full_prompt, system_prompt):
        if st.session_state.should_stop:
            was_stopped = True
            break
        
        if chunk["type"] == "text":
            full_response += chunk["content"]
            output_container.markdown(full_response + "â–Œ")
        elif chunk["type"] == "error":
            has_error = True
            error_message = chunk["content"]
            break
        elif chunk["type"] == "retry":
            st.info(chunk["content"])
    
    # ç§»é™¤å…‰æ ‡
    if full_response:
        output_container.markdown(full_response)
    
    # å¤„ç†ç»“æœ
    if has_error:
        return None, error_message
    elif was_stopped:
        if full_response:
            add_chat_message(chat_key, "assistant", full_response)
        return full_response, "å·²ä¸­æ­¢"
    else:
        add_chat_message(chat_key, "assistant", full_response)
        return full_response, None

def add_to_history(function_type: str, input_data: dict, output_data: str, 
                   download_data: bytes = None, download_filename: str = None,
                   download_mime: str = None):
    """
    æ·»åŠ è®°å½•åˆ°ä¼šè¯å†å²
    
    Args:
        function_type: åŠŸèƒ½ç±»å‹ï¼ˆç”Ÿæˆç­–åˆ’æ¡ˆ/ä¼˜åŒ–ç­–åˆ’æ¡ˆ/æ±‡æŠ¥åŠ©æ‰‹/å‘¨æŠ¥åŠ©æ‰‹/ç™½çš®ä¹¦åŠ©æ‰‹ï¼‰
        input_data: è¾“å…¥æ•°æ®å­—å…¸
        output_data: è¾“å‡ºå†…å®¹
        download_data: å¯ä¸‹è½½çš„æ–‡ä»¶æ•°æ®ï¼ˆå¯é€‰ï¼‰
        download_filename: ä¸‹è½½æ–‡ä»¶åï¼ˆå¯é€‰ï¼‰
        download_mime: æ–‡ä»¶MIMEç±»å‹ï¼ˆå¯é€‰ï¼‰
    """
    init_session_history()
    
    history_item = {
        "id": len(st.session_state.session_history) + 1,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "function_type": function_type,
        "input_data": input_data,
        "output_data": output_data,
        # å°†äºŒè¿›åˆ¶æ•°æ®è½¬ä¸ºbase64å­—ç¬¦ä¸²ä»¥ä¾¿å­˜å‚¨åˆ°JSON
        "download_data": base64.b64encode(download_data).decode('utf-8') if download_data else None,
        "download_filename": download_filename,
        "download_mime": download_mime
    }
    
    st.session_state.session_history.append(history_item)
    
    # ä¿å­˜åˆ°æœ¬åœ°æ–‡ä»¶
    save_history_to_file(st.session_state.session_history)

def get_history_summary(item: dict) -> str:
    """
    è·å–å†å²è®°å½•çš„æ‘˜è¦æè¿°
    
    Args:
        item: å†å²è®°å½•é¡¹
    
    Returns:
        æ‘˜è¦å­—ç¬¦ä¸²
    """
    func_type = item.get("function_type", "æœªçŸ¥")
    input_data = item.get("input_data", {})
    
    # æ ¹æ®ä¸åŒåŠŸèƒ½ç±»å‹ç”Ÿæˆä¸åŒçš„æ‘˜è¦
    if func_type == "ç”Ÿæˆç­–åˆ’æ¡ˆ":
        desc = input_data.get("åŠŸèƒ½æè¿°", "")[:30]
        return f"ğŸ“ {desc}..." if len(input_data.get("åŠŸèƒ½æè¿°", "")) > 30 else f"ğŸ“ {desc}"
    elif func_type == "ä¼˜åŒ–ç­–åˆ’æ¡ˆ":
        return f"ğŸ”„ ç­–åˆ’æ¡ˆä¼˜åŒ–"
    elif func_type == "æ±‡æŠ¥åŠ©æ‰‹":
        problem = input_data.get("å½“å‰é—®é¢˜", "")[:20]
        return f"ğŸ“Š {problem}..." if len(input_data.get("å½“å‰é—®é¢˜", "")) > 20 else f"ğŸ“Š {problem}"
    elif func_type == "å‘¨æŠ¥åŠ©æ‰‹":
        return f"ğŸ“… å‘¨æŠ¥ç”Ÿæˆ"
    elif func_type == "ç™½çš®ä¹¦åŠ©æ‰‹":
        keyword = input_data.get("åŠŸèƒ½å…³é”®è¯", "")
        return f"ğŸ“– {keyword}"
    else:
        return f"ğŸ“„ {func_type}"

def clear_session_history():
    """æ¸…ç©ºä¼šè¯å†å²"""
    st.session_state.session_history = []
    # åŒæ—¶æ¸…ç©ºæœ¬åœ°æ–‡ä»¶
    save_history_to_file([])

def render_history_sidebar():
    """
    åœ¨ä¾§è¾¹æ æ¸²æŸ“ä¼šè¯å†å²é¢æ¿
    """
    init_session_history()
    
    st.sidebar.markdown("---")
    st.sidebar.subheader("ğŸ“œ ä¼šè¯å†å²")
    
    # æ˜¾ç¤ºç”¨æˆ·IDå’Œå†å²æ–‡ä»¶ä¿¡æ¯
    user_id = get_user_id()
    history_path = get_user_history_path()
    
    # ç”¨æˆ·ä¿¡æ¯æ˜¾ç¤ºåŒº
    st.sidebar.caption(f"ğŸ†” æ‚¨çš„ç”¨æˆ·ID: `{user_id}`")
    
    # ä¸‹è½½æŒ‰é’®æ”¾åœ¨æœ€æ˜¾çœ¼ä½ç½®
    if os.path.exists(history_path):
        try:
            with open(history_path, 'r', encoding='utf-8') as f:
                history_content = f.read()
            st.sidebar.download_button(
                label="ğŸ’¾ ä¸‹è½½æˆ‘çš„å†å²è®°å½•",
                data=history_content,
                file_name=f"history_{user_id}.json",
                mime="application/json",
                key="download_history_file",
                use_container_width=True
            )
        except Exception as e:
            st.sidebar.error(f"è¯»å–æ–‡ä»¶å¤±è´¥: {e}")
    else:
        st.sidebar.caption("ğŸ“ æš‚æ— å†å²è®°å½•å¯ä¸‹è½½")
    
    # å­˜å‚¨ä¿¡æ¯æŠ˜å é¢æ¿
    with st.sidebar.expander("ğŸ“ å­˜å‚¨ä¿¡æ¯è¯¦æƒ…", expanded=False):
        st.caption(f"ğŸ“‚ **å­˜å‚¨æ–‡ä»¶**: `history_{user_id}.json`")
        st.caption(f"ğŸ“ **å­˜å‚¨ç›®å½•**: `{HISTORY_DIR}`")
        st.info("ğŸ’¡ åˆ·æ–°é¡µé¢ä¼šç”Ÿæˆæ–°çš„ç”¨æˆ·IDï¼Œå»ºè®®åŠæ—¶ä¸‹è½½å¤‡ä»½å†å²è®°å½•")
    
    history = st.session_state.session_history
    
    if not history:
        st.sidebar.caption("æš‚æ— å†å²è®°å½•")
        return
    
    # æ˜¾ç¤ºå†å²è®°å½•æ•°é‡å’Œæ¸…ç©ºæŒ‰é’®
    col1, col2 = st.sidebar.columns([2, 1])
    with col1:
        st.caption(f"å…± {len(history)} æ¡è®°å½•")
    with col2:
        if st.button("ğŸ—‘ï¸ æ¸…ç©º", key="clear_history", use_container_width=True):
            clear_session_history()
            st.rerun()
    
    # å€’åºæ˜¾ç¤ºå†å²è®°å½•ï¼ˆæœ€æ–°çš„åœ¨å‰ï¼‰
    for item in reversed(history):
        item_id = item.get("id", 0)
        timestamp = item.get("timestamp", "")
        func_type = item.get("function_type", "")
        summary = get_history_summary(item)
        
        # ä½¿ç”¨expanderæ˜¾ç¤ºæ¯æ¡è®°å½•
        with st.sidebar.expander(f"#{item_id} {summary}", expanded=False):
            st.caption(f"ğŸ• {timestamp}")
            st.caption(f"ğŸ“Œ {func_type}")
            
            # æŸ¥çœ‹è¯¦æƒ…æŒ‰é’®
            if st.button("ğŸ“„ æŸ¥çœ‹è¯¦æƒ…", key=f"view_{item_id}", use_container_width=True):
                st.session_state.viewing_history_id = item_id
                st.session_state.show_history_detail = True
                st.rerun()
            
            # å¦‚æœæœ‰ä¸‹è½½æ•°æ®ï¼Œæ˜¾ç¤ºä¸‹è½½æŒ‰é’®
            if item.get("download_data"):
                st.download_button(
                    label="ğŸ“¥ ä¸‹è½½",
                    data=get_download_data(item),
                    file_name=item.get("download_filename", "download.txt"),
                    mime=item.get("download_mime", "text/plain"),
                    key=f"download_{item_id}",
                    use_container_width=True
                )


# AIè‡ªæ£€çš„System Prompt
SELF_CHECK_SYSTEM_PROMPT = """ä½ æ˜¯èµ„æ·±æ¸¸æˆç­–åˆ’"é…¸å¥¶"ï¼Œæ­£åœ¨å¯¹ç­–åˆ’æ¡ˆè¿›è¡Œå¤æ£€æ¸…å•æ£€æŸ¥ã€‚

ã€å›å¤è¯­è¨€ã€‘
- è¯·å§‹ç»ˆä½¿ç”¨ä¸­æ–‡è¿›è¡Œå›ç­”å’Œè¾“å‡º

è¯·æ ¹æ®ä»¥ä¸‹10é¡¹å¤æ£€æ¸…å•ï¼Œé€ä¸€æ£€æŸ¥ç­–åˆ’æ¡ˆçš„å®Œæ•´æ€§å’Œè§„èŒƒæ€§ï¼š

ã€å¤æ£€æ¸…å•ã€‘
1. æ˜¯å¦ç”¨ä¸€å¥è¯è¯´æ¸…åŠŸèƒ½æ ¸å¿ƒï¼Ÿ
2. æ˜¯å¦æ˜ç¡®å®šä¹‰ç›®æ ‡ç”¨æˆ·å’Œä½¿ç”¨åœºæ™¯ï¼Ÿ
3. æ˜¯å¦æè¿°æ¸…æ¥šç”¨æˆ·è§¦å‘è·¯å¾„ï¼Ÿ
4. æ˜¯å¦å®šä¹‰è¾“å…¥è¦æ±‚ï¼ˆæ ¼å¼ã€é™åˆ¶ï¼‰ï¼Ÿ
5. æ˜¯å¦è¯´æ˜AIå¤„ç†é€»è¾‘ï¼ˆæ¨¡å‹ã€æµç¨‹ï¼‰ï¼Ÿ
6. æ˜¯å¦å®šä¹‰è¾“å‡ºæ ¼å¼ï¼ˆæ˜¯å¦å¯ç¼–è¾‘ï¼‰ï¼Ÿ
7. æ˜¯å¦è®¾è®¡ç”¨æˆ·ä½“éªŒæµè½¬ï¼ˆä¿®æ”¹ã€é‡è¯•ï¼‰ï¼Ÿ
8. æ˜¯å¦è®¾å®šé‡åŒ–éªŒæ”¶æ ‡å‡†ï¼Ÿ
9. æ˜¯å¦å£°æ˜èƒ½åŠ›è¾¹ç•Œï¼Ÿ
10. æ˜¯å¦åˆ—å‡ºæŠ€æœ¯ä¾èµ–ï¼Ÿ

ã€è¾“å‡ºè¦æ±‚ã€‘
è¯·æŒ‰ä»¥ä¸‹æ ¼å¼è¾“å‡ºæ£€æŸ¥ç»“æœï¼š
- å¯¹æ¯ä¸€é¡¹ï¼Œå…ˆæ ‡æ˜æ£€æŸ¥é¡¹ç¼–å·å’Œåç§°
- ç»™å‡ºåˆ¤æ–­ï¼šâœ… é€šè¿‡ / âš ï¸ éƒ¨åˆ†æ»¡è¶³ / âŒ ç¼ºå¤±
- å¦‚æœæ˜¯"éƒ¨åˆ†æ»¡è¶³"æˆ–"ç¼ºå¤±"ï¼Œè¯·è¯´æ˜å…·ä½“ç¼ºå°‘ä»€ä¹ˆå†…å®¹æˆ–æ”¹è¿›å»ºè®®
- æœ€åç»™å‡ºæ€»ä½“è¯„ä»·å’Œä¼˜å…ˆæ”¹è¿›å»ºè®®

è¯·ç”¨ä¸­æ–‡è¾“å‡ºï¼Œæ ¼å¼æ¸…æ™°æ˜“è¯»ã€‚"""


def parse_prd_to_excel_data(prd_content: str) -> list:
    """
    è§£æç­–åˆ’æ¡ˆæ–‡æœ¬ï¼Œè½¬æ¢ä¸ºExcelæ•°æ®æ ¼å¼
    æŒ‰æ ‡é¢˜å±‚çº§åˆ†é…åˆ°ä¸åŒåˆ—ï¼š
    - ä¸€çº§æ ‡é¢˜ï¼ˆå¦‚ 1ã€xxxï¼‰åœ¨ç¬¬1åˆ—
    - äºŒçº§æ ‡é¢˜ï¼ˆå¦‚ 1.1ã€xxxï¼‰åœ¨ç¬¬2åˆ—
    - ä¸‰çº§æ ‡é¢˜ï¼ˆå¦‚ 1.1.1ã€xxxï¼‰åœ¨ç¬¬3åˆ—
    - æ™®é€šå†…å®¹åœ¨æœ€è¿‘æ ‡é¢˜çš„ä¸‹ä¸€åˆ—
    
    Returns:
        list: [(row_data, level), ...] æ¯è¡Œæ•°æ®å’Œå…¶å±‚çº§
    """
    lines = prd_content.strip().split('\n')
    excel_data = []
    current_level = 0
    
    # åŒ¹é…å„çº§æ ‡é¢˜çš„æ­£åˆ™è¡¨è¾¾å¼
    # ä¸€çº§æ ‡é¢˜: 1ã€ æˆ– 1. æˆ– 1  å¼€å¤´ï¼ˆçº¯æ•°å­—ï¼‰
    level1_pattern = re.compile(r'^(\d+)[ã€\.ï¼]\s*(.+)$')
    # äºŒçº§æ ‡é¢˜: 1.1ã€ æˆ– 1.1. æˆ– 1.1 å¼€å¤´
    level2_pattern = re.compile(r'^(\d+\.\d+)[ã€\.ï¼]?\s*(.+)$')
    # ä¸‰çº§æ ‡é¢˜: 1.1.1ã€ æˆ– 1.1.1. æˆ– 1.1.1 å¼€å¤´
    level3_pattern = re.compile(r'^(\d+\.\d+\.\d+)[ã€\.ï¼]?\s*(.+)$')
    # å››çº§æ ‡é¢˜: 1.1.1.1 å¼€å¤´
    level4_pattern = re.compile(r'^(\d+\.\d+\.\d+\.\d+)[ã€\.ï¼]?\s*(.+)$')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # æ£€æŸ¥æ˜¯å¦æ˜¯æ ‡é¢˜è¡Œï¼Œä»é«˜çº§åˆ«å¾€ä½çº§åˆ«æ£€æŸ¥
        level4_match = level4_pattern.match(line)
        level3_match = level3_pattern.match(line)
        level2_match = level2_pattern.match(line)
        level1_match = level1_pattern.match(line)
        
        if level4_match:
            # å››çº§æ ‡é¢˜ -> ç¬¬4åˆ—
            current_level = 4
            excel_data.append((line, 4))
        elif level3_match:
            # ä¸‰çº§æ ‡é¢˜ -> ç¬¬3åˆ—
            current_level = 3
            excel_data.append((line, 3))
        elif level2_match:
            # äºŒçº§æ ‡é¢˜ -> ç¬¬2åˆ—
            current_level = 2
            excel_data.append((line, 2))
        elif level1_match:
            # ä¸€çº§æ ‡é¢˜ -> ç¬¬1åˆ—
            current_level = 1
            excel_data.append((line, 1))
        else:
            # æ™®é€šå†…å®¹ -> å½“å‰æ ‡é¢˜çš„ä¸‹ä¸€åˆ—ï¼Œè‡³å°‘åœ¨ç¬¬2åˆ—
            content_level = max(current_level + 1, 2) if current_level > 0 else 1
            excel_data.append((line, content_level))
    
    return excel_data


def create_excel_file(prd_content: str, check_result: str = "") -> bytes:
    """
    åˆ›å»ºExcelæ–‡ä»¶
    
    Args:
        prd_content: ç­–åˆ’æ¡ˆå†…å®¹
        check_result: AIå¤æ£€ç»“æœï¼ˆå¯é€‰ï¼‰
    
    Returns:
        bytes: Excelæ–‡ä»¶çš„äºŒè¿›åˆ¶æ•°æ®
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "ç­–åˆ’æ¡ˆ"
    
    # å®šä¹‰æ ·å¼
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    level1_font = Font(bold=True, size=12, color="1F4E79")
    level2_font = Font(bold=True, size=11, color="2E75B6")
    level3_font = Font(bold=False, size=10, color="5B9BD5")
    normal_font = Font(size=10)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # è®¾ç½®åˆ—å®½
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50
    
    # æ·»åŠ è¡¨å¤´
    headers = ["ä¸€çº§æ ‡é¢˜", "äºŒçº§æ ‡é¢˜/å†…å®¹", "ä¸‰çº§æ ‡é¢˜/è¯¦æƒ…", "å››çº§æ ‡é¢˜/è¯´æ˜", "è¯¦ç»†å†…å®¹"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # è§£æå¹¶å¡«å……ç­–åˆ’æ¡ˆå†…å®¹
    excel_data = parse_prd_to_excel_data(prd_content)
    
    row_num = 2
    for content, level in excel_data:
        # å°†å†…å®¹æ”¾åˆ°å¯¹åº”å±‚çº§çš„åˆ—
        cell = ws.cell(row=row_num, column=level, value=content)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        
        # æ ¹æ®å±‚çº§è®¾ç½®å­—ä½“æ ·å¼
        if level == 1:
            cell.font = level1_font
        elif level == 2:
            cell.font = level2_font
        elif level == 3:
            cell.font = level3_font
        else:
            cell.font = normal_font
        
        # ä¸ºè¯¥è¡Œçš„æ‰€æœ‰åˆ—æ·»åŠ è¾¹æ¡†
        for col in range(1, 6):
            if col != level:
                empty_cell = ws.cell(row=row_num, column=col, value="")
                empty_cell.border = thin_border
        
        row_num += 1
    
    # å¦‚æœæœ‰å¤æ£€ç»“æœï¼Œæ·»åŠ åˆ°æ–°çš„sheet
    if check_result:
        ws_check = wb.create_sheet(title="AIå¤æ£€ç»“æœ")
        ws_check.column_dimensions['A'].width = 100
        
        # æ·»åŠ æ ‡é¢˜
        title_cell = ws_check.cell(row=1, column=1, value="AIå¤æ£€æ¸…å•æ£€æŸ¥ç»“æœ")
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal='center')
        
        # è§£æå¤æ£€ç»“æœ
        check_lines = check_result.strip().split('\n')
        for idx, line in enumerate(check_lines, 2):
            cell = ws_check.cell(row=idx, column=1, value=line)
            cell.alignment = wrap_alignment
            
            # æ ¹æ®å†…å®¹è®¾ç½®æ ·å¼
            if 'âœ…' in line:
                cell.font = Font(color="228B22")  # ç»¿è‰²
            elif 'âš ï¸' in line:
                cell.font = Font(color="FF8C00")  # æ©™è‰²
            elif 'âŒ' in line:
                cell.font = Font(color="DC143C")  # çº¢è‰²
    
    # ä¿å­˜åˆ°å†…å­˜
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def extract_text_from_pdf(file_content: bytes) -> str:
    """ä»PDFæ–‡ä»¶æå–æ–‡æœ¬"""
    try:
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text.strip()
    except Exception as e:
        return f"[PDFè§£æå¤±è´¥: {str(e)}]"


def extract_text_from_docx(file_content: bytes) -> str:
    """ä»Wordæ–‡æ¡£æå–æ–‡æœ¬"""
    try:
        doc = docx.Document(io.BytesIO(file_content))
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text.strip()
    except Exception as e:
        return f"[Wordæ–‡æ¡£è§£æå¤±è´¥: {str(e)}]"


def extract_text_from_file(uploaded_file) -> str:
    """
    ä»ä¸Šä¼ çš„æ–‡ä»¶ä¸­æå–æ–‡æœ¬å†…å®¹
    
    Args:
        uploaded_file: Streamlitä¸Šä¼ çš„æ–‡ä»¶å¯¹è±¡
    
    Returns:
        str: æå–çš„æ–‡æœ¬å†…å®¹
    """
    if uploaded_file is None:
        return ""
    
    file_name = uploaded_file.name.lower()
    file_content = uploaded_file.read()
    
    # é‡ç½®æ–‡ä»¶æŒ‡é’ˆï¼Œä»¥ä¾¿åç»­å¯èƒ½çš„é‡å¤è¯»å–
    uploaded_file.seek(0)
    
    if file_name.endswith('.pdf'):
        return extract_text_from_pdf(file_content)
    elif file_name.endswith('.docx'):
        return extract_text_from_docx(file_content)
    elif file_name.endswith('.txt') or file_name.endswith('.md'):
        # å°è¯•å¤šç§ç¼–ç 
        for encoding in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
            try:
                return file_content.decode(encoding)
            except UnicodeDecodeError:
                continue
        return "[æ–‡æœ¬æ–‡ä»¶è§£ç å¤±è´¥]"
    else:
        return "[ä¸æ”¯æŒçš„æ–‡ä»¶ç±»å‹]"


def format_prd_content(content: str) -> str:
    """
    æ ¼å¼åŒ–ç­–åˆ’æ¡ˆå†…å®¹ï¼Œå¢å¼ºMarkdownæ˜¾ç¤ºæ•ˆæœ
    å°†æ•°å­—æ ‡é¢˜è½¬æ¢ä¸ºæ›´ç¾è§‚çš„æ ¼å¼
    
    Args:
        content: åŸå§‹ç­–åˆ’æ¡ˆå†…å®¹
    
    Returns:
        str: æ ¼å¼åŒ–åçš„Markdownå†…å®¹
    """
    import re
    
    # å¤„ç†å†…å®¹ï¼Œå¢å¼ºæ ¼å¼
    lines = content.split('\n')
    formatted_lines = []
    
    # ç”¨äºåˆ¤æ–­æ˜¯å¦åœ¨åˆ—è¡¨ä¸Šä¸‹æ–‡ä¸­
    in_list_context = False
    
    for i, line in enumerate(lines):
        stripped = line.strip()
        
        # è·³è¿‡ç©ºè¡Œ
        if not stripped:
            formatted_lines.append(line)
            in_list_context = False
            continue
        
        # æ¸…ç†æ ‡é¢˜ä¸­çš„ ** ç¬¦å·
        clean_line = re.sub(r'\*\*', '', stripped)
        
        # åŒ¹é…ä¸‰çº§æ ‡é¢˜ï¼š1.1.1ã€xxx æˆ– 1.1.1 xxxï¼ˆä¼˜å…ˆåŒ¹é…æ›´é•¿çš„æ¨¡å¼ï¼‰
        level3_match = re.match(r'^(\d+\.\d+\.\d+)[ã€\.ï¼]?\s*(.+)$', clean_line)
        # åŒ¹é…äºŒçº§æ ‡é¢˜ï¼š1.1ã€xxx æˆ– 1.1 xxx
        level2_match = re.match(r'^(\d+\.\d+)[ã€\.ï¼]?\s*(.+)$', clean_line)
        # åŒ¹é…ä¸€çº§æ ‡é¢˜ï¼šä»…è¡Œé¦–ä¸ºå•ä¸ªæ•°å­— + é¡¿å·/ç‚¹å· + æ ‡é¢˜æ–‡å­—ï¼ˆä¸å«å†’å·ç»“å°¾ï¼Œé¿å…åŒ¹é…åˆ—è¡¨ï¼‰
        level1_match = re.match(r'^(\d+)[ã€\.ï¼]\s*([^ï¼š:]+)$', clean_line)
        
        # æ£€æŸ¥æ˜¯å¦æ˜¯åˆ—è¡¨é¡¹ï¼ˆåœ¨ç‰¹å®šä¸Šä¸‹æ–‡ä¸­çš„æ•°å­—å¼€å¤´è¡Œï¼‰
        # åˆ—è¡¨é¡¹ç‰¹å¾ï¼šå‰é¢æœ‰ - æˆ– * å¼€å¤´ï¼Œæˆ–è€…åœ¨æµç¨‹/æ­¥éª¤æè¿°ä¸­
        is_list_item = False
        
        # æ£€æŸ¥å‰ä¸€è¡Œæ˜¯å¦æš—ç¤ºè¿™æ˜¯åˆ—è¡¨
        if i > 0:
            prev_line = lines[i-1].strip() if i > 0 else ""
            # å¦‚æœå‰ä¸€è¡Œä»¥å†’å·ç»“å°¾ï¼Œæˆ–åŒ…å«"æµç¨‹"ã€"æ­¥éª¤"ç­‰è¯ï¼Œåç»­çš„æ•°å­—è¡Œå¯èƒ½æ˜¯åˆ—è¡¨
            if prev_line.endswith('ï¼š') or prev_line.endswith(':') or \
               'æµç¨‹' in prev_line or 'æ­¥éª¤' in prev_line or in_list_context:
                # æ£€æŸ¥å½“å‰è¡Œæ˜¯å¦çœ‹èµ·æ¥åƒåˆ—è¡¨é¡¹ï¼ˆè¾ƒé•¿çš„æè¿°æ€§æ–‡å­—ï¼‰
                if level1_match and len(clean_line) > 20:
                    is_list_item = True
                    in_list_context = True
        
        # æ£€æŸ¥æ˜¯å¦æ˜¯ä»¥ - æˆ– * å¼€å¤´çš„åˆ—è¡¨é¡¹
        if stripped.startswith('-') or stripped.startswith('*'):
            # ä¿æŒåŸæ ·ï¼Œåªæ¸…ç†å¤šä½™çš„ **
            formatted_lines.append(re.sub(r'\*\*([^*]+)\*\*', r'**\1**', line))
            in_list_context = True
            continue
        
        if level3_match:
            num, title = level3_match.groups()
            title = title.strip()
            formatted_lines.append(f'\n#### {num} {title}\n')
            in_list_context = False
        elif level2_match:
            num, title = level2_match.groups()
            title = title.strip()
            formatted_lines.append(f'\n### {num} {title}\n')
            in_list_context = False
        elif level1_match and not is_list_item:
            num, title = level1_match.groups()
            title = title.strip()
            # ä¸€çº§æ ‡é¢˜ä½¿ç”¨ç‰¹æ®Šæ ·å¼
            formatted_lines.append(f'\n## {num}ã€{title}\n')
            in_list_context = False
        else:
            # å¯¹äºæ™®é€šè¡Œï¼Œä¿æŒåŸæ ·ä½†æ¸…ç†æ ¼å¼
            # å¤„ç†åˆ—è¡¨é¡¹æ ¼å¼ï¼Œç¡®ä¿ **xxx** æ ¼å¼æ­£ç¡®
            processed_line = line
            # å¦‚æœæ˜¯æ•°å­—å¼€å¤´çš„åˆ—è¡¨é¡¹ï¼Œè½¬æ¢ä¸ºæœ‰åºåˆ—è¡¨æ ¼å¼
            list_item_match = re.match(r'^(\d+)[ã€\.ï¼]\s*(.+)$', clean_line)
            if list_item_match and is_list_item:
                num, text = list_item_match.groups()
                processed_line = f'{num}. {text}'
            formatted_lines.append(processed_line)
    
    return '\n'.join(formatted_lines)


def render_prd_document(content: str, title: str = "ç­–åˆ’æ¡ˆ"):
    """
    ä»¥ç¾è§‚çš„æ–‡æ¡£æ ¼å¼æ¸²æŸ“ç­–åˆ’æ¡ˆå†…å®¹
    
    Args:
        content: ç­–åˆ’æ¡ˆå†…å®¹
        title: æ–‡æ¡£æ ‡é¢˜
    """
    import re
    
    # æ ¼å¼åŒ–å†…å®¹
    formatted_content = format_prd_content(content)
    
    # å°†Markdownè½¬æ¢ä¸ºHTMLä»¥ä¾¿åœ¨è‡ªå®šä¹‰å®¹å™¨ä¸­æ­£ç¡®æ˜¾ç¤º
    # å¤„ç†æ ‡é¢˜
    html_content = formatted_content
    
    # è½¬æ¢ ## æ ‡é¢˜ä¸º h2
    html_content = re.sub(r'^## (.+)$', r'<h2>\1</h2>', html_content, flags=re.MULTILINE)
    # è½¬æ¢ ### æ ‡é¢˜ä¸º h3
    html_content = re.sub(r'^### (.+)$', r'<h3>\1</h3>', html_content, flags=re.MULTILINE)
    # è½¬æ¢ #### æ ‡é¢˜ä¸º h4
    html_content = re.sub(r'^#### (.+)$', r'<h4>\1</h4>', html_content, flags=re.MULTILINE)
    
    # è½¬æ¢åŠ ç²—æ–‡æœ¬
    html_content = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', html_content)
    
    # è½¬æ¢åˆ—è¡¨é¡¹ (- å¼€å¤´)
    html_content = re.sub(r'^- (.+)$', r'<li>\1</li>', html_content, flags=re.MULTILINE)
    
    # è½¬æ¢æœ‰åºåˆ—è¡¨é¡¹ (1. å¼€å¤´)
    html_content = re.sub(r'^(\d+)\. (.+)$', r'<li>\2</li>', html_content, flags=re.MULTILINE)
    
    # å°†è¿ç»­çš„ <li> åŒ…è£¹åœ¨ <ul> ä¸­
    html_content = re.sub(r'((?:<li>.*?</li>\s*)+)', r'<ul>\1</ul>', html_content, flags=re.DOTALL)
    
    # è½¬æ¢æ®µè½ï¼ˆéç©ºè¡Œä¸”ä¸æ˜¯HTMLæ ‡ç­¾å¼€å¤´çš„è¡Œï¼‰
    lines = html_content.split('\n')
    processed_lines = []
    for line in lines:
        stripped = line.strip()
        if stripped and not stripped.startswith('<') and not stripped.startswith('#'):
            processed_lines.append(f'<p>{stripped}</p>')
        else:
            processed_lines.append(line)
    html_content = '\n'.join(processed_lines)
    
    # æ¸…ç†å¤šä½™çš„ç©ºè¡Œ
    html_content = re.sub(r'\n{3,}', '\n\n', html_content)
    
    # ä½¿ç”¨Streamlitæ¸²æŸ“æ•´ä¸ªæ–‡æ¡£ï¼ˆåŒ…æ‹¬æ ‡é¢˜å’Œå†…å®¹ï¼‰åœ¨åŒä¸€ä¸ªå®¹å™¨ä¸­
    st.markdown(f"""
    <div class="prd-document">
        <div style="text-align: center; margin-bottom: 25px;">
            <h1 style="color: #1a73e8; border-bottom: 2px solid #1a73e8; padding-bottom: 10px; display: inline-block; margin: 0;">
                ğŸ“„ {title}
            </h1>
        </div>
        <div class="prd-content">
            {html_content}
        </div>
        <hr style="border: none; border-top: 1px dashed #ccc; margin-top: 30px;">
    </div>
    """, unsafe_allow_html=True)


def is_file_upload_supported() -> bool:
    """æ£€æŸ¥å½“å‰é€‰æ‹©çš„æ¨¡å‹æ˜¯å¦æ”¯æŒæ–‡ä»¶ä¸Šä¼ """
    current_model = get_selected_model()
    # æ£€æŸ¥æ¨¡å‹åç§°æ˜¯å¦åœ¨æ”¯æŒåˆ—è¡¨ä¸­ï¼ˆéƒ¨åˆ†åŒ¹é…ï¼‰
    for supported_model in FILE_UPLOAD_SUPPORTED_MODELS:
        if supported_model in current_model or current_model in supported_model:
            return True
    return False


def get_gemini_client():
    """è·å–Geminiå®¢æˆ·ç«¯å®ä¾‹"""
    api_key = st.session_state.get("api_key", "")
    if not api_key:
        st.error("âš ï¸ è¯·å…ˆåœ¨ä¾§è¾¹æ é…ç½® API Key")
        return None
    try:
        client = genai.Client(api_key=api_key)
        return client
    except Exception as e:
        st.error(f"APIåˆå§‹åŒ–å¤±è´¥: {str(e)}")
        return None


def get_selected_model():
    """è·å–å½“å‰é€‰æ‹©çš„æ¨¡å‹"""
    return st.session_state.get("selected_model", AVAILABLE_MODELS[0])


def fetch_available_models():
    """ä»APIè·å–å¯ç”¨çš„æ¨¡å‹åˆ—è¡¨"""
    api_key = st.session_state.get("api_key", "")
    if not api_key:
        return []
    try:
        client = genai.Client(api_key=api_key)
        models = []
        for model in client.models.list():
            # åªè·å–æ”¯æŒgenerateContentçš„æ¨¡å‹
            if hasattr(model, 'supported_actions') and 'generateContent' in model.supported_actions:
                models.append(model.name.replace("models/", ""))
            elif not hasattr(model, 'supported_actions'):
                # å¦‚æœæ²¡æœ‰supported_actionså±æ€§ï¼Œä¹Ÿæ·»åŠ ï¼ˆå…¼å®¹æ€§å¤„ç†ï¼‰
                model_name = model.name.replace("models/", "")
                if 'gemini' in model_name.lower():
                    models.append(model_name)
        return sorted(models) if models else AVAILABLE_MODELS
    except Exception as e:
        st.sidebar.warning(f"è·å–æ¨¡å‹åˆ—è¡¨å¤±è´¥ï¼Œä½¿ç”¨é»˜è®¤åˆ—è¡¨: {str(e)}")
        return AVAILABLE_MODELS


def call_gemini(prompt: str, system_prompt: str = "") -> Optional[str]:
    """
    è°ƒç”¨Gemini APIï¼ˆéæµå¼ï¼Œç”¨äºå†…éƒ¨å¤„ç†ï¼‰
    
    Args:
        prompt: ç”¨æˆ·è¾“å…¥çš„æç¤ºè¯
        system_prompt: ç³»ç»Ÿæç¤ºè¯
    
    Returns:
        APIè¿”å›çš„æ–‡æœ¬å†…å®¹ï¼Œå¤±è´¥è¿”å›None
    """
    try:
        client = get_gemini_client()
        if client is None:
            return None
        
        # æ„å»ºé…ç½®
        config = types.GenerateContentConfig(
            system_instruction=system_prompt if system_prompt else None
        )
        
        response = client.models.generate_content(
            model=get_selected_model(),
            contents=prompt,
            config=config
        )
        return response.text
    except Exception as e:
        st.error(f"APIè°ƒç”¨å¤±è´¥: {str(e)}")
        return None


def call_gemini_with_image(image_data: bytes, prompt: str, system_prompt: str = "", mime_type: str = "image/png") -> Optional[str]:
    """
    è°ƒç”¨Gemini APIå¤„ç†å›¾ç‰‡ï¼ˆéæµå¼ï¼‰
    
    Args:
        image_data: å›¾ç‰‡çš„å­—èŠ‚æ•°æ®
        prompt: ç”¨æˆ·è¾“å…¥çš„æç¤ºè¯
        system_prompt: ç³»ç»Ÿæç¤ºè¯
        mime_type: å›¾ç‰‡çš„MIMEç±»å‹ï¼ˆimage/png, image/jpeg, application/pdfï¼‰
    
    Returns:
        APIè¿”å›çš„æ–‡æœ¬å†…å®¹ï¼Œå¤±è´¥è¿”å›None
    """
    try:
        client = get_gemini_client()
        if client is None:
            return None
        
        # æ„å»ºé…ç½®
        config = types.GenerateContentConfig(
            system_instruction=system_prompt if system_prompt else None
        )
        
        # æ„å»ºåŒ…å«å›¾ç‰‡çš„å†…å®¹
        contents = [
            types.Part.from_bytes(data=image_data, mime_type=mime_type),
            prompt
        ]
        
        response = client.models.generate_content(
            model=get_selected_model(),
            contents=contents,
            config=config
        )
        return response.text
    except Exception as e:
        st.error(f"å›¾ç‰‡å¤„ç†APIè°ƒç”¨å¤±è´¥: {str(e)}")
        st.session_state.last_error = str(e)
        return None


def call_gemini_with_image_stream(image_data: bytes, prompt: str, system_prompt: str = "", mime_type: str = "image/png", thinking_container=None) -> Generator[dict, None, None]:
    """
    æµå¼è°ƒç”¨Gemini APIå¤„ç†å›¾ç‰‡ï¼Œæ”¯æŒä¸­æ­¢ã€é”™è¯¯å±•ç¤ºå’Œè‡ªåŠ¨é‡è¯•
    
    Args:
        image_data: å›¾ç‰‡çš„å­—èŠ‚æ•°æ®
        prompt: ç”¨æˆ·è¾“å…¥çš„æç¤ºè¯
        system_prompt: ç³»ç»Ÿæç¤ºè¯
        mime_type: å›¾ç‰‡çš„MIMEç±»å‹
        thinking_container: ç”¨äºæ˜¾ç¤ºæ€è€ƒè¿‡ç¨‹çš„å®¹å™¨ï¼ˆå¯é€‰ï¼‰
    
    Yields:
        dict: {"type": "text"|"thinking"|"error"|"retry", "content": str}
    """
    # æ¸…ç©ºä¹‹å‰çš„é”™è¯¯
    st.session_state.last_error = ""
    st.session_state.thinking_content = ""
    
    # é‡è¯•é…ç½®
    max_retries = 3
    retry_delay = 5
    retryable_errors = ["503", "429", "overloaded", "UNAVAILABLE", "RESOURCE_EXHAUSTED", "rate limit"]
    
    for attempt in range(max_retries):
        try:
            client = get_gemini_client()
            if client is None:
                yield {"type": "error", "content": "APIå®¢æˆ·ç«¯åˆå§‹åŒ–å¤±è´¥ï¼Œè¯·æ£€æŸ¥API Key"}
                return
            
            # æ„å»ºé…ç½®
            config = types.GenerateContentConfig(
                system_instruction=system_prompt if system_prompt else None,
                thinking_config=types.ThinkingConfig(
                    thinking_budget=10000
                ) if "2.5" in get_selected_model() or "think" in get_selected_model().lower() else None
            )
            
            # æ„å»ºåŒ…å«å›¾ç‰‡çš„å†…å®¹
            contents = [
                types.Part.from_bytes(data=image_data, mime_type=mime_type),
                prompt
            ]
            
            # ä½¿ç”¨æµå¼API
            response_stream = client.models.generate_content_stream(
                model=get_selected_model(),
                contents=contents,
                config=config
            )
            
            for chunk in response_stream:
                # æ£€æŸ¥æ˜¯å¦éœ€è¦ä¸­æ­¢
                if st.session_state.should_stop:
                    yield {"type": "stopped", "content": "ç”¨æˆ·å·²ä¸­æ­¢ç”Ÿæˆ"}
                    st.session_state.should_stop = False
                    return
                
                # å¤„ç†æ€è€ƒè¿‡ç¨‹ï¼ˆå¦‚æœæœ‰ï¼‰
                if hasattr(chunk, 'candidates') and chunk.candidates:
                    for candidate in chunk.candidates:
                        if hasattr(candidate, 'content') and candidate.content:
                            for part in candidate.content.parts:
                                # æ£€æŸ¥æ˜¯å¦æ˜¯æ€è€ƒå†…å®¹ - thought å±æ€§ç›´æ¥åŒ…å«æ€è€ƒæ–‡æœ¬
                                thinking_text = ""
                                
                                # thought å±æ€§ç›´æ¥åŒ…å«æ€è€ƒæ–‡æœ¬
                                if hasattr(part, 'thought') and part.thought:
                                    thinking_text = part.thought
                                
                                if thinking_text:
                                    st.session_state.thinking_content += thinking_text
                                    yield {"type": "thinking", "content": thinking_text}
                                elif hasattr(part, 'text') and part.text:
                                    yield {"type": "text", "content": part.text}
                elif chunk.text:
                    yield {"type": "text", "content": chunk.text}
            
            return
                    
        except Exception as e:
            error_msg = str(e)
            st.session_state.last_error = error_msg
            
            is_retryable = any(err_key in error_msg for err_key in retryable_errors)
            
            if is_retryable and attempt < max_retries - 1:
                remaining = max_retries - attempt - 1
                yield {
                    "type": "retry", 
                    "content": f"âš ï¸ æœåŠ¡æš‚æ—¶ä¸å¯ç”¨ ({error_msg[:50]}...)ï¼Œ{retry_delay}ç§’åè‡ªåŠ¨é‡è¯•ï¼ˆå‰©ä½™{remaining}æ¬¡ï¼‰..."
                }
                time.sleep(retry_delay)
                retry_delay = min(retry_delay * 2, 30)
                continue
            else:
                yield {"type": "error", "content": error_msg}
                return


def call_gemini_stream(prompt: str, system_prompt: str = "", thinking_container=None) -> Generator[dict, None, None]:
    """
    æµå¼è°ƒç”¨Gemini APIï¼Œæ”¯æŒä¸­æ­¢ã€é”™è¯¯å±•ç¤ºã€æ€è€ƒè¿‡ç¨‹å’Œè‡ªåŠ¨é‡è¯•
    
    Args:
        prompt: ç”¨æˆ·è¾“å…¥çš„æç¤ºè¯
        system_prompt: ç³»ç»Ÿæç¤ºè¯
        thinking_container: ç”¨äºæ˜¾ç¤ºæ€è€ƒè¿‡ç¨‹çš„å®¹å™¨ï¼ˆå¯é€‰ï¼‰
    
    Yields:
        dict: {"type": "text"|"thinking"|"error"|"retry", "content": str}
    """
    # æ¸…ç©ºä¹‹å‰çš„é”™è¯¯
    st.session_state.last_error = ""
    st.session_state.thinking_content = ""
    
    # é‡è¯•é…ç½®
    max_retries = 3
    retry_delay = 5  # ç§’
    retryable_errors = ["503", "429", "overloaded", "UNAVAILABLE", "RESOURCE_EXHAUSTED", "rate limit"]
    
    for attempt in range(max_retries):
        try:
            client = get_gemini_client()
            if client is None:
                yield {"type": "error", "content": "APIå®¢æˆ·ç«¯åˆå§‹åŒ–å¤±è´¥ï¼Œè¯·æ£€æŸ¥API Key"}
                return
            
            # è·å–å½“å‰é€‰æ‹©çš„æ¨¡å‹
            selected_model = get_selected_model()
            
            # åˆ¤æ–­æ˜¯å¦å¯ç”¨æ€è€ƒæ¨¡å¼
            enable_thinking = "2.5" in selected_model or "think" in selected_model.lower()
            print(f"[DEBUG] Selected model: {selected_model}")
            print(f"[DEBUG] Enable thinking: {enable_thinking}")
            
            # æ„å»ºé…ç½® - å¯ç”¨æ€è€ƒè¿‡ç¨‹ï¼ˆå¦‚æœæ¨¡å‹æ”¯æŒï¼‰
            config = types.GenerateContentConfig(
                system_instruction=system_prompt if system_prompt else None,
                # å°è¯•å¯ç”¨æ€è€ƒæ¨¡å¼ï¼ˆéƒ¨åˆ†æ¨¡å‹æ”¯æŒï¼‰
                thinking_config=types.ThinkingConfig(
                    thinking_budget=10000  # å…è®¸çš„æ€è€ƒtokenæ•°
                ) if enable_thinking else None
            )
            
            print(f"[DEBUG] Config thinking_config: {config.thinking_config}")
            
            # ä½¿ç”¨æµå¼API
            response_stream = client.models.generate_content_stream(
                model=get_selected_model(),
                contents=prompt,
                config=config
            )
            
            # è°ƒè¯•æ ‡è®°ï¼Œåªæ‰“å°ä¸€æ¬¡
            debug_printed = False
            
            for chunk in response_stream:
                # æ£€æŸ¥æ˜¯å¦éœ€è¦ä¸­æ­¢
                if st.session_state.should_stop:
                    yield {"type": "stopped", "content": "ç”¨æˆ·å·²ä¸­æ­¢ç”Ÿæˆ"}
                    st.session_state.should_stop = False
                    return
                
                # å¤„ç†æ€è€ƒè¿‡ç¨‹ï¼ˆå¦‚æœæœ‰ï¼‰
                if hasattr(chunk, 'candidates') and chunk.candidates:
                    for candidate in chunk.candidates:
                        if hasattr(candidate, 'content') and candidate.content:
                            for part in candidate.content.parts:
                                # è·å– part çš„ç±»å‹åï¼ˆç”¨äºè°ƒè¯•å’Œæ£€æµ‹ï¼‰
                                part_type = type(part).__name__
                                
                                # è°ƒè¯•ï¼šæ‰“å° part çš„æ‰€æœ‰å±æ€§ï¼ˆä»…é¦–æ¬¡ï¼‰
                                if not debug_printed:
                                    part_attrs = [attr for attr in dir(part) if not attr.startswith('_')]
                                    print(f"[DEBUG call_gemini_stream] Part type: {part_type}")
                                    print(f"[DEBUG call_gemini_stream] Part attributes: {part_attrs}")
                                    # æ‰“å°ä¸€äº›å…³é”®å±æ€§çš„å€¼
                                    for attr in ['thought', 'thinking', 'text']:
                                        if hasattr(part, attr):
                                            val = getattr(part, attr)
                                            print(f"[DEBUG call_gemini_stream] part.{attr} = {repr(val)[:100] if val else None}")
                                    debug_printed = True
                                
                                # æ£€æŸ¥æ˜¯å¦æ˜¯æ€è€ƒå†…å®¹ - thought å±æ€§ç›´æ¥åŒ…å«æ€è€ƒæ–‡æœ¬
                                thinking_text = ""
                                
                                # æ–¹å¼1: æ£€æŸ¥ thought å±æ€§ï¼ˆç›´æ¥åŒ…å«æ€è€ƒæ–‡æœ¬ï¼‰
                                if hasattr(part, 'thought') and part.thought:
                                    thinking_text = part.thought
                                    print(f"[DEBUG] Found thinking content: {thinking_text[:50]}...")
                                
                                if thinking_text:
                                    st.session_state.thinking_content += thinking_text
                                    yield {"type": "thinking", "content": thinking_text}
                                elif hasattr(part, 'text') and part.text:
                                    yield {"type": "text", "content": part.text}
                elif chunk.text:
                    yield {"type": "text", "content": chunk.text}
            
            # æˆåŠŸå®Œæˆï¼Œé€€å‡ºé‡è¯•å¾ªç¯
            return
                    
        except Exception as e:
            error_msg = str(e)
            st.session_state.last_error = error_msg
            
            # æ£€æŸ¥æ˜¯å¦æ˜¯å¯é‡è¯•çš„é”™è¯¯
            is_retryable = any(err_key in error_msg for err_key in retryable_errors)
            
            if is_retryable and attempt < max_retries - 1:
                # é€šçŸ¥ç”¨æˆ·æ­£åœ¨é‡è¯•
                remaining = max_retries - attempt - 1
                yield {
                    "type": "retry", 
                    "content": f"âš ï¸ æœåŠ¡æš‚æ—¶ä¸å¯ç”¨ ({error_msg[:50]}...)ï¼Œ{retry_delay}ç§’åè‡ªåŠ¨é‡è¯•ï¼ˆå‰©ä½™{remaining}æ¬¡ï¼‰..."
                }
                time.sleep(retry_delay)
                # å¢åŠ ä¸‹æ¬¡é‡è¯•çš„ç­‰å¾…æ—¶é—´ï¼ˆæŒ‡æ•°é€€é¿ï¼‰
                retry_delay = min(retry_delay * 2, 30)
                continue
            else:
                # ä¸å¯é‡è¯•æˆ–å·²ç”¨å®Œé‡è¯•æ¬¡æ•°
                yield {"type": "error", "content": error_msg}
                return


def stream_to_container(prompt: str, system_prompt: str, container, thinking_container=None, status_container=None) -> tuple:
    """
    æµå¼è¾“å‡ºåˆ°Streamlitå®¹å™¨ï¼Œå®æ—¶æ˜¾ç¤ºæ‰“å­—æ•ˆæœï¼Œæ”¯æŒä¸­æ­¢ã€é”™è¯¯å±•ç¤ºå’Œæ€è€ƒè¿‡ç¨‹
    
    Args:
        prompt: ç”¨æˆ·è¾“å…¥çš„æç¤ºè¯
        system_prompt: ç³»ç»Ÿæç¤ºè¯
        container: Streamlitå®¹å™¨å¯¹è±¡ï¼ˆå¦‚st.empty()æˆ–st.container()ï¼‰
        thinking_container: ç”¨äºæ˜¾ç¤ºæ€è€ƒè¿‡ç¨‹çš„å®¹å™¨ï¼ˆå¯é€‰ï¼‰
        status_container: ç”¨äºæ˜¾ç¤ºçŠ¶æ€ä¿¡æ¯çš„å®¹å™¨ï¼ˆå¯é€‰ï¼‰
    
    Returns:
        tuple: (å®Œæ•´çš„å“åº”æ–‡æœ¬, æ˜¯å¦æˆåŠŸ, é”™è¯¯ä¿¡æ¯)
    """
    full_response = ""
    thinking_text = ""
    error_msg = ""
    was_stopped = False
    
    # ä½¿ç”¨ç”Ÿæˆå™¨è¿›è¡Œæµå¼è¾“å‡º
    for chunk_data in call_gemini_stream(prompt, system_prompt, thinking_container):
        chunk_type = chunk_data.get("type", "text")
        chunk_content = chunk_data.get("content", "")
        
        if chunk_type == "text":
            full_response += chunk_content
            # å®æ—¶æ›´æ–°æ˜¾ç¤ºå†…å®¹ï¼Œæ·»åŠ å…‰æ ‡æ•ˆæœ
            container.markdown(full_response + " â–Œ")
        elif chunk_type == "thinking":
            thinking_text += chunk_content
            # æ˜¾ç¤ºæ€è€ƒè¿‡ç¨‹
            if thinking_container:
                thinking_container.markdown(f"ğŸ’­ **æ¨¡å‹æ€è€ƒä¸­...**\n\n{thinking_text}")
        elif chunk_type == "retry":
            # æ˜¾ç¤ºé‡è¯•çŠ¶æ€
            if status_container:
                status_container.warning(chunk_content)
            else:
                st.warning(chunk_content)
        elif chunk_type == "error":
            error_msg = chunk_content
            if status_container:
                status_container.error(f"âŒ APIè°ƒç”¨å¤±è´¥: {error_msg}")
            else:
                st.error(f"âŒ APIè°ƒç”¨å¤±è´¥: {error_msg}")
            break
        elif chunk_type == "stopped":
            was_stopped = True
            if status_container:
                status_container.warning("â¹ï¸ ç”Ÿæˆå·²ä¸­æ­¢")
            break
        
        # å¼ºåˆ¶åˆ·æ–°æ˜¾ç¤º
        time.sleep(0.01)
    
    # ç§»é™¤å…‰æ ‡ï¼Œæ˜¾ç¤ºæœ€ç»ˆç»“æœ
    if full_response:
        container.markdown(full_response)
    
    # åˆ¤æ–­æ˜¯å¦æˆåŠŸ
    success = bool(full_response) and not error_msg and not was_stopped
    
    return (full_response, success, error_msg)


def stream_generator(prompt: str, system_prompt: str):
    """
    åˆ›å»ºæµå¼è¾“å‡ºç”Ÿæˆå™¨ï¼Œé…åˆst.write_streamä½¿ç”¨
    
    Args:
        prompt: ç”¨æˆ·è¾“å…¥çš„æç¤ºè¯
        system_prompt: ç³»ç»Ÿæç¤ºè¯
    
    Yields:
        æ–‡æœ¬ç‰‡æ®µ
    """
    for chunk_data in call_gemini_stream(prompt, system_prompt):
        if chunk_data.get("type") == "text":
            yield chunk_data.get("content", "")


def generate_prd(user_input: str, use_stream: bool = False, container=None, thinking_container=None, status_container=None) -> tuple:
    """
    åŠŸèƒ½æ¨¡å—1ï¼šç”Ÿæˆç­–åˆ’æ¡ˆï¼ˆæ”¯æŒæµå¼è¾“å‡ºï¼‰
    
    Args:
        user_input: ç”¨æˆ·è¾“å…¥çš„åŠŸèƒ½æè¿°
        use_stream: æ˜¯å¦ä½¿ç”¨æµå¼è¾“å‡º
        container: Streamlitå®¹å™¨å¯¹è±¡ï¼Œç”¨äºæµå¼æ˜¾ç¤º
        thinking_container: ç”¨äºæ˜¾ç¤ºæ€è€ƒè¿‡ç¨‹çš„å®¹å™¨
        status_container: ç”¨äºæ˜¾ç¤ºçŠ¶æ€ä¿¡æ¯çš„å®¹å™¨
    
    Returns:
        tuple: (ç”Ÿæˆçš„ç­–åˆ’æ¡ˆæ–‡æœ¬, æ˜¯å¦æˆåŠŸ, é”™è¯¯ä¿¡æ¯)
    """
    prompt = f"è¯·æ ¹æ®ä»¥ä¸‹åŠŸèƒ½æè¿°ç”Ÿæˆå®Œæ•´çš„ç­–åˆ’æ¡ˆï¼š\n\n{user_input}"
    
    if use_stream and container:
        return stream_to_container(prompt, get_system_prompt_with_date(GENERATE_PRD_SYSTEM_PROMPT), container, thinking_container, status_container)
    else:
        result = call_gemini(prompt, get_system_prompt_with_date(GENERATE_PRD_SYSTEM_PROMPT))
        return (result, result is not None, st.session_state.last_error if not result else "")


def ai_self_check(prd_content: str, use_stream: bool = False, container=None, thinking_container=None, status_container=None) -> tuple:
    """
    AIè‡ªæ£€åŠŸèƒ½ï¼šå¯¹ç­–åˆ’æ¡ˆè¿›è¡Œå¤æ£€æ¸…å•æ£€æŸ¥ï¼ˆæ”¯æŒæµå¼è¾“å‡ºï¼‰
    
    Args:
        prd_content: ç­–åˆ’æ¡ˆå†…å®¹
        use_stream: æ˜¯å¦ä½¿ç”¨æµå¼è¾“å‡º
        container: Streamlitå®¹å™¨å¯¹è±¡ï¼Œç”¨äºæµå¼æ˜¾ç¤º
        thinking_container: ç”¨äºæ˜¾ç¤ºæ€è€ƒè¿‡ç¨‹çš„å®¹å™¨
        status_container: ç”¨äºæ˜¾ç¤ºçŠ¶æ€ä¿¡æ¯çš„å®¹å™¨
    
    Returns:
        tuple: (æ£€æŸ¥ç»“æœæŠ¥å‘Š, æ˜¯å¦æˆåŠŸ, é”™è¯¯ä¿¡æ¯)
    """
    prompt = f"""è¯·å¯¹ä»¥ä¸‹ç­–åˆ’æ¡ˆè¿›è¡Œå¤æ£€æ¸…å•æ£€æŸ¥ï¼š

{prd_content}

è¯·é€ä¸€æ£€æŸ¥æ¯ä¸€é¡¹ï¼Œç»™å‡ºè¯¦ç»†çš„æ£€æŸ¥ç»“æœã€‚"""
    
    if use_stream and container:
        return stream_to_container(prompt, SELF_CHECK_SYSTEM_PROMPT, container, thinking_container, status_container)
    else:
        result = call_gemini(prompt, SELF_CHECK_SYSTEM_PROMPT)
        return (result, result is not None, st.session_state.last_error if not result else "")


def optimize_prd_initial(old_prd: str, feedback: str, use_stream: bool = False, container=None, thinking_container=None, status_container=None) -> tuple:
    """
    ä¼˜åŒ–ç­–åˆ’æ¡ˆ - åˆå§‹ä¿®æ­£ï¼ˆæ”¯æŒæµå¼è¾“å‡ºï¼‰
    
    Args:
        old_prd: æ—§ç­–åˆ’æ¡ˆ
        feedback: ç”¨æˆ·çš„ä¿®æ”¹æ„è§
        use_stream: æ˜¯å¦ä½¿ç”¨æµå¼è¾“å‡º
        container: Streamlitå®¹å™¨å¯¹è±¡ï¼Œç”¨äºæµå¼æ˜¾ç¤º
        thinking_container: ç”¨äºæ˜¾ç¤ºæ€è€ƒè¿‡ç¨‹çš„å®¹å™¨
        status_container: ç”¨äºæ˜¾ç¤ºçŠ¶æ€ä¿¡æ¯çš„å®¹å™¨
    
    Returns:
        tuple: (åˆæ­¥ä¿®æ­£åçš„ç­–åˆ’æ¡ˆ, æ˜¯å¦æˆåŠŸ, é”™è¯¯ä¿¡æ¯)
    """
    prompt = f"""ã€æ—§ç­–åˆ’æ¡ˆã€‘
{old_prd}

ã€ç”¨æˆ·ä¿®æ”¹æ„è§ã€‘
{feedback if feedback else "æ— ç‰¹åˆ«æ„è§ï¼Œè¯·æ ¹æ®å¤æ£€æ¸…å•è¿›è¡Œæ£€æŸ¥å’Œå®Œå–„"}

è¯·æ ¹æ®å¤æ£€æ¸…å•æ£€æŸ¥æ—§æ¡ˆï¼Œç»“åˆç”¨æˆ·æ„è§è¿›è¡Œä¿®æ”¹å’Œå¡«è¡¥ã€‚"""
    
    if use_stream and container:
        return stream_to_container(prompt, INITIAL_FIX_SYSTEM_PROMPT, container, thinking_container, status_container)
    else:
        result = call_gemini(prompt, INITIAL_FIX_SYSTEM_PROMPT)
        return (result, result is not None, st.session_state.last_error if not result else "")


def developer_review(current_prd: str, use_stream: bool = False, container=None, thinking_container=None, status_container=None) -> tuple:
    """
    å¼€å‘äººå‘˜è§’è‰²å®¡æŸ¥ç­–åˆ’æ¡ˆï¼ˆæ”¯æŒæµå¼è¾“å‡ºï¼‰
    
    Args:
        current_prd: å½“å‰ç‰ˆæœ¬çš„ç­–åˆ’æ¡ˆ
        use_stream: æ˜¯å¦ä½¿ç”¨æµå¼è¾“å‡º
        container: Streamlitå®¹å™¨å¯¹è±¡ï¼Œç”¨äºæµå¼æ˜¾ç¤º
        thinking_container: ç”¨äºæ˜¾ç¤ºæ€è€ƒè¿‡ç¨‹çš„å®¹å™¨
        status_container: ç”¨äºæ˜¾ç¤ºçŠ¶æ€ä¿¡æ¯çš„å®¹å™¨
    
    Returns:
        tuple: (å¼€å‘äººå‘˜æå‡ºçš„é—®é¢˜åˆ—è¡¨, æ˜¯å¦æˆåŠŸ, é”™è¯¯ä¿¡æ¯)
    """
    prompt = f"""è¯·å®¡æŸ¥ä»¥ä¸‹ç­–åˆ’æ¡ˆï¼Œæå‡ºä½ çš„é—®é¢˜å’Œç–‘è™‘ï¼š

{current_prd}"""
    
    if use_stream and container:
        return stream_to_container(prompt, DEVELOPER_REVIEW_PROMPT, container, thinking_container, status_container)
    else:
        result = call_gemini(prompt, DEVELOPER_REVIEW_PROMPT)
        return (result, result is not None, st.session_state.last_error if not result else "")


def planner_fix(current_prd: str, dev_questions: str, use_stream: bool = False, container=None, thinking_container=None, status_container=None) -> tuple:
    """
    ç­–åˆ’è§’è‰²æ ¹æ®å¼€å‘äººå‘˜é—®é¢˜ä¿®æ”¹ç­–åˆ’æ¡ˆï¼ˆæ”¯æŒæµå¼è¾“å‡ºï¼‰
    
    Args:
        current_prd: å½“å‰ç‰ˆæœ¬çš„ç­–åˆ’æ¡ˆ
        dev_questions: å¼€å‘äººå‘˜æå‡ºçš„é—®é¢˜
        use_stream: æ˜¯å¦ä½¿ç”¨æµå¼è¾“å‡º
        container: Streamlitå®¹å™¨å¯¹è±¡ï¼Œç”¨äºæµå¼æ˜¾ç¤º
        thinking_container: ç”¨äºæ˜¾ç¤ºæ€è€ƒè¿‡ç¨‹çš„å®¹å™¨
        status_container: ç”¨äºæ˜¾ç¤ºçŠ¶æ€ä¿¡æ¯çš„å®¹å™¨
    
    Returns:
        tuple: (ä¿®æ”¹åçš„ç­–åˆ’æ¡ˆ, æ˜¯å¦æˆåŠŸ, é”™è¯¯ä¿¡æ¯)
    """
    prompt = f"""ã€å½“å‰ç­–åˆ’æ¡ˆã€‘
{current_prd}

ã€å¼€å‘äººå‘˜æå‡ºçš„é—®é¢˜ã€‘
{dev_questions}

è¯·é’ˆå¯¹ä»¥ä¸Šé—®é¢˜ä¿®æ”¹å’Œå®Œå–„ç­–åˆ’æ¡ˆã€‚"""
    
    if use_stream and container:
        return stream_to_container(prompt, PLANNER_FIX_PROMPT, container, thinking_container, status_container)
    else:
        result = call_gemini(prompt, PLANNER_FIX_PROMPT)
        return (result, result is not None, st.session_state.last_error if not result else "")


def reflection_loop(initial_prd: str, max_iterations: int) -> tuple:
    """
    Reflectionå¾ªç¯ä¼˜åŒ–ç­–åˆ’æ¡ˆï¼ˆæµå¼è¾“å‡ºç‰ˆæœ¬ï¼Œæ”¯æŒä¸­æ­¢ï¼‰
    
    Args:
        initial_prd: åˆå§‹ä¿®æ­£åçš„ç­–åˆ’æ¡ˆ
        max_iterations: æœ€å¤§è¿­ä»£è½®æ¬¡
    
    Returns:
        tuple: (æœ€ç»ˆä¼˜åŒ–åçš„ç­–åˆ’æ¡ˆ, æ˜¯å¦è¢«ä¸­æ­¢)
    """
    current_prd = initial_prd
    was_stopped = False
    
    for i in range(max_iterations):
        # æ£€æŸ¥æ˜¯å¦éœ€è¦ä¸­æ­¢
        if st.session_state.should_stop:
            was_stopped = True
            st.warning(f"â¹ï¸ è¿­ä»£å·²åœ¨ç¬¬ {i + 1} è½®å‰ä¸­æ­¢")
            break
            
        st.markdown(f"### ğŸ”„ ç¬¬ {i + 1} è½®è¿­ä»£")
        
        # æ˜¾ç¤ºä¸­æ­¢æŒ‰é’®
        col_status, col_stop = st.columns([4, 1])
        with col_stop:
            if st.button(f"â¹ï¸ ä¸­æ­¢è¿­ä»£", key=f"stop_iteration_{i}", type="secondary"):
                st.session_state.should_stop = True
                was_stopped = True
                st.warning("â¹ï¸ è¿­ä»£å·²ä¸­æ­¢")
                return (current_prd, True)
        
        # è§’è‰²A: å¼€å‘äººå‘˜å®¡æŸ¥
        with st.expander(f"ğŸ“‹ ç¬¬ {i + 1} è½® - å¼€å‘äººå‘˜å®¡æŸ¥", expanded=True):
            st.markdown("**ğŸ” å¼€å‘äººå‘˜æ­£åœ¨å®¡æŸ¥ç­–åˆ’æ¡ˆ...**")
            
            # æ€è€ƒè¿‡ç¨‹å±•ç¤º
            thinking_expander = st.expander("ğŸ’­ æŸ¥çœ‹æ€è€ƒè¿‡ç¨‹", expanded=False)
            with thinking_expander:
                thinking_container = st.empty()
            
            status_container = st.empty()
            dev_container = st.empty()
            
            dev_questions, success, error = developer_review(
                current_prd, 
                use_stream=True, 
                container=dev_container,
                thinking_container=thinking_container,
                status_container=status_container
            )
            
            if st.session_state.should_stop:
                was_stopped = True
                st.warning("â¹ï¸ å·²ä¸­æ­¢")
                break
                
            if success and dev_questions:
                st.success("å®¡æŸ¥å®Œæˆï¼")
            elif error:
                st.error(f"âŒ å®¡æŸ¥å¤±è´¥: {error}")
                st.warning("è·³è¿‡æœ¬è½®")
                continue
            else:
                st.warning("å¼€å‘äººå‘˜å®¡æŸ¥å¤±è´¥ï¼Œè·³è¿‡æœ¬è½®")
                continue
        
        # è§’è‰²B: ç­–åˆ’ä¿®æ”¹
        with st.expander(f"âœï¸ ç¬¬ {i + 1} è½® - ç­–åˆ’ä¼˜åŒ–", expanded=True):
            st.markdown("**âœï¸ ç­–åˆ’é…¸å¥¶æ­£åœ¨ä¼˜åŒ–ç­–åˆ’æ¡ˆ...**")
            
            # æ€è€ƒè¿‡ç¨‹å±•ç¤º
            thinking_expander2 = st.expander("ğŸ’­ æŸ¥çœ‹æ€è€ƒè¿‡ç¨‹", expanded=False)
            with thinking_expander2:
                thinking_container2 = st.empty()
            
            status_container2 = st.empty()
            fix_container = st.empty()
            
            updated_prd, success, error = planner_fix(
                current_prd, 
                dev_questions, 
                use_stream=True, 
                container=fix_container,
                thinking_container=thinking_container2,
                status_container=status_container2
            )
            
            if st.session_state.should_stop:
                was_stopped = True
                st.warning("â¹ï¸ å·²ä¸­æ­¢")
                break
                
            if success and updated_prd:
                current_prd = updated_prd
                st.success(f"ç¬¬ {i + 1} è½®ä¼˜åŒ–å®Œæˆï¼")
            elif error:
                st.error(f"âŒ ä¼˜åŒ–å¤±è´¥: {error}")
                st.warning("ä¿æŒå½“å‰ç‰ˆæœ¬")
            else:
                st.warning("ç­–åˆ’ä¼˜åŒ–å¤±è´¥ï¼Œä¿æŒå½“å‰ç‰ˆæœ¬")
        
        st.markdown("---")
    
    return (current_prd, was_stopped)


def main():
    """ä¸»å‡½æ•°"""
    # é¡µé¢é…ç½®
    st.set_page_config(
        page_title="æ¸¸æˆç­–åˆ’Agentï¼ˆé…¸å¥¶ï¼‰",
        page_icon="ğŸ®",
        layout="wide"
    )
    
    # è‡ªå®šä¹‰CSSæ ·å¼ - ä¼˜åŒ–æ–‡æ¡£æ˜¾ç¤ºæ•ˆæœ
    st.markdown("""
    <style>
    /* ç­–åˆ’æ¡ˆæ–‡æ¡£å®¹å™¨æ ·å¼ */
    .prd-document {
        background-color: #ffffff;
        border: 1px solid #e0e0e0;
        border-radius: 10px;
        padding: 30px 40px;
        margin: 20px 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        line-height: 1.8;
        font-size: 15px;
    }
    
    /* æ·±è‰²æ¨¡å¼é€‚é… */
    @media (prefers-color-scheme: dark) {
        .prd-document {
            background-color: #1e1e1e;
            border-color: #3a3a3a;
        }
    }
    
    /* æ ‡é¢˜æ ·å¼ */
    .prd-document h1 {
        font-size: 24px;
        color: #1a73e8;
        border-bottom: 2px solid #1a73e8;
        padding-bottom: 10px;
        margin: 30px 0 20px 0;
    }
    
    .prd-document h2 {
        font-size: 20px;
        color: #1a73e8;
        border-left: 4px solid #1a73e8;
        padding-left: 12px;
        margin: 25px 0 15px 0;
    }
    
    .prd-document h3 {
        font-size: 17px;
        color: #333;
        margin: 20px 0 12px 0;
    }
    
    .prd-document h4 {
        font-size: 15px;
        color: #555;
        margin: 15px 0 10px 0;
        font-weight: 600;
    }
    
    /* å†…å®¹åŒºåŸŸ */
    .prd-content {
        padding: 10px 0;
    }
    
    /* æ®µè½æ ·å¼ */
    .prd-document p {
        margin: 12px 0;
        text-align: justify;
        line-height: 1.8;
    }
    
    /* åˆ—è¡¨æ ·å¼ */
    .prd-document ul, .prd-document ol {
        margin: 15px 0;
        padding-left: 25px;
    }
    
    .prd-document li {
        margin: 8px 0;
        line-height: 1.7;
    }
    
    /* åŠ ç²—æ–‡æœ¬é«˜äº® */
    .prd-document strong {
        color: #d93025;
        font-weight: 600;
    }
    
    /* ä»£ç å—æ ·å¼ */
    .prd-document code {
        background-color: #f5f5f5;
        padding: 2px 6px;
        border-radius: 4px;
        font-family: 'Consolas', monospace;
    }
    
    /* åˆ†éš”çº¿ */
    .prd-document hr {
        border: none;
        border-top: 1px dashed #ccc;
        margin: 25px 0;
    }
    
    /* ä¸€çº§ç« èŠ‚æ ‡é¢˜ï¼ˆæ•°å­—å¼€å¤´å¦‚ 1ã€åŠŸèƒ½æ¦‚è¿°ï¼‰*/
    .prd-section-title {
        font-size: 18px;
        font-weight: bold;
        color: #1a73e8;
        background: linear-gradient(90deg, #e8f0fe 0%, transparent 100%);
        padding: 10px 15px;
        margin: 25px 0 15px 0;
        border-left: 4px solid #1a73e8;
        border-radius: 0 6px 6px 0;
    }
    
    /* äºŒçº§æ ‡é¢˜ */
    .prd-subsection-title {
        font-size: 16px;
        font-weight: 600;
        color: #333;
        margin: 18px 0 10px 0;
        padding-left: 10px;
        border-left: 3px solid #4285f4;
    }
    
    /* å†…å®¹å— */
    .prd-content-block {
        padding: 10px 15px;
        margin: 10px 0;
        background-color: #fafafa;
        border-radius: 6px;
    }
    
    /* Streamlité»˜è®¤markdownå¢å¼º */
    .stMarkdown {
        line-height: 1.8;
    }
    
    .stMarkdown p {
        margin-bottom: 12px;
    }
    
    /* ç»Ÿä¸€æ ‡é¢˜æ ·å¼ - æ¸…æ™°çš„å±‚çº§åŒºåˆ†ï¼Œå»é™¤çº¢è‰²ä¸»é¢˜ */
    .stMarkdown h1 {
        font-size: 1.75em;
        font-weight: 700;
        color: #1f2937 !important;
        border-bottom: 2px solid #e5e7eb;
        padding-bottom: 8px;
        margin-top: 28px;
        margin-bottom: 16px;
    }
    
    .stMarkdown h2 {
        font-size: 1.4em;
        font-weight: 600;
        color: #374151 !important;
        border-bottom: 1px solid #e5e7eb;
        padding-bottom: 6px;
        margin-top: 24px;
        margin-bottom: 14px;
    }
    
    .stMarkdown h3 {
        font-size: 1.2em;
        font-weight: 600;
        color: #4b5563 !important;
        margin-top: 20px;
        margin-bottom: 12px;
    }
    
    .stMarkdown h4 {
        font-size: 1.1em;
        font-weight: 600;
        color: #6b7280 !important;
        margin-top: 16px;
        margin-bottom: 10px;
    }
    
    .stMarkdown h5, .stMarkdown h6 {
        font-size: 1em;
        font-weight: 600;
        color: #6b7280 !important;
        margin-top: 14px;
        margin-bottom: 8px;
    }
    
    /* æ·±è‰²æ¨¡å¼æ ‡é¢˜é€‚é… */
    @media (prefers-color-scheme: dark) {
        .stMarkdown h1 {
            color: #f3f4f6 !important;
            border-bottom-color: #4b5563;
        }
        .stMarkdown h2 {
            color: #e5e7eb !important;
            border-bottom-color: #4b5563;
        }
        .stMarkdown h3 {
            color: #d1d5db !important;
        }
        .stMarkdown h4, .stMarkdown h5, .stMarkdown h6 {
            color: #9ca3af !important;
        }
    }
    
    .stMarkdown ul, .stMarkdown ol {
        margin: 12px 0;
        padding-left: 24px;
    }
    
    .stMarkdown li {
        margin: 6px 0;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # åˆå§‹åŒ–session_state
    if "generated_prd" not in st.session_state:
        st.session_state.generated_prd = ""
    if "optimized_prd" not in st.session_state:
        st.session_state.optimized_prd = ""
    if "is_processing" not in st.session_state:
        st.session_state.is_processing = False
    
    # åˆå§‹åŒ–ä¼šè¯å†å²
    init_session_history()
    
    # å†å²è¯¦æƒ…æŸ¥çœ‹çŠ¶æ€
    if "viewing_history_id" not in st.session_state:
        st.session_state.viewing_history_id = None
    if "show_history_detail" not in st.session_state:
        st.session_state.show_history_detail = False
    
    # å°è¯•ä» Streamlit Secrets è·å– API Keyï¼ˆç”¨äºäº‘éƒ¨ç½²ï¼‰
    default_api_key = ""
    secrets_api_key_loaded = False
    try:
        if "GOOGLE_API_KEY" in st.secrets:
            default_api_key = st.secrets["GOOGLE_API_KEY"]
            secrets_api_key_loaded = True
        elif "GEMINI_API_KEY" in st.secrets:
            default_api_key = st.secrets["GEMINI_API_KEY"]
            secrets_api_key_loaded = True
    except Exception:
        # æœ¬åœ°è¿è¡Œæ—¶å¯èƒ½æ²¡æœ‰ secrets æ–‡ä»¶
        pass
    
    if "api_key" not in st.session_state:
        st.session_state.api_key = default_api_key
    if "secrets_api_key_loaded" not in st.session_state:
        st.session_state.secrets_api_key_loaded = secrets_api_key_loaded
    if "selected_model" not in st.session_state:
        st.session_state.selected_model = AVAILABLE_MODELS[0]
    if "models_list" not in st.session_state:
        st.session_state.models_list = AVAILABLE_MODELS
    if "api_key_validated" not in st.session_state:
        st.session_state.api_key_validated = False
    # è‡ªåŠ¨éªŒè¯æ ‡å¿—
    if "auto_validate_api_key" not in st.session_state:
        st.session_state.auto_validate_api_key = False
    # ä¸­æ­¢æ§åˆ¶
    if "should_stop" not in st.session_state:
        st.session_state.should_stop = False
    # é”™è¯¯ä¿¡æ¯
    if "last_error" not in st.session_state:
        st.session_state.last_error = ""
    # æ€è€ƒè¿‡ç¨‹
    if "thinking_content" not in st.session_state:
        st.session_state.thinking_content = ""
    
    # ========== ä¾§è¾¹æ  - APIé…ç½® ==========
    with st.sidebar:
        st.header("âš™ï¸ API é…ç½®")
        
        # å¦‚æœä» Secrets åŠ è½½äº† API Keyï¼Œæ˜¾ç¤ºæç¤º
        if st.session_state.secrets_api_key_loaded and st.session_state.api_key:
            st.success("ğŸ” å·²ä»äº‘ç«¯é…ç½®åŠ è½½ API Key")
            # æ˜¾ç¤ºéšè—çš„ API Key çŠ¶æ€
            st.text_input(
                "ğŸ”‘ Gemini API Key",
                type="password",
                value="********ï¼ˆäº‘ç«¯é…ç½®ï¼‰",
                disabled=True,
                help="API Key å·²ä» Streamlit Secrets è‡ªåŠ¨åŠ è½½"
            )
            # æä¾›æ‰‹åŠ¨è¦†ç›–é€‰é¡¹
            with st.expander("âœï¸ ä½¿ç”¨è‡ªå®šä¹‰ API Key"):
                custom_api_key = st.text_input(
                    "è¾“å…¥è‡ªå®šä¹‰ API Key",
                    type="password",
                    placeholder="ç•™ç©ºåˆ™ä½¿ç”¨äº‘ç«¯é…ç½®çš„ Key",
                    help="å¦‚éœ€ä½¿ç”¨è‡ªå·±çš„ API Keyï¼Œè¯·åœ¨æ­¤è¾“å…¥"
                )
                if custom_api_key and custom_api_key != st.session_state.api_key:
                    st.session_state.api_key = custom_api_key
                    st.session_state.secrets_api_key_loaded = False
                    st.session_state.api_key_validated = False
                    # è‡ªåŠ¨è§¦å‘éªŒè¯
                    st.session_state.auto_validate_api_key = True
                    st.rerun()
            api_key_input = st.session_state.api_key
        else:
            # æ‰‹åŠ¨è¾“å…¥ API Key
            api_key_input = st.text_input(
                "ğŸ”‘ Gemini API Key",
                type="password",
                value=st.session_state.api_key,
                placeholder="è¯·è¾“å…¥æ‚¨çš„ Gemini API Key",
                help="è¯·å‰å¾€ Google AI Studio è·å– API Key: https://aistudio.google.com/apikey"
            )
            
            # æ£€æµ‹API Keyå˜åŒ– - è‡ªåŠ¨è§¦å‘éªŒè¯
            if api_key_input != st.session_state.api_key:
                st.session_state.api_key = api_key_input
                st.session_state.api_key_validated = False
                st.session_state.models_list = AVAILABLE_MODELS
                # å¦‚æœæ–°çš„API Keyéç©ºï¼Œè‡ªåŠ¨è§¦å‘éªŒè¯
                if api_key_input:
                    st.session_state.auto_validate_api_key = True
                    st.rerun()
        
        # è‡ªåŠ¨éªŒè¯API Keyï¼ˆå½“æ£€æµ‹åˆ°éœ€è¦è‡ªåŠ¨éªŒè¯æ—¶ï¼‰
        if st.session_state.get('auto_validate_api_key', False) and api_key_input:
            st.session_state.auto_validate_api_key = False
            with st.spinner("æ­£åœ¨è‡ªåŠ¨éªŒè¯API Keyå¹¶è·å–æ¨¡å‹åˆ—è¡¨..."):
                models = fetch_available_models()
                if models:
                    st.session_state.models_list = models
                    st.session_state.api_key_validated = True
                    st.success(f"âœ… éªŒè¯æˆåŠŸï¼è·å–åˆ° {len(models)} ä¸ªå¯ç”¨æ¨¡å‹")
                else:
                    st.error("âŒ API Key æ— æ•ˆæˆ–æ— æ³•è·å–æ¨¡å‹åˆ—è¡¨")
                    st.session_state.api_key_validated = False
        
        # éªŒè¯å¹¶è·å–æ¨¡å‹åˆ—è¡¨æŒ‰é’®ï¼ˆæ‰‹åŠ¨åˆ·æ–°ï¼‰
        col1, col2 = st.columns(2)
        with col1:
            if st.button("ğŸ”„ åˆ·æ–°æ¨¡å‹åˆ—è¡¨", disabled=not api_key_input):
                if api_key_input:
                    with st.spinner("æ­£åœ¨éªŒè¯API Keyå¹¶è·å–æ¨¡å‹åˆ—è¡¨..."):
                        models = fetch_available_models()
                        if models:
                            st.session_state.models_list = models
                            st.session_state.api_key_validated = True
                            st.success(f"âœ… éªŒè¯æˆåŠŸï¼è·å–åˆ° {len(models)} ä¸ªå¯ç”¨æ¨¡å‹")
                        else:
                            st.error("âŒ API Key æ— æ•ˆæˆ–æ— æ³•è·å–æ¨¡å‹åˆ—è¡¨")
                            st.session_state.api_key_validated = False
        
        with col2:
            if st.session_state.api_key_validated:
                st.markdown("âœ… å·²éªŒè¯")
            elif api_key_input:
                st.markdown("âš ï¸ æœªéªŒè¯")
        
        # äº‘ç«¯éƒ¨ç½²æç¤º
        if st.session_state.secrets_api_key_loaded:
            st.caption("ğŸ’¡ äº‘ç«¯éƒ¨ç½²æ¨¡å¼ï¼šAPI Key å·²å®‰å…¨å­˜å‚¨")
        
        st.markdown("---")
        
        # æ¨¡å‹é€‰æ‹©
        st.subheader("ğŸ¤– æ¨¡å‹é€‰æ‹©")
        
        # æ¨¡å‹ä¸‹æ‹‰é€‰æ‹©æ¡†
        selected_model = st.selectbox(
            "é€‰æ‹©æ¨¡å‹",
            options=st.session_state.models_list,
            index=0 if st.session_state.selected_model not in st.session_state.models_list 
                  else st.session_state.models_list.index(st.session_state.selected_model),
            help="é€‰æ‹©è¦ä½¿ç”¨çš„ Gemini æ¨¡å‹"
        )
        st.session_state.selected_model = selected_model
        
        # æ˜¾ç¤ºå½“å‰é€‰æ‹©çš„æ¨¡å‹
        st.info(f"å½“å‰æ¨¡å‹: **{selected_model}**")
        
        st.markdown("---")
        
        # å¸®åŠ©ä¿¡æ¯
        with st.expander("ğŸ“– ä½¿ç”¨å¸®åŠ©"):
            st.markdown("""
            **å¦‚ä½•è·å– API Keyï¼š**
            1. è®¿é—® [Google AI Studio](https://aistudio.google.com/apikey)
            2. ç™»å½•æ‚¨çš„ Google è´¦å·
            3. ç‚¹å‡» "Create API Key" åˆ›å»ºå¯†é’¥
            4. å¤åˆ¶å¯†é’¥å¹¶ç²˜è´´åˆ°ä¸Šæ–¹è¾“å…¥æ¡†
            
            **æ¨¡å‹è¯´æ˜ï¼š**
            - `gemini-2.5-*`: æœ€æ–°ä¸€ä»£æ¨¡å‹ï¼Œèƒ½åŠ›æœ€å¼º
            - `gemini-2.0-flash`: é€Ÿåº¦å¿«ï¼Œé€‚åˆå¤§å¤šæ•°åœºæ™¯
            - `gemini-1.5-pro`: ä¸Šä¸€ä»£Proæ¨¡å‹ï¼Œç¨³å®šå¯é 
            - `gemini-1.5-flash`: è½»é‡å¿«é€Ÿæ¨¡å‹
            
            **æ³¨æ„äº‹é¡¹ï¼š**
            - ç‚¹å‡»"éªŒè¯ & åˆ·æ–°æ¨¡å‹"å¯è·å–æœ€æ–°çš„å¯ç”¨æ¨¡å‹åˆ—è¡¨
            - ä¸åŒæ¨¡å‹çš„èƒ½åŠ›å’Œå“åº”é€Ÿåº¦æœ‰æ‰€ä¸åŒ
            - API Key ä»…å­˜å‚¨åœ¨æœ¬åœ°æµè§ˆå™¨ä¼šè¯ä¸­
            
            **äº‘ç«¯éƒ¨ç½²ï¼ˆStreamlit Cloudï¼‰ï¼š**
            - æ”¯æŒé€šè¿‡ Secrets å®‰å…¨é…ç½® API Key
            - åœ¨ Streamlit Cloud çš„ Settings â†’ Secrets ä¸­æ·»åŠ ï¼š
            ```
            GOOGLE_API_KEY = "your-api-key"
            ```
            - æœ¬åœ°å¼€å‘æ—¶ï¼Œå¯åœ¨é¡¹ç›®æ ¹ç›®å½•åˆ›å»º `.streamlit/secrets.toml`
            """)
        
        st.markdown("---")
        st.caption("Powered by Google Gemini API")
        
        # æ¸²æŸ“ä¼šè¯å†å²ä¾§è¾¹æ 
        render_history_sidebar()
    
    # ========== ä¸»ç•Œé¢ ==========
    # æ ‡é¢˜
    st.title("ğŸ® æ¸¸æˆç­–åˆ’Agentï¼ˆé…¸å¥¶ï¼‰")
    st.markdown("*åŸºäºGemini APIçš„æ™ºèƒ½ç­–åˆ’è¾…åŠ©å·¥å…·*")
    st.markdown("---")
    
    # æ£€æŸ¥API Key
    if not st.session_state.api_key:
        st.warning("âš ï¸ è¯·åœ¨å·¦ä¾§ä¾§è¾¹æ é…ç½® API Key åä½¿ç”¨æœ¬å·¥å…·")
        st.info("ğŸ‘ˆ ç‚¹å‡»å·¦ä¾§ä¾§è¾¹æ è¾“å…¥æ‚¨çš„ Gemini API Key")
        
        # æ˜¾ç¤ºå¿«é€ŸæŒ‡å—
        with st.expander("ğŸš€ å¿«é€Ÿå¼€å§‹æŒ‡å—", expanded=True):
            st.markdown("""
            ### ç¬¬ä¸€æ­¥ï¼šè·å– API Key
            1. è®¿é—® [Google AI Studio](https://aistudio.google.com/apikey)
            2. ä½¿ç”¨ Google è´¦å·ç™»å½•
            3. ç‚¹å‡» "Create API Key" æŒ‰é’®
            4. å¤åˆ¶ç”Ÿæˆçš„ API Key
            
            ### ç¬¬äºŒæ­¥ï¼šé…ç½®å·¥å…·
            1. åœ¨å·¦ä¾§ä¾§è¾¹æ çš„ "API Key" è¾“å…¥æ¡†ä¸­ç²˜è´´æ‚¨çš„å¯†é’¥
            2. ç‚¹å‡» "éªŒè¯ & åˆ·æ–°æ¨¡å‹" æŒ‰é’®éªŒè¯å¯†é’¥
            3. é€‰æ‹©æ‚¨æƒ³è¦ä½¿ç”¨çš„æ¨¡å‹
            
            ### ç¬¬ä¸‰æ­¥ï¼šå¼€å§‹ä½¿ç”¨
            - **ç”Ÿæˆç­–åˆ’æ¡ˆ**ï¼šè¾“å…¥åŠŸèƒ½æè¿°ï¼ŒAIå°†ç”Ÿæˆå®Œæ•´çš„ç­–åˆ’æ¡ˆ
            - **ä¼˜åŒ–ç­–åˆ’æ¡ˆ**ï¼šè¾“å…¥ç°æœ‰ç­–åˆ’æ¡ˆï¼ŒAIå°†é€šè¿‡å¤šè½®è¿­ä»£ä¼˜åŒ–
            - **æ±‡æŠ¥åŠ©æ‰‹**ï¼šå°†å·¥ä½œä¿¡æ¯è½¬åŒ–ä¸ºç»“æ„åŒ–æ±‡æŠ¥æ–‡æ¡ˆ
            """)
        st.stop()
    
    # ========== å†å²è¯¦æƒ…æŸ¥çœ‹åŒºåŸŸ ==========
    if st.session_state.get("show_history_detail") and st.session_state.get("viewing_history_id"):
        history_id = st.session_state.viewing_history_id
        # æŸ¥æ‰¾å¯¹åº”çš„å†å²è®°å½•
        history_item = None
        for item in st.session_state.session_history:
            if item.get("id") == history_id:
                history_item = item
                break
        
        if history_item:
            st.markdown("---")
            st.markdown(f"### ğŸ“œ å†å²è®°å½•è¯¦æƒ… #{history_id}")
            
            # å…³é—­æŒ‰é’®
            if st.button("âŒ å…³é—­è¯¦æƒ…", key="close_history_detail"):
                st.session_state.show_history_detail = False
                st.session_state.viewing_history_id = None
                st.rerun()
            
            col_info1, col_info2 = st.columns(2)
            with col_info1:
                st.markdown(f"**åŠŸèƒ½ç±»å‹ï¼š** {history_item.get('function_type', 'æœªçŸ¥')}")
            with col_info2:
                st.markdown(f"**ç”Ÿæˆæ—¶é—´ï¼š** {history_item.get('timestamp', 'æœªçŸ¥')}")
            
            # æ˜¾ç¤ºè¾“å…¥æ•°æ®
            with st.expander("ğŸ“¥ è¾“å…¥å†…å®¹", expanded=False):
                input_data = history_item.get("input_data", {})
                for key, value in input_data.items():
                    st.markdown(f"**{key}ï¼š**")
                    st.text(str(value)[:500] + ("..." if len(str(value)) > 500 else ""))
            
            # æ˜¾ç¤ºè¾“å‡ºæ•°æ®
            with st.expander("ğŸ“¤ è¾“å‡ºå†…å®¹", expanded=True):
                st.markdown(history_item.get("output_data", ""))
            
            # ä¸‹è½½æŒ‰é’®
            if history_item.get("download_data"):
                st.download_button(
                    label=f"ğŸ“¥ ä¸‹è½½ {history_item.get('download_filename', 'æ–‡ä»¶')}",
                    data=get_download_data(history_item),
                    file_name=history_item.get("download_filename", "download.txt"),
                    mime=history_item.get("download_mime", "text/plain"),
                    key=f"history_download_{history_id}"
                )
            
            st.markdown("---")
    
    # åˆå§‹åŒ–åŠŸèƒ½é€‰æ‹©çš„session state
    if "selected_function" not in st.session_state:
        st.session_state.selected_function = "ç”Ÿæˆç­–åˆ’æ¡ˆ"
    
    # åŠŸèƒ½é€‰é¡¹åˆ—è¡¨
    function_options = ["ç”Ÿæˆç­–åˆ’æ¡ˆ", "è„‘å›¾ç”Ÿæˆç­–åˆ’æ¡ˆ", "ä¼˜åŒ–ç­–åˆ’æ¡ˆ", "æ±‡æŠ¥åŠ©æ‰‹", "å‘¨æŠ¥åŠ©æ‰‹", "ç™½çš®ä¹¦åŠ©æ‰‹", "æ¸¸æˆç­–åˆ’(lina)", "è¡¨æ ¼å¤„ç†åŠ©æ‰‹", "æ€è·¯å¼•å¯¼åŠ©æ‰‹ (linmo)", "PUBGM WoW ç©æ³•è¯„å®¡"]
    
    # è·å–å½“å‰é€‰ä¸­çš„ç´¢å¼•
    current_index = function_options.index(st.session_state.selected_function) if st.session_state.selected_function in function_options else 0
    
    # åŠŸèƒ½é€‰æ‹©
    function_mode = st.selectbox(
        "ğŸ”§ åŠŸèƒ½é€‰æ‹©",
        options=function_options,
        index=current_index,
        help="é€‰æ‹©è¦ä½¿ç”¨çš„åŠŸèƒ½",
        key="function_selectbox"
    )
    
    # æ›´æ–°session state
    st.session_state.selected_function = function_mode
    
    # æ ¹æ®åŠŸèƒ½æ¨¡å¼æ˜¾ç¤ºä¸åŒçš„è¾“å…¥ç•Œé¢
    if function_mode == "ç”Ÿæˆç­–åˆ’æ¡ˆ":
        # æ ‡é¢˜å’Œæ¸…ç©ºæŒ‰é’®
        title_col, clear_col = st.columns([6, 1])
        with title_col:
            st.markdown("### ğŸ“ ç”Ÿæˆæ–°ç­–åˆ’æ¡ˆ")
        with clear_col:
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºä¼šè¯", key="clear_generate_session", use_container_width=True):
                clear_module_session("ç”Ÿæˆç­–åˆ’æ¡ˆ")
                st.rerun()
        st.markdown("è¯·è¾“å…¥åŠŸèƒ½æè¿°ï¼ŒAIå°†ä¸ºæ‚¨ç”Ÿæˆå®Œæ•´çš„ç­–åˆ’æ¡ˆã€‚")
        
        user_input = st.text_area(
            "åŠŸèƒ½æè¿°",
            height=300,
            placeholder="è¯·è¯¦ç»†æè¿°æ‚¨è¦è®¾è®¡çš„æ¸¸æˆåŠŸèƒ½...\n\nä¾‹å¦‚ï¼š\nè®¾è®¡ä¸€ä¸ªæ¸¸æˆå†…çš„å¥½å‹ç³»ç»Ÿï¼ŒåŒ…æ‹¬æ·»åŠ å¥½å‹ã€åˆ é™¤å¥½å‹ã€å¥½å‹åˆ—è¡¨å±•ç¤ºã€åœ¨çº¿çŠ¶æ€æ˜¾ç¤ºç­‰åŠŸèƒ½...",
            key="generate_input"
        )
        
        # ========== æ–‡ä»¶ä¸Šä¼ åŒºåŸŸï¼ˆè¾“å…¥æ¡†å³ä¸‹æ–¹ï¼‰==========
        if is_file_upload_supported():
            # åˆ›å»ºå¸ƒå±€ï¼šå·¦è¾¹æ˜¯ç©ºçš„å ä½ï¼Œå³è¾¹æ˜¯æ–‡ä»¶ä¸Šä¼ 
            upload_col1, upload_col2 = st.columns([2, 1])
            
            with upload_col2:
                uploaded_file = st.file_uploader(
                    "ğŸ“ ä¸Šä¼ é™„ä»¶",
                    type=SUPPORTED_FILE_TYPES,
                    help="ä¸Šä¼ å‚è€ƒæ–‡æ¡£ä¾›AIå‚è€ƒï¼ˆPDF/Word/TXT/MDï¼‰",
                    key="generate_file_uploader"
                )
                
                # å¤„ç†ä¸Šä¼ çš„æ–‡ä»¶
                if uploaded_file is not None:
                    if "uploaded_file_content" not in st.session_state or \
                       st.session_state.get("uploaded_file_name") != uploaded_file.name:
                        with st.spinner("è§£æä¸­..."):
                            file_text = extract_text_from_file(uploaded_file)
                            st.session_state.uploaded_file_content = file_text
                            st.session_state.uploaded_file_name = uploaded_file.name
                    
                    # æ˜¾ç¤ºæ–‡ä»¶ä¿¡æ¯å’Œæ“ä½œ
                    st.caption(f"âœ… {uploaded_file.name}")
                    
                    # é¢„è§ˆå’Œæ¸…é™¤æŒ‰é’®æ”¾åœ¨ä¸€è¡Œ
                    btn_col1, btn_col2 = st.columns(2)
                    with btn_col1:
                        if st.button("ï¿½ï¸ é¢„è§ˆ", key="preview_gen", use_container_width=True):
                            st.session_state.show_preview_gen = not st.session_state.get("show_preview_gen", False)
                    with btn_col2:
                        if st.button("ğŸ—‘ï¸ æ¸…é™¤", key="clear_gen", use_container_width=True):
                            st.session_state.uploaded_file_content = ""
                            st.session_state.uploaded_file_name = ""
                            st.session_state.show_preview_gen = False
                            st.rerun()
                    
                    # é¢„è§ˆå†…å®¹
                    if st.session_state.get("show_preview_gen", False):
                        with st.expander("ğŸ“„ æ–‡ä»¶å†…å®¹é¢„è§ˆ", expanded=True):
                            preview_text = st.session_state.uploaded_file_content
                            if len(preview_text) > 500:
                                st.text(preview_text[:500] + "\n\n... [å·²æˆªæ–­] ...")
                            else:
                                st.text(preview_text)
                else:
                    # æ¸…é™¤ä¹‹å‰çš„æ–‡ä»¶å†…å®¹
                    if "uploaded_file_content" in st.session_state and st.session_state.uploaded_file_content:
                        pass  # ä¿ç•™å·²ä¸Šä¼ çš„å†…å®¹ï¼Œé™¤éç”¨æˆ·æ‰‹åŠ¨æ¸…é™¤
            
            with upload_col1:
                # æ˜¾ç¤ºé™„ä»¶çŠ¶æ€æç¤º
                if st.session_state.get("uploaded_file_content"):
                    st.info(f"ğŸ“ å·²æ·»åŠ é™„ä»¶: **{st.session_state.get('uploaded_file_name', 'æœªçŸ¥æ–‡ä»¶')}**")
        else:
            # æ¨¡å‹ä¸æ”¯æŒæ–‡ä»¶ä¸Šä¼ æ—¶æ˜¾ç¤ºæç¤º
            st.caption("ğŸ’¡ å½“å‰æ¨¡å‹ä¸æ”¯æŒæ–‡ä»¶ä¸Šä¼ ï¼Œå¦‚éœ€ä¸Šä¼ é™„ä»¶è¯·åˆ‡æ¢è‡³æ”¯æŒçš„æ¨¡å‹")
        
        # åˆå§‹åŒ–è‡ªæ£€ç»“æœçš„session_state
        if "generated_check_result" not in st.session_state:
            st.session_state.generated_check_result = ""
        
        # ä½¿ç”¨session_stateè·Ÿè¸ªå½“å‰å¤„ç†é˜¶æ®µ
        if "current_stage" not in st.session_state:
            st.session_state.current_stage = "idle"  # idle, generating, checking, done
        
        if st.button("ğŸš€ ç”Ÿæˆç­–åˆ’æ¡ˆ", type="primary", disabled=st.session_state.is_processing):
            if not user_input.strip():
                st.error("è¯·è¾“å…¥åŠŸèƒ½æè¿°ï¼")
            else:
                st.session_state.is_processing = True
                st.session_state.should_stop = False  # é‡ç½®ä¸­æ­¢æ ‡å¿—
                st.session_state.generated_check_result = ""  # æ¸…ç©ºä¹‹å‰çš„æ£€æŸ¥ç»“æœ
                st.session_state.generated_prd = ""  # æ¸…ç©ºä¹‹å‰çš„ç»“æœ
                st.session_state.last_error = ""  # æ¸…ç©ºé”™è¯¯
                st.session_state.current_stage = "generating"
                st.session_state.generate_saved_to_history = False  # é‡ç½®å†å²ä¿å­˜æ ‡è®°
                # ä¿å­˜ç”¨æˆ·è¾“å…¥å’Œé™„ä»¶å†…å®¹åˆ°session_state
                st.session_state.saved_user_input = user_input
                st.session_state.saved_attachment_content = st.session_state.get("uploaded_file_content", "")
                st.session_state.saved_attachment_name = st.session_state.get("uploaded_file_name", "")
                st.rerun()  # è§¦å‘é‡æ–°æ¸²æŸ“
        
        # å¤„ç†ç”Ÿæˆé˜¶æ®µ
        if st.session_state.is_processing and st.session_state.current_stage == "generating":
            # ä»session_stateè·å–ä¿å­˜çš„è¾“å…¥
            user_input_saved = st.session_state.get("saved_user_input", user_input)
            attachment_content = st.session_state.get("saved_attachment_content", "")
            attachment_name = st.session_state.get("saved_attachment_name", "")
            
            # æµå¼ç”Ÿæˆç­–åˆ’æ¡ˆ
            st.markdown("### ğŸ“„ ç”Ÿæˆçš„ç­–åˆ’æ¡ˆ")
            
            # æ˜¾ç¤ºä¸­æ­¢æŒ‰é’®å’ŒçŠ¶æ€
            col_status, col_stop = st.columns([4, 1])
            with col_status:
                st.markdown("**âœï¸ ç­–åˆ’é…¸å¥¶æ­£åœ¨æ’°å†™ç­–åˆ’æ¡ˆ...**")
            with col_stop:
                if st.button("â¹ï¸ ä¸­æ­¢ç”Ÿæˆ", key="stop_generate", type="secondary"):
                    st.session_state.should_stop = True
                    st.session_state.is_processing = False
                    st.session_state.current_stage = "idle"
                    st.warning("â¹ï¸ ç”Ÿæˆå·²ä¸­æ­¢")
                    st.rerun()
            
            # æ€è€ƒè¿‡ç¨‹å±•ç¤ºåŒºåŸŸï¼ˆå¯æŠ˜å ï¼‰
            thinking_expander = st.expander("ğŸ’­ æŸ¥çœ‹æ¨¡å‹æ€è€ƒè¿‡ç¨‹", expanded=False)
            with thinking_expander:
                thinking_container = st.empty()
            
            # çŠ¶æ€å’Œé”™è¯¯æ˜¾ç¤ºå®¹å™¨
            status_container = st.empty()
            
            # æ„å»ºæœ€ç»ˆçš„è¾“å…¥ï¼ˆåŒ…å«é™„ä»¶å†…å®¹ï¼‰
            final_input = user_input_saved
            if attachment_content:
                final_input = f"""ã€ç”¨æˆ·åŠŸèƒ½æè¿°ã€‘
{user_input_saved}

ã€é™„ä»¶å†…å®¹ã€‘ï¼ˆæ–‡ä»¶å: {attachment_name}ï¼‰
{attachment_content}

è¯·å‚è€ƒä»¥ä¸ŠåŠŸèƒ½æè¿°å’Œé™„ä»¶å†…å®¹ï¼Œç”Ÿæˆå®Œæ•´çš„ç­–åˆ’æ¡ˆã€‚"""
                st.info(f"ğŸ“ å·²åŒ…å«é™„ä»¶: {attachment_name}")
            
            # åˆ›å»ºå®¹å™¨ç”¨äºæµå¼æ˜¾ç¤º
            prd_container = st.empty()
            result, success, error = generate_prd(
                final_input, 
                use_stream=True, 
                container=prd_container,
                thinking_container=thinking_container,
                status_container=status_container
            )
            
            if success and result:
                st.session_state.generated_prd = result
                st.success("âœ… ç­–åˆ’æ¡ˆç”Ÿæˆå®Œæˆï¼")
                st.session_state.current_stage = "checking"
                st.rerun()  # è¿›å…¥ä¸‹ä¸€é˜¶æ®µ
            elif error:
                st.error(f"âŒ ç”Ÿæˆå¤±è´¥: {error}")
                st.session_state.is_processing = False
                st.session_state.current_stage = "idle"
            elif st.session_state.should_stop:
                st.warning("â¹ï¸ ç”Ÿæˆå·²ä¸­æ­¢")
                if result:  # å¦‚æœæœ‰éƒ¨åˆ†ç»“æœï¼Œä¿å­˜å®ƒ
                    st.session_state.generated_prd = result
                st.session_state.is_processing = False
                st.session_state.current_stage = "idle"
                st.session_state.should_stop = False
            else:
                st.error("ç”Ÿæˆå¤±è´¥ï¼Œè¯·é‡è¯•")
                st.session_state.is_processing = False
                st.session_state.current_stage = "idle"
        
        # å¤„ç†æ£€æŸ¥é˜¶æ®µ
        elif st.session_state.is_processing and st.session_state.current_stage == "checking":
            # æ˜¾ç¤ºå·²ç”Ÿæˆçš„ç­–åˆ’æ¡ˆï¼ˆæ ¼å¼åŒ–æ˜¾ç¤ºï¼‰
            render_prd_document(st.session_state.generated_prd, "ç”Ÿæˆçš„ç­–åˆ’æ¡ˆ")
            st.success("âœ… ç­–åˆ’æ¡ˆç”Ÿæˆå®Œæˆï¼")
            
            # AIè‡ªæ£€ - æµå¼è¾“å‡º
            st.markdown("### ğŸ” AIå¤æ£€æ¸…å•æ£€æŸ¥ç»“æœ")
            
            # æ˜¾ç¤ºä¸­æ­¢æŒ‰é’®å’ŒçŠ¶æ€
            col_status, col_stop = st.columns([4, 1])
            with col_status:
                st.markdown("**ğŸ” AIæ­£åœ¨è¿›è¡Œå¤æ£€æ¸…å•æ£€æŸ¥...**")
            with col_stop:
                if st.button("â¹ï¸ ä¸­æ­¢æ£€æŸ¥", key="stop_check", type="secondary"):
                    st.session_state.should_stop = True
                    st.session_state.is_processing = False
                    st.session_state.current_stage = "idle"
                    st.warning("â¹ï¸ æ£€æŸ¥å·²ä¸­æ­¢")
                    st.rerun()
            
            # æ€è€ƒè¿‡ç¨‹å±•ç¤ºåŒºåŸŸ
            thinking_expander = st.expander("ğŸ’­ æŸ¥çœ‹æ¨¡å‹æ€è€ƒè¿‡ç¨‹", expanded=False)
            with thinking_expander:
                thinking_container = st.empty()
            
            # çŠ¶æ€å’Œé”™è¯¯æ˜¾ç¤ºå®¹å™¨
            status_container = st.empty()
            
            check_container = st.empty()
            check_result, success, error = ai_self_check(
                st.session_state.generated_prd, 
                use_stream=True, 
                container=check_container,
                thinking_container=thinking_container,
                status_container=status_container
            )
            
            if success and check_result:
                st.session_state.generated_check_result = check_result
                st.success("âœ… å¤æ£€å®Œæˆï¼")
            elif error:
                st.error(f"âŒ å¤æ£€å¤±è´¥: {error}")
            
            st.session_state.is_processing = False
            st.session_state.current_stage = "done"
            st.session_state.should_stop = False
            st.rerun()  # åˆ·æ–°ä»¥æ˜¾ç¤ºæœ€ç»ˆç»“æœå’Œä¸‹è½½æŒ‰é’®
        
        # æ˜¾ç¤ºå·²ä¿å­˜çš„ç”Ÿæˆç»“æœï¼ˆéå¤„ç†ä¸­çŠ¶æ€ï¼‰
        if st.session_state.generated_prd and not st.session_state.is_processing:
            # ä½¿ç”¨æ ¼å¼åŒ–æ˜¾ç¤ºå‡½æ•°
            render_prd_document(st.session_state.generated_prd, "ç”Ÿæˆçš„ç­–åˆ’æ¡ˆ")
            
            # æ˜¾ç¤ºAIè‡ªæ£€ç»“æœ
            if st.session_state.generated_check_result:
                st.markdown("### ğŸ” AIå¤æ£€æ¸…å•æ£€æŸ¥ç»“æœ")
                with st.expander("æŸ¥çœ‹è¯¦ç»†æ£€æŸ¥ç»“æœ", expanded=True):
                    st.markdown(st.session_state.generated_check_result)
            
            st.markdown(CHECKLIST)
            
            # ä¸‹è½½æŒ‰é’® - Excelæ ¼å¼
            excel_data = create_excel_file(
                st.session_state.generated_prd,
                st.session_state.generated_check_result
            )
            
            st.download_button(
                label="ğŸ“¥ ä¸‹è½½ç­–åˆ’æ¡ˆ (Excel)",
                data=excel_data,
                file_name="ç­–åˆ’æ¡ˆ.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # ä¿å­˜åˆ°ä¼šè¯å†å²ï¼ˆä»…åœ¨é¦–æ¬¡å®Œæˆæ—¶ä¿å­˜ï¼Œé¿å…é‡å¤ï¼‰
            if st.session_state.get("current_stage") == "done" and not st.session_state.get("generate_saved_to_history"):
                add_to_history(
                    function_type="ç”Ÿæˆç­–åˆ’æ¡ˆ",
                    input_data={"åŠŸèƒ½æè¿°": st.session_state.get("saved_user_input", "")},
                    output_data=st.session_state.generated_prd,
                    download_data=excel_data,
                    download_filename="ç­–åˆ’æ¡ˆ.xlsx",
                    download_mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.session_state.generate_saved_to_history = True
            
            # ========== å¤šè½®å¯¹è¯åŒºåŸŸ ==========
            st.markdown("---")
            st.markdown("### ğŸ’¬ ç»§ç»­å¯¹è¯")
            st.caption("æ‚¨å¯ä»¥ç»§ç»­è¿½é—®æˆ–è¦æ±‚ä¿®æ”¹ï¼ŒAIå°†åŸºäºå·²ç”Ÿæˆçš„ç­–åˆ’æ¡ˆè¿›è¡Œå›ç­”ã€‚")
            
            # åˆå§‹åŒ–å¯¹è¯å†å²
            chat_key = "generate_prd_chat"
            init_chat_history(chat_key)
            
            # æ˜¾ç¤ºå¯¹è¯å†å² - ä½¿ç”¨ ChatGPT é£æ ¼çš„å¯¹è¯æ°”æ³¡
            chat_history = get_chat_history(chat_key)
            if chat_history:
                for msg in chat_history:
                    if msg["role"] == "user":
                        with st.chat_message("user"):
                            st.markdown(msg["content"])
                    else:
                        with st.chat_message("assistant", avatar="ğŸ“"):
                            st.markdown(msg["content"])
            
            # å¯¹è¯è¾“å…¥ - ä½¿ç”¨ chat_input
            chat_input = st.chat_input(
                placeholder="ä¾‹å¦‚ï¼šè¯·è¯¦ç»†è¯´æ˜ç¬¬3ç« çš„éªŒæ”¶æ ‡å‡†...",
                key="generate_chat_input"
            )
            
            # æ¸…ç©ºæŒ‰é’®æ”¾åœ¨å•ç‹¬ä¸€è¡Œ
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºå¯¹è¯å†å²", key="generate_chat_clear", use_container_width=False):
                clear_chat_history(chat_key)
                st.rerun()
            
            # å¤„ç†å¯¹è¯
            if chat_input and chat_input.strip():
                add_chat_message(chat_key, "user", chat_input)
                
                # æ„å»ºä¸Šä¸‹æ–‡
                function_context = f"""ã€å·²ç”Ÿæˆçš„ç­–åˆ’æ¡ˆã€‘
{st.session_state.generated_prd}"""
                
                history_context = build_chat_context(chat_key, get_system_prompt_with_date(GENERATE_PRD_SYSTEM_PROMPT))
                full_prompt = f"""{function_context}

{history_context}

ã€å½“å‰ç”¨æˆ·è¾“å…¥ã€‘
{chat_input}

è¯·åŸºäºä»¥ä¸Šç­–åˆ’æ¡ˆå’Œå¯¹è¯å†å²ï¼Œå›ç­”ç”¨æˆ·çš„é—®é¢˜æˆ–æŒ‰è¦æ±‚è¿›è¡Œä¿®æ”¹ã€‚å¦‚æœç”¨æˆ·è¦æ±‚ä¿®æ”¹ç­–åˆ’æ¡ˆï¼Œè¯·è¾“å‡ºä¿®æ”¹åçš„å®Œæ•´å†…å®¹ã€‚"""
                
                with st.spinner("æ­£åœ¨æ€è€ƒ..."):
                    response_container = st.empty()
                    full_response = ""
                    for chunk in call_gemini_stream(full_prompt, get_system_prompt_with_date(GENERATE_PRD_SYSTEM_PROMPT)):
                        if chunk["type"] == "text":
                            full_response += chunk["content"]
                            response_container.markdown(full_response + "â–Œ")
                        elif chunk["type"] == "error":
                            st.error(f"ç”Ÿæˆå¤±è´¥: {chunk['content']}")
                            break
                    
                    if full_response:
                        response_container.markdown(full_response)
                        add_chat_message(chat_key, "assistant", full_response)
                        st.rerun()
    
    elif function_mode == "è„‘å›¾ç”Ÿæˆç­–åˆ’æ¡ˆ":
        # æ ‡é¢˜å’Œæ¸…ç©ºæŒ‰é’®
        title_col, clear_col = st.columns([6, 1])
        with title_col:
            st.markdown("### ğŸ§  è„‘å›¾ç”Ÿæˆç­–åˆ’æ¡ˆ")
        with clear_col:
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºä¼šè¯", key="clear_mindmap_session", use_container_width=True):
                clear_module_session("è„‘å›¾ç”Ÿæˆç­–åˆ’æ¡ˆ")
                st.rerun()
        st.markdown("ä¸Šä¼ æ€ç»´è„‘å›¾å›¾ç‰‡æˆ–è¾“å…¥Mermaidä»£ç ï¼ŒAIå°†è¯†åˆ«ç»“æ„å¹¶ç”Ÿæˆå®Œæ•´çš„ç­–åˆ’æ¡ˆã€‚")
        
        # åˆå§‹åŒ–è„‘å›¾ç›¸å…³çš„session state
        if "mindmap_parsed_structure" not in st.session_state:
            st.session_state.mindmap_parsed_structure = None
        if "mindmap_generated_prd" not in st.session_state:
            st.session_state.mindmap_generated_prd = None
        if "mindmap_image_data" not in st.session_state:
            st.session_state.mindmap_image_data = None
        if "mindmap_saved" not in st.session_state:
            st.session_state.mindmap_saved = False
        if "mindmap_mermaid_code" not in st.session_state:
            st.session_state.mindmap_mermaid_code = ""
        if "mindmap_input_mode" not in st.session_state:
            st.session_state.mindmap_input_mode = "å›¾ç‰‡ä¸Šä¼ "
        
        # æ£€æŸ¥æ˜¯å¦æœ‰ä»linmoæ¨¡å—ä¼ å…¥çš„mermaidä»£ç 
        if "linmo_to_mindmap_mermaid" in st.session_state and st.session_state.linmo_to_mindmap_mermaid:
            st.session_state.mindmap_mermaid_code = st.session_state.linmo_to_mindmap_mermaid
            st.session_state.mindmap_input_mode = "Mermaidä»£ç "
            st.session_state.linmo_to_mindmap_mermaid = ""  # æ¸…ç©ºä¼ å…¥æ•°æ®
            st.success("âœ… å·²ä»æ€è·¯å¼•å¯¼åŠ©æ‰‹å¯¼å…¥Mermaidä»£ç ï¼")
        
        # è¾“å…¥æ¨¡å¼é€‰æ‹©
        input_mode = st.radio(
            "é€‰æ‹©è¾“å…¥æ–¹å¼",
            ["å›¾ç‰‡ä¸Šä¼ ", "Mermaidä»£ç "],
            index=0 if st.session_state.mindmap_input_mode == "å›¾ç‰‡ä¸Šä¼ " else 1,
            horizontal=True,
            key="mindmap_input_mode_radio"
        )
        st.session_state.mindmap_input_mode = input_mode
        
        st.markdown("---")
        
        if input_mode == "å›¾ç‰‡ä¸Šä¼ ":
            # æ–‡ä»¶ä¸Šä¼ åŒºåŸŸ
            uploaded_mindmap = st.file_uploader(
                "ğŸ“¤ ä¸Šä¼ æ€ç»´è„‘å›¾",
                type=["jpg", "jpeg", "png", "pdf"],
                help="æ”¯æŒ JPGã€PNG æ ¼å¼çš„å›¾ç‰‡æˆ– PDF æ–‡ä»¶",
                key="mindmap_uploader"
            )
            
            # æ˜¾ç¤ºä¸Šä¼ çš„å›¾ç‰‡é¢„è§ˆ
            if uploaded_mindmap:
                file_type = uploaded_mindmap.type
                file_data = uploaded_mindmap.read()
                
                # å›¾ç‰‡é¢„è§ˆ
                if file_type in ["image/jpeg", "image/png"]:
                    st.image(file_data, caption="ä¸Šä¼ çš„æ€ç»´è„‘å›¾", use_container_width=True)
                elif file_type == "application/pdf":
                    st.info("ğŸ“„ å·²ä¸Šä¼  PDF æ–‡ä»¶ï¼ŒAIå°†å°è¯•è§£æå…¶ä¸­çš„æ€ç»´è„‘å›¾å†…å®¹")
                
                # ä¿å­˜å›¾ç‰‡æ•°æ®åˆ°session state
                st.session_state.mindmap_image_data = {
                    "data": file_data,
                    "mime_type": file_type,
                    "name": uploaded_mindmap.name
                }
        else:
            # Mermaidä»£ç è¾“å…¥åŒºåŸŸ
            st.markdown("#### ğŸ“ è¾“å…¥Mermaidä»£ç ")
            st.markdown("*æ”¯æŒä»æ€è·¯å¼•å¯¼åŠ©æ‰‹ç”Ÿæˆçš„Mermaidæ€ç»´å¯¼å›¾ä»£ç *")
            
            mermaid_code = st.text_area(
                "Mermaidä»£ç ",
                value=st.session_state.mindmap_mermaid_code,
                height=250,
                placeholder="""è¯·è¾“å…¥Mermaidæ ¼å¼çš„æ€ç»´å¯¼å›¾ä»£ç ï¼Œä¾‹å¦‚ï¼š
graph LR
    A[æ ¸å¿ƒåŠŸèƒ½] --> B(å­åŠŸèƒ½1)
    A --> C(å­åŠŸèƒ½2)
    B --> B1[å…·ä½“ç»†èŠ‚]
    C --> C1[å…·ä½“ç»†èŠ‚]""",
                key="mindmap_mermaid_input"
            )
            st.session_state.mindmap_mermaid_code = mermaid_code
            
            # æ˜¾ç¤ºMermaidä»£ç é¢„è§ˆæç¤º
            if mermaid_code.strip():
                st.info("ğŸ’¡ æç¤ºï¼šæ‚¨å¯ä»¥å°†æ­¤ä»£ç å¤åˆ¶åˆ° [Mermaid Live Editor](https://mermaid-live.nodejs.cn/edit) é¢„è§ˆæ•ˆæœ")
        
        # è¡¥å……è¯´æ˜è¾“å…¥
        additional_info = st.text_area(
            "ğŸ“ è¡¥å……è¯´æ˜ï¼ˆå¯é€‰ï¼‰",
            height=100,
            placeholder="å¦‚æœ‰å…¶ä»–éœ€æ±‚æˆ–èƒŒæ™¯ä¿¡æ¯ï¼Œè¯·åœ¨æ­¤è¾“å…¥...\nä¾‹å¦‚ï¼šè¿™æ˜¯ä¸€ä¸ªMMORPGæ¸¸æˆçš„ç¤¾äº¤ç³»ç»Ÿè®¾è®¡",
            key="mindmap_additional_info"
        )
        
        # æ“ä½œæŒ‰é’® - æ ¹æ®è¾“å…¥æ¨¡å¼è°ƒæ•´
        if input_mode == "å›¾ç‰‡ä¸Šä¼ ":
            col1, col2, col3 = st.columns([1, 1, 1])
            
            with col1:
                parse_btn = st.button(
                    "ğŸ” è§£æè„‘å›¾ç»“æ„",
                    disabled=not st.session_state.mindmap_image_data,
                    use_container_width=True
                )
            
            with col2:
                generate_btn = st.button(
                    "ğŸ“ ç”Ÿæˆç­–åˆ’æ¡ˆ",
                    disabled=not st.session_state.mindmap_parsed_structure,
                    use_container_width=True
                )
            
            with col3:
                clear_btn = st.button(
                    "ğŸ—‘ï¸ æ¸…ç©ºé‡æ¥",
                    use_container_width=True
                )
            
            # Mermaidæ¨¡å¼ä¸éœ€è¦è§£ææŒ‰é’®
            mermaid_parse_btn = False
        else:
            col1, col2, col3 = st.columns([1, 1, 1])
            
            with col1:
                mermaid_parse_btn = st.button(
                    "ğŸ” è§£æMermaidç»“æ„",
                    disabled=not st.session_state.mindmap_mermaid_code.strip(),
                    use_container_width=True
                )
                parse_btn = False
            
            with col2:
                generate_btn = st.button(
                    "ğŸ“ ç”Ÿæˆç­–åˆ’æ¡ˆ",
                    disabled=not st.session_state.mindmap_parsed_structure,
                    use_container_width=True
                )
            
            with col3:
                clear_btn = st.button(
                    "ğŸ—‘ï¸ æ¸…ç©ºé‡æ¥",
                    use_container_width=True
                )
        
        if clear_btn:
            st.session_state.mindmap_parsed_structure = None
            st.session_state.mindmap_generated_prd = None
            st.session_state.mindmap_image_data = None
            st.session_state.mindmap_mermaid_code = ""
            st.session_state.mindmap_saved = False
            st.rerun()
        
        # è§£æMermaidä»£ç ç»“æ„
        if mermaid_parse_btn and st.session_state.mindmap_mermaid_code.strip():
            st.markdown("---")
            st.markdown("#### ğŸ”„ æ­£åœ¨è§£æMermaidä»£ç ç»“æ„...")
            
            # åˆ›å»ºæ˜¾ç¤ºå®¹å™¨
            thinking_container = st.expander("ğŸ’­ AIæ€è€ƒè¿‡ç¨‹", expanded=False)
            status_container = st.empty()
            result_container = st.empty()
            
            mermaid_parse_prompt = f"""è¯·åˆ†æä»¥ä¸‹Mermaidæ ¼å¼çš„æ€ç»´å¯¼å›¾ä»£ç ï¼Œå°†å…¶è½¬æ¢ä¸ºç»“æ„åŒ–çš„æ–‡æœ¬æ ¼å¼ï¼Œä¾¿äºç”Ÿæˆç­–åˆ’æ¡ˆã€‚

ã€Mermaidä»£ç ã€‘
```mermaid
{st.session_state.mindmap_mermaid_code}
```

è¯·è¯†åˆ«å‡ºï¼š
1. æ ¸å¿ƒä¸»é¢˜/åŠŸèƒ½
2. å„ä¸ªåˆ†æ”¯èŠ‚ç‚¹åŠå…¶å±‚çº§å…³ç³»
3. èŠ‚ç‚¹ä¹‹é—´çš„é€»è¾‘å…³ç³»

è¾“å‡ºæ ¼å¼è¦æ±‚ï¼šä½¿ç”¨å±‚çº§ç¼©è¿›çš„æ–‡æœ¬å½¢å¼å±•ç¤ºç»“æ„ã€‚"""

            if additional_info:
                mermaid_parse_prompt += f"\n\nè¡¥å……èƒŒæ™¯ä¿¡æ¯ï¼š{additional_info}"
            
            # æµå¼è§£æ
            full_response = ""
            thinking_text = ""
            
            for chunk_data in call_gemini_stream(mermaid_parse_prompt, MINDMAP_PARSE_SYSTEM_PROMPT):
                chunk_type = chunk_data.get("type", "text")
                chunk_content = chunk_data.get("content", "")
                
                if chunk_type == "text":
                    full_response += chunk_content
                    result_container.markdown(full_response + " â–Œ")
                elif chunk_type == "thinking":
                    thinking_text += chunk_content
                    with thinking_container:
                        st.markdown(thinking_text)
                elif chunk_type == "error":
                    status_container.error(f"âŒ è§£æå¤±è´¥: {chunk_content}")
            
            if full_response:
                result_container.markdown(full_response)
                st.session_state.mindmap_parsed_structure = full_response
                status_container.success('âœ… Mermaidç»“æ„è§£æå®Œæˆï¼è¯·ç‚¹å‡»"ç”Ÿæˆç­–åˆ’æ¡ˆ"æŒ‰é’®ç»§ç»­ã€‚')
                st.rerun()
        
        # è§£æè„‘å›¾ç»“æ„
        if parse_btn and st.session_state.mindmap_image_data:
            st.markdown("---")
            st.markdown("#### ğŸ”„ æ­£åœ¨è§£ææ€ç»´è„‘å›¾...")
            
            image_info = st.session_state.mindmap_image_data
            
            # åˆ›å»ºæ˜¾ç¤ºå®¹å™¨
            thinking_container = st.expander("ğŸ’­ AIæ€è€ƒè¿‡ç¨‹", expanded=False)
            status_container = st.empty()
            result_container = st.empty()
            
            parse_prompt = "è¯·ä»”ç»†åˆ†æè¿™å¼ æ€ç»´è„‘å›¾å›¾ç‰‡ï¼Œè¯†åˆ«å‡ºæ‰€æœ‰çš„èŠ‚ç‚¹ã€å±‚çº§å…³ç³»å’Œè¿æ¥ï¼Œå°†å…¶è½¬æ¢ä¸ºç»“æ„åŒ–çš„æ–‡æœ¬æ ¼å¼ã€‚"
            
            if additional_info:
                parse_prompt += f"\n\nè¡¥å……èƒŒæ™¯ä¿¡æ¯ï¼š{additional_info}"
            
            # æµå¼è§£æ
            full_response = ""
            thinking_text = ""
            
            for chunk_data in call_gemini_with_image_stream(
                image_info["data"],
                parse_prompt,
                MINDMAP_PARSE_SYSTEM_PROMPT,
                image_info["mime_type"],
                thinking_container
            ):
                chunk_type = chunk_data.get("type", "text")
                chunk_content = chunk_data.get("content", "")
                
                if chunk_type == "text":
                    full_response += chunk_content
                    result_container.markdown(full_response + " â–Œ")
                elif chunk_type == "thinking":
                    thinking_text += chunk_content
                    with thinking_container:
                        st.markdown(thinking_text)
                elif chunk_type == "retry":
                    status_container.warning(chunk_content)
                elif chunk_type == "error":
                    status_container.error(f"âŒ è§£æå¤±è´¥: {chunk_content}")
                elif chunk_type == "stopped":
                    status_container.warning("âš ï¸ ç”¨æˆ·å·²ä¸­æ­¢")
            
            if full_response:
                result_container.markdown(full_response)
                st.session_state.mindmap_parsed_structure = full_response
                status_container.success('âœ… è„‘å›¾ç»“æ„è§£æå®Œæˆï¼è¯·ç‚¹å‡»"ç”Ÿæˆç­–åˆ’æ¡ˆ"æŒ‰é’®ç»§ç»­ã€‚')
                st.rerun()
        
        # æ˜¾ç¤ºå·²è§£æçš„ç»“æ„
        if st.session_state.mindmap_parsed_structure:
            st.markdown("---")
            st.markdown("#### ğŸ“‹ è§£æå‡ºçš„è„‘å›¾ç»“æ„")
            with st.expander("æŸ¥çœ‹/ç¼–è¾‘è§£æç»“æœ", expanded=True):
                edited_structure = st.text_area(
                    "è§£æç»“æœï¼ˆå¯æ‰‹åŠ¨ç¼–è¾‘ä¿®æ­£ï¼‰",
                    value=st.session_state.mindmap_parsed_structure,
                    height=300,
                    key="mindmap_structure_editor"
                )
                if edited_structure != st.session_state.mindmap_parsed_structure:
                    st.session_state.mindmap_parsed_structure = edited_structure
        
        # ç”Ÿæˆç­–åˆ’æ¡ˆ
        if generate_btn and st.session_state.mindmap_parsed_structure:
            st.markdown("---")
            st.markdown("#### ğŸ”„ æ­£åœ¨ç”Ÿæˆç­–åˆ’æ¡ˆ...")
            
            # åˆ›å»ºæ˜¾ç¤ºå®¹å™¨
            thinking_container = st.expander("ğŸ’­ AIæ€è€ƒè¿‡ç¨‹", expanded=False)
            status_container = st.empty()
            result_container = st.empty()
            
            generate_prompt = f"""è¯·æ ¹æ®ä»¥ä¸‹æ€ç»´è„‘å›¾ç»“æ„ç”Ÿæˆå®Œæ•´çš„ç­–åˆ’æ¡ˆï¼š

ã€æ€ç»´è„‘å›¾ç»“æ„ã€‘
{st.session_state.mindmap_parsed_structure}
"""
            
            if additional_info:
                generate_prompt += f"\nã€è¡¥å……è¯´æ˜ã€‘\n{additional_info}"
            
            # æµå¼ç”Ÿæˆ
            full_response = ""
            thinking_text = ""
            
            for chunk_data in call_gemini_stream(generate_prompt, get_system_prompt_with_date(MINDMAP_TO_PRD_SYSTEM_PROMPT), thinking_container):
                chunk_type = chunk_data.get("type", "text")
                chunk_content = chunk_data.get("content", "")
                
                if chunk_type == "text":
                    full_response += chunk_content
                    result_container.markdown(full_response + " â–Œ")
                elif chunk_type == "thinking":
                    thinking_text += chunk_content
                    with thinking_container:
                        st.markdown(thinking_text)
                elif chunk_type == "retry":
                    status_container.warning(chunk_content)
                elif chunk_type == "error":
                    status_container.error(f"âŒ ç”Ÿæˆå¤±è´¥: {chunk_content}")
                elif chunk_type == "stopped":
                    status_container.warning("âš ï¸ ç”¨æˆ·å·²ä¸­æ­¢")
            
            if full_response:
                result_container.empty()
                st.session_state.mindmap_generated_prd = full_response
                st.session_state.mindmap_saved = False
                status_container.success("âœ… ç­–åˆ’æ¡ˆç”Ÿæˆå®Œæˆï¼")
                st.rerun()
        
        # æ˜¾ç¤ºç”Ÿæˆçš„ç­–åˆ’æ¡ˆ
        if st.session_state.mindmap_generated_prd:
            st.markdown("---")
            render_prd_document(st.session_state.mindmap_generated_prd, "ç”Ÿæˆçš„ç­–åˆ’æ¡ˆï¼ˆåŸºäºæ€ç»´è„‘å›¾ï¼‰")
            
            # ä¿å­˜åˆ°å†å²è®°å½•
            if not st.session_state.mindmap_saved:
                mindmap_name = st.session_state.mindmap_image_data.get("name", "æ€ç»´è„‘å›¾") if st.session_state.mindmap_image_data else "æ€ç»´è„‘å›¾"
                excel_data = create_excel_file(st.session_state.mindmap_generated_prd)
                add_to_history(
                    function_type="è„‘å›¾ç”Ÿæˆç­–åˆ’æ¡ˆ",
                    input_data={
                        "è„‘å›¾æ–‡ä»¶": mindmap_name,
                        "è§£æç»“æ„": st.session_state.mindmap_parsed_structure[:200] + "..." if len(st.session_state.mindmap_parsed_structure) > 200 else st.session_state.mindmap_parsed_structure,
                        "è¡¥å……è¯´æ˜": additional_info if additional_info else "æ— "
                    },
                    output_data=st.session_state.mindmap_generated_prd,
                    download_data=excel_data,
                    download_filename=f"è„‘å›¾ç­–åˆ’æ¡ˆ_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    download_mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.session_state.mindmap_saved = True
            
            # ä¸‹è½½æŒ‰é’®
            col1, col2 = st.columns(2)
            with col1:
                st.download_button(
                    label="ğŸ“¥ ä¸‹è½½ç­–åˆ’æ¡ˆ (Excel)",
                    data=create_excel_file(st.session_state.mindmap_generated_prd),
                    file_name=f"è„‘å›¾ç­–åˆ’æ¡ˆ_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            with col2:
                st.download_button(
                    label="ğŸ“¥ ä¸‹è½½ç­–åˆ’æ¡ˆ (Markdown)",
                    data=st.session_state.mindmap_generated_prd,
                    file_name=f"è„‘å›¾ç­–åˆ’æ¡ˆ_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md",
                    mime="text/markdown",
                    use_container_width=True
                )
            
            # å¤šè½®å¯¹è¯åŒºåŸŸ
            st.markdown("---")
            st.markdown("#### ğŸ’¬ ç»§ç»­å¯¹è¯")
            
            chat_key = "mindmap_prd_chat"
            init_chat_history(chat_key)
            
            # æ˜¾ç¤ºå¯¹è¯å†å² - ä½¿ç”¨ ChatGPT é£æ ¼çš„å¯¹è¯æ°”æ³¡
            chat_history = get_chat_history(chat_key)
            if chat_history:
                for msg in chat_history:
                    if msg["role"] == "user":
                        with st.chat_message("user"):
                            st.markdown(msg["content"])
                    else:
                        with st.chat_message("assistant", avatar="ğŸ—ºï¸"):
                            st.markdown(msg["content"])
            
            # å¯¹è¯è¾“å…¥ - ä½¿ç”¨ chat_input
            chat_input = st.chat_input(
                placeholder="ä¾‹å¦‚ï¼šè¯·è¡¥å……ä¸€ä¸‹æŠ€æœ¯å®ç°æ–¹æ¡ˆ...",
                key="mindmap_chat_input"
            )
            
            # æ¸…ç©ºæŒ‰é’®
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºå¯¹è¯å†å²", key="mindmap_clear_chat", use_container_width=False):
                clear_chat_history(chat_key)
                st.rerun()
            
            if chat_input and chat_input.strip():
                add_chat_message(chat_key, "user", chat_input)
                
                # æ„å»ºä¸Šä¸‹æ–‡
                context_prompt = f"""å½“å‰ç­–åˆ’æ¡ˆå†…å®¹ï¼š

{st.session_state.mindmap_generated_prd}

ç”¨æˆ·è¿½é—®ï¼š{chat_input}

è¯·æ ¹æ®ç­–åˆ’æ¡ˆå†…å®¹å›ç­”ç”¨æˆ·çš„é—®é¢˜æˆ–è¿›è¡Œç›¸åº”ä¿®æ”¹ã€‚"""
                
                history_context = build_chat_context(chat_key, get_system_prompt_with_date(MINDMAP_TO_PRD_SYSTEM_PROMPT))
                full_prompt = history_context + "\n\n" + context_prompt
                
                response_container = st.empty()
                full_response = ""
                
                for chunk_data in call_gemini_stream(full_prompt, get_system_prompt_with_date(MINDMAP_TO_PRD_SYSTEM_PROMPT)):
                    chunk_type = chunk_data.get("type", "text")
                    chunk_content = chunk_data.get("content", "")
                    
                    if chunk_type == "text":
                        full_response += chunk_content
                        response_container.markdown(f"**ğŸ¤– AIï¼š** {full_response} â–Œ")
                
                if full_response:
                    response_container.markdown(f"**ğŸ¤– AIï¼š** {full_response}")
                    add_chat_message(chat_key, "assistant", full_response)
                    st.rerun()
    
    elif function_mode == "ä¼˜åŒ–ç­–åˆ’æ¡ˆ":
        # æ ‡é¢˜å’Œæ¸…ç©ºæŒ‰é’®
        title_col, clear_col = st.columns([6, 1])
        with title_col:
            st.markdown("### ğŸ”„ ä¼˜åŒ–ç°æœ‰ç­–åˆ’æ¡ˆ")
        with clear_col:
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºä¼šè¯", key="clear_optimize_session", use_container_width=True):
                clear_module_session("ä¼˜åŒ–ç­–åˆ’æ¡ˆ")
                st.rerun()
        st.markdown("è¯·è¾“å…¥åŸç­–åˆ’æ¡ˆå’Œä¿®æ”¹æ„è§ï¼ŒAIå°†é€šè¿‡å¤šè½®è¿­ä»£è¿›è¡Œä¼˜åŒ–ã€‚")
        
        col1, col2 = st.columns([3, 1])
        
        with col1:
            old_prd = st.text_area(
                "åŸç­–åˆ’æ¡ˆ",
                height=400,
                placeholder="è¯·ç²˜è´´éœ€è¦ä¼˜åŒ–çš„ç­–åˆ’æ¡ˆå†…å®¹...",
                key="optimize_input"
            )
            
            # ========== æ–‡ä»¶ä¸Šä¼ åŒºåŸŸï¼ˆè¾“å…¥æ¡†å³ä¸‹æ–¹ï¼‰==========
            if is_file_upload_supported():
                # åˆ›å»ºå¸ƒå±€ï¼šå·¦è¾¹æ˜¯çŠ¶æ€æç¤ºï¼Œå³è¾¹æ˜¯æ–‡ä»¶ä¸Šä¼ 
                opt_upload_col1, opt_upload_col2 = st.columns([2, 1])
                
                with opt_upload_col2:
                    uploaded_file_opt = st.file_uploader(
                        "ğŸ“ ä¸Šä¼ é™„ä»¶",
                        type=SUPPORTED_FILE_TYPES,
                        help="ä¸Šä¼ å‚è€ƒæ–‡æ¡£ä¾›AIå‚è€ƒï¼ˆPDF/Word/TXT/MDï¼‰",
                        key="optimize_file_uploader"
                    )
                    
                    # å¤„ç†ä¸Šä¼ çš„æ–‡ä»¶
                    if uploaded_file_opt is not None:
                        if "uploaded_file_content" not in st.session_state or \
                           st.session_state.get("uploaded_file_name") != uploaded_file_opt.name:
                            with st.spinner("è§£æä¸­..."):
                                file_text = extract_text_from_file(uploaded_file_opt)
                                st.session_state.uploaded_file_content = file_text
                                st.session_state.uploaded_file_name = uploaded_file_opt.name
                        
                        # æ˜¾ç¤ºæ–‡ä»¶ä¿¡æ¯å’Œæ“ä½œ
                        st.caption(f"âœ… {uploaded_file_opt.name}")
                        
                        # é¢„è§ˆå’Œæ¸…é™¤æŒ‰é’®æ”¾åœ¨ä¸€è¡Œ
                        opt_btn_col1, opt_btn_col2 = st.columns(2)
                        with opt_btn_col1:
                            if st.button("ğŸ‘ï¸ é¢„è§ˆ", key="preview_opt", use_container_width=True):
                                st.session_state.show_preview_opt = not st.session_state.get("show_preview_opt", False)
                        with opt_btn_col2:
                            if st.button("ğŸ—‘ï¸ æ¸…é™¤", key="clear_opt", use_container_width=True):
                                st.session_state.uploaded_file_content = ""
                                st.session_state.uploaded_file_name = ""
                                st.session_state.show_preview_opt = False
                                st.rerun()
                        
                        # é¢„è§ˆå†…å®¹
                        if st.session_state.get("show_preview_opt", False):
                            with st.expander("ğŸ“„ æ–‡ä»¶å†…å®¹é¢„è§ˆ", expanded=True):
                                preview_text = st.session_state.uploaded_file_content
                                if len(preview_text) > 500:
                                    st.text(preview_text[:500] + "\n\n... [å·²æˆªæ–­] ...")
                                else:
                                    st.text(preview_text)
                
                with opt_upload_col1:
                    # æ˜¾ç¤ºé™„ä»¶çŠ¶æ€æç¤º
                    if st.session_state.get("uploaded_file_content"):
                        st.info(f"ğŸ“ å·²æ·»åŠ é™„ä»¶: **{st.session_state.get('uploaded_file_name', 'æœªçŸ¥æ–‡ä»¶')}**")
            else:
                # æ¨¡å‹ä¸æ”¯æŒæ–‡ä»¶ä¸Šä¼ æ—¶æ˜¾ç¤ºæç¤º
                st.caption("ğŸ’¡ å½“å‰æ¨¡å‹ä¸æ”¯æŒæ–‡ä»¶ä¸Šä¼ ï¼Œå¦‚éœ€ä¸Šä¼ é™„ä»¶è¯·åˆ‡æ¢è‡³æ”¯æŒçš„æ¨¡å‹")
        
        with col2:
            feedback = st.text_area(
                "ä¿®æ”¹æ„è§ï¼ˆå¯é€‰ï¼‰",
                height=200,
                placeholder="è¯·è¾“å…¥æ‚¨çš„ä¿®æ”¹æ„è§æˆ–å…³æ³¨ç‚¹...",
                key="feedback_input"
            )
            
            max_iterations = st.number_input(
                "è¿­ä»£è½®æ¬¡",
                min_value=1,
                max_value=10,
                value=3,
                help="è®¾ç½®Reflectionå¾ªç¯çš„è¿­ä»£æ¬¡æ•°ï¼ˆ1-10è½®ï¼‰"
            )
        
        # ä½¿ç”¨session_stateè·Ÿè¸ªä¼˜åŒ–é˜¶æ®µ
        if "optimize_stage" not in st.session_state:
            st.session_state.optimize_stage = "idle"  # idle, initial, reflection, checking, done
        if "initial_fixed_prd" not in st.session_state:
            st.session_state.initial_fixed_prd = ""
        if "saved_old_prd" not in st.session_state:
            st.session_state.saved_old_prd = ""
        if "saved_feedback" not in st.session_state:
            st.session_state.saved_feedback = ""
        if "saved_max_iterations" not in st.session_state:
            st.session_state.saved_max_iterations = 3
        
        if st.button("ğŸ”„ å¼€å§‹ä¼˜åŒ–", type="primary", disabled=st.session_state.is_processing):
            if not old_prd.strip():
                st.error("è¯·è¾“å…¥åŸç­–åˆ’æ¡ˆï¼")
            else:
                st.session_state.is_processing = True
                st.session_state.should_stop = False  # é‡ç½®ä¸­æ­¢æ ‡å¿—
                st.session_state.last_error = ""  # æ¸…ç©ºé”™è¯¯
                st.session_state.optimized_prd = ""
                st.session_state.optimized_check_result = ""
                st.session_state.initial_fixed_prd = ""
                st.session_state.saved_old_prd = old_prd
                st.session_state.saved_feedback = feedback
                st.session_state.saved_max_iterations = max_iterations
                st.session_state.optimize_saved_to_history = False  # é‡ç½®å†å²ä¿å­˜æ ‡è®°
                # ä¿å­˜é™„ä»¶å†…å®¹
                st.session_state.saved_optimize_attachment = st.session_state.get("uploaded_file_content", "")
                st.session_state.saved_optimize_attachment_name = st.session_state.get("uploaded_file_name", "")
                st.session_state.optimize_stage = "initial"
                st.rerun()  # è§¦å‘é‡æ–°æ¸²æŸ“
        
        # å¤„ç†åˆå§‹ä¿®æ­£é˜¶æ®µ
        if st.session_state.is_processing and st.session_state.optimize_stage == "initial":
            st.markdown("### ğŸ“Œ Step 1: åˆå§‹ä¿®æ­£")
            
            # æ˜¾ç¤ºé™„ä»¶ä½¿ç”¨ä¿¡æ¯
            optimize_attachment = st.session_state.get("saved_optimize_attachment", "")
            optimize_attachment_name = st.session_state.get("saved_optimize_attachment_name", "")
            if optimize_attachment:
                st.info(f"ğŸ“ å‚è€ƒé™„ä»¶: {optimize_attachment_name}")
            
            # æ˜¾ç¤ºä¸­æ­¢æŒ‰é’®å’ŒçŠ¶æ€
            col_status, col_stop = st.columns([4, 1])
            with col_status:
                st.markdown("**âœï¸ æ­£åœ¨è¿›è¡Œåˆå§‹ä¿®æ­£...**")
            with col_stop:
                if st.button("â¹ï¸ ä¸­æ­¢", key="stop_initial", type="secondary"):
                    st.session_state.should_stop = True
                    st.session_state.optimize_processing = False
                    st.session_state.optimize_stage = "idle"
                    st.warning("â¹ï¸ ä¼˜åŒ–å·²ä¸­æ­¢")
                    st.rerun()
            
            # æ€è€ƒè¿‡ç¨‹å±•ç¤ºåŒºåŸŸ
            thinking_expander = st.expander("ğŸ’­ æŸ¥çœ‹æ¨¡å‹æ€è€ƒè¿‡ç¨‹", expanded=False)
            with thinking_expander:
                thinking_container = st.empty()
            
            # çŠ¶æ€å’Œé”™è¯¯æ˜¾ç¤ºå®¹å™¨
            status_container = st.empty()
            
            # æ„å»ºåŒ…å«é™„ä»¶çš„feedback
            final_feedback = st.session_state.saved_feedback
            if optimize_attachment:
                final_feedback = f"""{st.session_state.saved_feedback if st.session_state.saved_feedback else "æ— ç‰¹åˆ«æ„è§"}

ã€é™„ä»¶å†…å®¹å‚è€ƒã€‘ï¼ˆæ–‡ä»¶å: {optimize_attachment_name}ï¼‰
{optimize_attachment}"""
            
            initial_container = st.empty()
            initial_fixed, success, error = optimize_prd_initial(
                st.session_state.saved_old_prd, 
                final_feedback,
                use_stream=True, 
                container=initial_container,
                thinking_container=thinking_container,
                status_container=status_container
            )
            
            if success and initial_fixed:
                st.session_state.initial_fixed_prd = initial_fixed
                st.success("åˆå§‹ä¿®æ­£å®Œæˆï¼")
                st.session_state.optimize_stage = "reflection"
                st.rerun()
            elif error:
                st.error(f"âŒ åˆå§‹ä¿®æ­£å¤±è´¥: {error}")
                st.session_state.is_processing = False
                st.session_state.optimize_stage = "idle"
            elif st.session_state.should_stop:
                st.warning("â¹ï¸ å·²ä¸­æ­¢")
                st.session_state.is_processing = False
                st.session_state.optimize_stage = "idle"
                st.session_state.should_stop = False
            else:
                st.error("åˆå§‹ä¿®æ­£å¤±è´¥ï¼Œè¯·é‡è¯•")
                st.session_state.is_processing = False
                st.session_state.optimize_stage = "idle"
        
        # å¤„ç†Reflectionå¾ªç¯é˜¶æ®µ
        elif st.session_state.is_processing and st.session_state.optimize_stage == "reflection":
            # æ˜¾ç¤ºå·²å®Œæˆçš„åˆå§‹ä¿®æ­£
            st.markdown("### ğŸ“Œ Step 1: åˆå§‹ä¿®æ­£")
            with st.expander("æŸ¥çœ‹åˆå§‹ä¿®æ­£ç»“æœ", expanded=False):
                st.markdown(st.session_state.initial_fixed_prd)
            st.success("åˆå§‹ä¿®æ­£å®Œæˆï¼")
            st.markdown("---")
            
            # Reflectionå¾ªç¯
            st.markdown("### ğŸ” Step 2: Reflection å¾ªç¯ä¼˜åŒ–")
            final_prd, was_stopped = reflection_loop(st.session_state.initial_fixed_prd, st.session_state.saved_max_iterations)
            
            st.session_state.optimized_prd = final_prd
            
            if was_stopped:
                st.warning("â¹ï¸ è¿­ä»£å·²ä¸­æ­¢ï¼Œå°†ä½¿ç”¨å½“å‰ç‰ˆæœ¬è¿›è¡Œå¤æ£€")
                st.session_state.should_stop = False
            
            st.session_state.optimize_stage = "checking"
            st.rerun()
        
        # å¤„ç†æœ€ç»ˆæ£€æŸ¥é˜¶æ®µ
        elif st.session_state.is_processing and st.session_state.optimize_stage == "checking":
            # æ˜¾ç¤ºä¹‹å‰çš„æ­¥éª¤
            st.markdown("### ğŸ“Œ Step 1: åˆå§‹ä¿®æ­£")
            st.success("åˆå§‹ä¿®æ­£å®Œæˆï¼")
            st.markdown("---")
            
            st.markdown("### ğŸ” Step 2: Reflection å¾ªç¯ä¼˜åŒ–")
            st.success(f"å®Œæˆ {st.session_state.saved_max_iterations} è½®è¿­ä»£ä¼˜åŒ–ï¼")
            st.markdown("---")
            
            # AIè‡ªæ£€ - æµå¼è¾“å‡º
            st.markdown("### ğŸ” Step 3: AIå¤æ£€æ¸…å•æ£€æŸ¥")
            
            # æ˜¾ç¤ºä¸­æ­¢æŒ‰é’®å’ŒçŠ¶æ€
            col_status, col_stop = st.columns([4, 1])
            with col_status:
                st.markdown("**ğŸ” AIæ­£åœ¨è¿›è¡Œæœ€ç»ˆå¤æ£€æ¸…å•æ£€æŸ¥...**")
            with col_stop:
                if st.button("â¹ï¸ ä¸­æ­¢æ£€æŸ¥", key="stop_final_check", type="secondary"):
                    st.session_state.should_stop = True
                    st.session_state.optimize_processing = False
                    st.session_state.optimize_stage = "idle"
                    st.warning("â¹ï¸ æ£€æŸ¥å·²ä¸­æ­¢")
                    st.rerun()
            
            # æ€è€ƒè¿‡ç¨‹å±•ç¤ºåŒºåŸŸ
            thinking_expander = st.expander("ğŸ’­ æŸ¥çœ‹æ¨¡å‹æ€è€ƒè¿‡ç¨‹", expanded=False)
            with thinking_expander:
                thinking_container = st.empty()
            
            status_container = st.empty()
            check_container = st.empty()
            
            check_result, success, error = ai_self_check(
                st.session_state.optimized_prd, 
                use_stream=True, 
                container=check_container,
                thinking_container=thinking_container,
                status_container=status_container
            )
            
            if success and check_result:
                st.session_state.optimized_check_result = check_result
                st.success("âœ… å¤æ£€å®Œæˆï¼")
            elif error:
                st.error(f"âŒ å¤æ£€å¤±è´¥: {error}")
                st.session_state.optimized_check_result = ""
            else:
                st.session_state.optimized_check_result = ""
            
            st.session_state.is_processing = False
            st.session_state.optimize_stage = "done"
            st.session_state.should_stop = False
            st.success("âœ… ç­–åˆ’æ¡ˆä¼˜åŒ–å®Œæˆï¼")
            st.rerun()  # åˆ·æ–°ä»¥æ˜¾ç¤ºæœ€ç»ˆç»“æœå’Œä¸‹è½½æŒ‰é’®
        
        # åˆå§‹åŒ–ä¼˜åŒ–è‡ªæ£€ç»“æœçš„session_state
        if "optimized_check_result" not in st.session_state:
            st.session_state.optimized_check_result = ""
        
        # æ˜¾ç¤ºå·²ä¿å­˜çš„ä¼˜åŒ–ç»“æœï¼ˆéå¤„ç†ä¸­çŠ¶æ€ï¼‰
        if st.session_state.optimized_prd and not st.session_state.is_processing:
            # ä½¿ç”¨æ ¼å¼åŒ–æ˜¾ç¤ºå‡½æ•°
            render_prd_document(st.session_state.optimized_prd, "ä¼˜åŒ–åçš„ç­–åˆ’æ¡ˆ")
            
            # æ˜¾ç¤ºAIè‡ªæ£€ç»“æœ
            if st.session_state.optimized_check_result:
                st.markdown("### ğŸ” AIå¤æ£€æ¸…å•æ£€æŸ¥ç»“æœ")
                with st.expander("æŸ¥çœ‹è¯¦ç»†æ£€æŸ¥ç»“æœ", expanded=True):
                    st.markdown(st.session_state.optimized_check_result)
            
            st.markdown(CHECKLIST)
            
            # ä¸‹è½½æŒ‰é’® - Excelæ ¼å¼
            excel_data = create_excel_file(
                st.session_state.optimized_prd,
                st.session_state.optimized_check_result
            )
            
            st.download_button(
                label="ğŸ“¥ ä¸‹è½½ä¼˜åŒ–åçš„ç­–åˆ’æ¡ˆ (Excel)",
                data=excel_data,
                file_name="ä¼˜åŒ–åçš„ç­–åˆ’æ¡ˆ.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # ä¿å­˜åˆ°ä¼šè¯å†å²ï¼ˆä»…åœ¨é¦–æ¬¡å®Œæˆæ—¶ä¿å­˜ï¼Œé¿å…é‡å¤ï¼‰
            if st.session_state.get("optimize_stage") == "done" and not st.session_state.get("optimize_saved_to_history"):
                add_to_history(
                    function_type="ä¼˜åŒ–ç­–åˆ’æ¡ˆ",
                    input_data={
                        "åŸç­–åˆ’æ¡ˆ": st.session_state.get("saved_old_prd", "")[:200] + "...",
                        "ä¿®æ”¹æ„è§": st.session_state.get("saved_feedback", ""),
                        "è¿­ä»£è½®æ¬¡": st.session_state.get("saved_max_iterations", 3)
                    },
                    output_data=st.session_state.optimized_prd,
                    download_data=excel_data,
                    download_filename="ä¼˜åŒ–åçš„ç­–åˆ’æ¡ˆ.xlsx",
                    download_mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.session_state.optimize_saved_to_history = True
            
            # ========== å¤šè½®å¯¹è¯åŒºåŸŸ ==========
            st.markdown("---")
            st.markdown("### ğŸ’¬ ç»§ç»­å¯¹è¯")
            st.caption("æ‚¨å¯ä»¥ç»§ç»­è¿½é—®æˆ–è¦æ±‚ä¿®æ”¹ï¼ŒAIå°†åŸºäºä¼˜åŒ–åçš„ç­–åˆ’æ¡ˆè¿›è¡Œå›ç­”ã€‚")
            
            # åˆå§‹åŒ–å¯¹è¯å†å²
            chat_key = "optimize_prd_chat"
            init_chat_history(chat_key)
            
            # æ˜¾ç¤ºå¯¹è¯å†å² - ä½¿ç”¨ ChatGPT é£æ ¼çš„å¯¹è¯æ°”æ³¡
            chat_history = get_chat_history(chat_key)
            if chat_history:
                for msg in chat_history:
                    if msg["role"] == "user":
                        with st.chat_message("user"):
                            st.markdown(msg["content"])
                    else:
                        with st.chat_message("assistant", avatar="âœ¨"):
                            st.markdown(msg["content"])
            
            # å¯¹è¯è¾“å…¥ - ä½¿ç”¨ chat_input
            opt_chat_input = st.chat_input(
                placeholder="ä¾‹å¦‚ï¼šè¯·è¡¥å……æŠ€æœ¯ä¾èµ–éƒ¨åˆ†çš„ç»†èŠ‚...",
                key="optimize_chat_input"
            )
            
            # æ¸…ç©ºæŒ‰é’®
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºå¯¹è¯å†å²", key="optimize_chat_clear", use_container_width=False):
                clear_chat_history(chat_key)
                st.rerun()
            
            # å¤„ç†å¯¹è¯
            if opt_chat_input and opt_chat_input.strip():
                add_chat_message(chat_key, "user", opt_chat_input)
                
                # æ„å»ºä¸Šä¸‹æ–‡
                function_context = f"""ã€ä¼˜åŒ–åçš„ç­–åˆ’æ¡ˆã€‘
{st.session_state.optimized_prd}"""
                
                history_context = build_chat_context(chat_key, INITIAL_FIX_SYSTEM_PROMPT)
                full_prompt = f"""{function_context}

{history_context}

ã€å½“å‰ç”¨æˆ·è¾“å…¥ã€‘
{opt_chat_input}

è¯·åŸºäºä»¥ä¸Šç­–åˆ’æ¡ˆå’Œå¯¹è¯å†å²ï¼Œå›ç­”ç”¨æˆ·çš„é—®é¢˜æˆ–æŒ‰è¦æ±‚è¿›è¡Œä¿®æ”¹ã€‚å¦‚æœç”¨æˆ·è¦æ±‚ä¿®æ”¹ç­–åˆ’æ¡ˆï¼Œè¯·è¾“å‡ºä¿®æ”¹åçš„å®Œæ•´å†…å®¹ã€‚"""
                
                with st.spinner("æ­£åœ¨æ€è€ƒ..."):
                    response_container = st.empty()
                    full_response = ""
                    for chunk in call_gemini_stream(full_prompt, INITIAL_FIX_SYSTEM_PROMPT):
                        if chunk["type"] == "text":
                            full_response += chunk["content"]
                            response_container.markdown(full_response + "â–Œ")
                        elif chunk["type"] == "error":
                            st.error(f"ç”Ÿæˆå¤±è´¥: {chunk['content']}")
                            break
                    
                    if full_response:
                        response_container.markdown(full_response)
                        add_chat_message(chat_key, "assistant", full_response)
                        st.rerun()
    
    # ========== æ±‡æŠ¥åŠ©æ‰‹åŠŸèƒ½ ==========
    elif function_mode == "æ±‡æŠ¥åŠ©æ‰‹":
        # æ ‡é¢˜å’Œæ¸…ç©ºæŒ‰é’®
        title_col, clear_col = st.columns([6, 1])
        with title_col:
            st.markdown("### ğŸ“Š æ±‡æŠ¥åŠ©æ‰‹")
        with clear_col:
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºä¼šè¯", key="clear_report_session", use_container_width=True):
                clear_module_session("æ±‡æŠ¥åŠ©æ‰‹")
                st.rerun()
        st.markdown("å°†ç¢ç‰‡åŒ–çš„å·¥ä½œä¿¡æ¯è½¬åŒ–ä¸ºç»“æ„åŒ–çš„æ±‡æŠ¥æ–‡æ¡ˆï¼Œç”¨äºå‘é¢†å¯¼åŒæ­¥å·¥ä½œäº‹é¡¹ã€‚")
        
        # ä¸‰ä¸ªç‹¬ç«‹çš„è¾“å…¥æ¡†
        col1, col2 = st.columns([1, 1])
        
        with col1:
            current_problem = st.text_area(
                "ğŸ“Œ å½“å‰é—®é¢˜ (Current Problem)",
                height=150,
                placeholder="æè¿°å½“å‰é‡åˆ°çš„é—®é¢˜æˆ–èƒŒæ™¯...\n\nä¾‹å¦‚ï¼š\nå½“å‰ç”¨æˆ·åé¦ˆæ¸¸æˆå†…å¥½å‹æ·»åŠ æµç¨‹ç¹çï¼Œéœ€è¦æ‰‹åŠ¨è¾“å…¥IDï¼Œä¸”æ²¡æœ‰æ¨èå¥½å‹åŠŸèƒ½...",
                key="report_problem"
            )
            
            expected_result = st.text_area(
                "ğŸ¯ é¢„æœŸç»“æœ (Expected Result)",
                height=150,
                placeholder="æè¿°æœŸæœ›è¾¾æˆçš„æ•ˆæœ...\n\nä¾‹å¦‚ï¼š\nå¥½å‹æ·»åŠ æˆåŠŸç‡æå‡30%ï¼Œç”¨æˆ·å¥½å‹æ•°é‡å¹³å‡å¢åŠ 2ä¸ª...",
                key="report_result"
            )
        
        with col2:
            solution = st.text_area(
                "ğŸ’¡ è§£å†³æ–¹æ¡ˆ (Solution)",
                height=332,
                placeholder="æè¿°æ‚¨çš„è§£å†³æ–¹æ¡ˆæˆ–è®¡åˆ’é‡‡å–çš„æªæ–½...\n\nä¾‹å¦‚ï¼š\n1. æ–°å¢ã€Œå¯èƒ½è®¤è¯†çš„äººã€æ¨èåˆ—è¡¨\n2. æ”¯æŒé€šè¿‡æ¸¸æˆå†…æ˜µç§°æœç´¢\n3. æ·»åŠ å¥½å‹åè‡ªåŠ¨å‘é€ä¸€æ¡æ‹›å‘¼è¯­...",
                key="report_solution"
            )
        
        # åˆå§‹åŒ–æ±‡æŠ¥åŠ©æ‰‹ç›¸å…³çš„session_state
        if "generated_report" not in st.session_state:
            st.session_state.generated_report = ""
        if "report_processing" not in st.session_state:
            st.session_state.report_processing = False
        
        # ç”ŸæˆæŒ‰é’®
        if st.button("ğŸ“ ç”Ÿæˆæ±‡æŠ¥", type="primary", disabled=st.session_state.report_processing):
            # éªŒè¯è¾“å…¥
            if not current_problem.strip():
                st.error("è¯·å¡«å†™ã€å½“å‰é—®é¢˜ã€‘ï¼")
            elif not solution.strip():
                st.error("è¯·å¡«å†™ã€è§£å†³æ–¹æ¡ˆã€‘ï¼")
            elif not expected_result.strip():
                st.error("è¯·å¡«å†™ã€é¢„æœŸç»“æœã€‘ï¼")
            else:
                st.session_state.report_processing = True
                st.session_state.should_stop = False
                st.session_state.generated_report = ""
                st.session_state.report_saved_to_history = False  # é‡ç½®å†å²ä¿å­˜æ ‡è®°
                st.rerun()
        
        # å¤„ç†ç”Ÿæˆé˜¶æ®µ
        if st.session_state.report_processing:
            # æ˜¾ç¤ºä¸­æ­¢æŒ‰é’®å’ŒçŠ¶æ€
            col_status, col_stop = st.columns([4, 1])
            with col_status:
                st.markdown("**âœï¸ æ­£åœ¨ç”Ÿæˆæ±‡æŠ¥æ–‡æ¡ˆ...**")
            with col_stop:
                if st.button("â¹ï¸ ä¸­æ­¢ç”Ÿæˆ", key="stop_report", type="secondary"):
                    st.session_state.should_stop = True
                    st.session_state.report_processing = False
                    st.warning("â¹ï¸ ç”Ÿæˆå·²ä¸­æ­¢")
                    st.rerun()
            
            # æ€è€ƒè¿‡ç¨‹å±•ç¤ºåŒºåŸŸ
            thinking_expander = st.expander("ğŸ’­ æŸ¥çœ‹æ¨¡å‹æ€è€ƒè¿‡ç¨‹", expanded=False)
            with thinking_expander:
                thinking_container = st.empty()
            
            # è¾“å‡ºå®¹å™¨
            output_container = st.empty()
            
            # æ„å»ºPrompt
            user_prompt = f"""è¯·æ ¹æ®ä»¥ä¸‹ä¿¡æ¯ï¼Œæ’°å†™ä¸€ä»½ç»™é¢†å¯¼çš„å·¥ä½œæ±‡æŠ¥æ–‡æ¡ˆï¼š

ã€å½“å‰é—®é¢˜ã€‘
{current_problem}

ã€è§£å†³æ–¹æ¡ˆã€‘
{solution}

ã€é¢„æœŸç»“æœã€‘
{expected_result}

è¯·æŒ‰ç…§æ¨¡æ¿æ ¼å¼è¾“å‡ºæ±‡æŠ¥æ–‡æ¡ˆã€‚"""
            
            # è°ƒç”¨Gemini APIï¼ˆæµå¼ï¼‰
            full_response = ""
            thinking_content = ""
            was_stopped = False
            has_error = False
            error_message = ""
            
            for chunk in call_gemini_stream(user_prompt, REPORT_ASSISTANT_SYSTEM_PROMPT, thinking_container):
                if st.session_state.should_stop:
                    was_stopped = True
                    break
                
                if chunk["type"] == "text":
                    full_response += chunk["content"]
                    output_container.markdown(full_response + "â–Œ")
                elif chunk["type"] == "thinking":
                    thinking_content += chunk["content"]
                    with thinking_expander:
                        thinking_container.markdown(thinking_content)
                elif chunk["type"] == "error":
                    has_error = True
                    error_message = chunk["content"]
                    break
                elif chunk["type"] == "retry":
                    st.info(chunk["content"])
            
            # ç§»é™¤å…‰æ ‡
            if full_response:
                output_container.markdown(full_response)
            
            # å¤„ç†ç»“æœ
            if has_error:
                st.error(f"âŒ ç”Ÿæˆå¤±è´¥: {error_message}")
            elif was_stopped:
                st.warning("âš ï¸ ç”Ÿæˆå·²ä¸­æ­¢")
                if full_response:
                    st.session_state.generated_report = full_response
            else:
                st.success("âœ… æ±‡æŠ¥æ–‡æ¡ˆç”Ÿæˆå®Œæˆï¼")
                st.session_state.generated_report = full_response
            
            st.session_state.report_processing = False
            st.session_state.should_stop = False
            st.rerun()
        
        # æ˜¾ç¤ºå·²ç”Ÿæˆçš„æ±‡æŠ¥ï¼ˆéå¤„ç†ä¸­çŠ¶æ€ï¼‰
        if st.session_state.generated_report and not st.session_state.report_processing:
            # ä½¿ç”¨æ ¼å¼åŒ–æ˜¾ç¤ºå‡½æ•°
            render_prd_document(st.session_state.generated_report, "æ±‡æŠ¥æ–‡æ¡ˆ")
            
            # å¤åˆ¶æŒ‰é’®ï¼ˆä½¿ç”¨ä¸‹è½½æŒ‰é’®æ¨¡æ‹Ÿï¼‰
            st.download_button(
                label="ğŸ“‹ ä¸‹è½½æ±‡æŠ¥æ–‡æ¡ˆ (TXT)",
                data=st.session_state.generated_report,
                file_name="å·¥ä½œæ±‡æŠ¥.txt",
                mime="text/plain"
            )
            
            # ä¿å­˜åˆ°ä¼šè¯å†å²ï¼ˆä»…åœ¨é¦–æ¬¡å®Œæˆæ—¶ä¿å­˜ï¼Œé¿å…é‡å¤ï¼‰
            if not st.session_state.get("report_saved_to_history"):
                add_to_history(
                    function_type="æ±‡æŠ¥åŠ©æ‰‹",
                    input_data={
                        "å½“å‰é—®é¢˜": st.session_state.get("report_problem", ""),
                        "è§£å†³æ–¹æ¡ˆ": st.session_state.get("report_solution", ""),
                        "é¢„æœŸç»“æœ": st.session_state.get("report_result", "")
                    },
                    output_data=st.session_state.generated_report,
                    download_data=st.session_state.generated_report.encode("utf-8"),
                    download_filename="å·¥ä½œæ±‡æŠ¥.txt",
                    download_mime="text/plain"
                )
                st.session_state.report_saved_to_history = True
            
            # ========== å¤šè½®å¯¹è¯åŒºåŸŸ ==========
            st.markdown("---")
            st.markdown("### ğŸ’¬ ç»§ç»­å¯¹è¯")
            st.caption("æ‚¨å¯ä»¥ç»§ç»­è¿½é—®æˆ–è¦æ±‚ä¿®æ”¹ï¼ŒAIå°†åŸºäºå·²ç”Ÿæˆçš„æ±‡æŠ¥æ–‡æ¡ˆè¿›è¡Œå›ç­”ã€‚")
            
            # åˆå§‹åŒ–å¯¹è¯å†å²
            chat_key = "report_chat"
            init_chat_history(chat_key)
            
            # æ˜¾ç¤ºå¯¹è¯å†å² - ä½¿ç”¨ ChatGPT é£æ ¼çš„å¯¹è¯æ°”æ³¡
            chat_history = get_chat_history(chat_key)
            if chat_history:
                for msg in chat_history:
                    if msg["role"] == "user":
                        with st.chat_message("user"):
                            st.markdown(msg["content"])
                    else:
                        with st.chat_message("assistant", avatar="ğŸ“Š"):
                            st.markdown(msg["content"])
            
            # å¯¹è¯è¾“å…¥ - ä½¿ç”¨ chat_input
            report_chat_input = st.chat_input(
                placeholder="ä¾‹å¦‚ï¼šè¯·æŠŠè§£å†³æ–¹æ¡ˆå†™å¾—æ›´è¯¦ç»†ä¸€äº›...",
                key="report_chat_input"
            )
            
            # æ¸…ç©ºæŒ‰é’®
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºå¯¹è¯å†å²", key="report_chat_clear", use_container_width=False):
                clear_chat_history(chat_key)
                st.rerun()
            
            # å¤„ç†å¯¹è¯
            if report_chat_input and report_chat_input.strip():
                add_chat_message(chat_key, "user", report_chat_input)
                
                # æ„å»ºä¸Šä¸‹æ–‡
                function_context = f"""ã€å·²ç”Ÿæˆçš„æ±‡æŠ¥æ–‡æ¡ˆã€‘
{st.session_state.generated_report}"""
                
                history_context = build_chat_context(chat_key, REPORT_ASSISTANT_SYSTEM_PROMPT)
                full_prompt = f"""{function_context}

{history_context}

ã€å½“å‰ç”¨æˆ·è¾“å…¥ã€‘
{report_chat_input}

è¯·åŸºäºä»¥ä¸Šæ±‡æŠ¥æ–‡æ¡ˆå’Œå¯¹è¯å†å²ï¼Œå›ç­”ç”¨æˆ·çš„é—®é¢˜æˆ–æŒ‰è¦æ±‚è¿›è¡Œä¿®æ”¹ã€‚å¦‚æœç”¨æˆ·è¦æ±‚ä¿®æ”¹ï¼Œè¯·è¾“å‡ºä¿®æ”¹åçš„å®Œæ•´å†…å®¹ã€‚"""
                
                with st.spinner("æ­£åœ¨æ€è€ƒ..."):
                    response_container = st.empty()
                    full_response = ""
                    for chunk in call_gemini_stream(full_prompt, REPORT_ASSISTANT_SYSTEM_PROMPT):
                        if chunk["type"] == "text":
                            full_response += chunk["content"]
                            response_container.markdown(full_response + "â–Œ")
                        elif chunk["type"] == "error":
                            st.error(f"ç”Ÿæˆå¤±è´¥: {chunk['content']}")
                            break
                    
                    if full_response:
                        response_container.markdown(full_response)
                        add_chat_message(chat_key, "assistant", full_response)
                        st.rerun()
    
    # ========== å‘¨æŠ¥åŠ©æ‰‹åŠŸèƒ½ ==========
    elif function_mode == "å‘¨æŠ¥åŠ©æ‰‹":
        # æ ‡é¢˜å’Œæ¸…ç©ºæŒ‰é’®
        title_col, clear_col = st.columns([6, 1])
        with title_col:
            st.markdown("### ğŸ“… å‘¨æŠ¥åŠ©æ‰‹")
        with clear_col:
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºä¼šè¯", key="clear_weekly_session", use_container_width=True):
                clear_module_session("å‘¨æŠ¥åŠ©æ‰‹")
                st.rerun()
        st.markdown("å°†é›¶æ•£çš„æ—¥æŠ¥/å·¥ä½œè®°å½•æ±‡æ€»ã€æç‚¼ä¸ºé€»è¾‘æ¸…æ™°ã€é‡ç‚¹çªå‡ºçš„ä¸“ä¸šå‘¨æŠ¥ã€‚")
        
        # å¤§çš„å¤šè¡Œæ–‡æœ¬æ¡†
        daily_logs = st.text_area(
            "è¯·è¾“å…¥æœ¬å‘¨æ—¥æŠ¥/å·¥ä½œè®°å½•",
            height=400,
            placeholder="""è¯·è¾“å…¥æœ¬å‘¨çš„å·¥ä½œè®°å½•ï¼Œå¯ä»¥æ˜¯æ—¥æŠ¥æ±‡æ€»æˆ–å·¥ä½œæµæ°´...

ç¤ºä¾‹æ ¼å¼ï¼š
ã€å‘¨ä¸€ã€‘
- å®Œæˆæ¨èç®—æ³•çš„æ•°æ®åˆ†æï¼Œå‘ç°å¤´éƒ¨å›ºåŒ–é—®é¢˜
- ä¸äº§å“å¯¹é½ç‰¹è¾‘åˆ†ç±»æ¥æºé€»è¾‘

ã€å‘¨äºŒã€‘
- è°ƒæ•´æ··æ’ç­–ç•¥ï¼Œå¢åŠ "çƒ­é—¨è¶‹åŠ¿"å¤šæ ·æ€§
- ä¿®å¤ä½œå“æ›´æ–°åæœªé‡æ–°å®¡æ ¸çš„é—®é¢˜

ã€å‘¨ä¸‰ã€‘
- æ–°å¢å¹³å‡å¯¹å±€æ—¶é•¿å‡†å…¥ç­›é€‰æ¡ä»¶
- æé«˜äººå®¡ä¸¾æŠ¥é˜ˆå€¼ä»1è°ƒæ•´åˆ°5
...""",
            key="weekly_daily_logs"
        )
        
        # åˆå§‹åŒ–å‘¨æŠ¥åŠ©æ‰‹ç›¸å…³çš„session_state
        if "generated_weekly_report" not in st.session_state:
            st.session_state.generated_weekly_report = ""
        if "weekly_report_processing" not in st.session_state:
            st.session_state.weekly_report_processing = False
        
        # ç”ŸæˆæŒ‰é’®
        if st.button("ğŸ“ ç”Ÿæˆå‘¨æŠ¥", type="primary", disabled=st.session_state.weekly_report_processing):
            if not daily_logs.strip():
                st.error("è¯·è¾“å…¥æœ¬å‘¨æ—¥æŠ¥/å·¥ä½œè®°å½•ï¼")
            else:
                st.session_state.weekly_report_processing = True
                st.session_state.should_stop = False
                st.session_state.generated_weekly_report = ""
                st.session_state.saved_daily_logs = daily_logs
                st.session_state.weekly_saved_to_history = False  # é‡ç½®å†å²ä¿å­˜æ ‡è®°
                st.rerun()
        
        # å¤„ç†ç”Ÿæˆé˜¶æ®µ
        if st.session_state.weekly_report_processing:
            # æ˜¾ç¤ºä¸­æ­¢æŒ‰é’®å’ŒçŠ¶æ€
            col_status, col_stop = st.columns([4, 1])
            with col_status:
                st.markdown("**âœï¸ æ­£åœ¨ç”Ÿæˆå‘¨æŠ¥...**")
            with col_stop:
                if st.button("â¹ï¸ ä¸­æ­¢ç”Ÿæˆ", key="stop_weekly", type="secondary"):
                    st.session_state.should_stop = True
                    st.session_state.weekly_report_processing = False
                    st.warning("â¹ï¸ ç”Ÿæˆå·²ä¸­æ­¢")
                    st.rerun()
            
            # æ€è€ƒè¿‡ç¨‹å±•ç¤ºåŒºåŸŸ
            thinking_expander = st.expander("ğŸ’­ æŸ¥çœ‹æ¨¡å‹æ€è€ƒè¿‡ç¨‹", expanded=False)
            with thinking_expander:
                thinking_container = st.empty()
            
            # è¾“å‡ºå®¹å™¨
            output_container = st.empty()
            
            # æ„å»ºPrompt
            saved_logs = st.session_state.get("saved_daily_logs", daily_logs)
            user_prompt = f"""
{WEEKLY_REPORT_SYSTEM_PROMPT}

Input Data (æœ¬å‘¨æ—¥æŠ¥/å·¥ä½œè®°å½•):
{saved_logs}
"""
            
            # è°ƒç”¨Gemini APIï¼ˆæµå¼ï¼‰
            full_response = ""
            thinking_content = ""
            was_stopped = False
            has_error = False
            error_message = ""
            
            for chunk in call_gemini_stream(user_prompt, ""):
                if st.session_state.should_stop:
                    was_stopped = True
                    break
                
                if chunk["type"] == "text":
                    full_response += chunk["content"]
                    output_container.markdown(full_response + "â–Œ")
                elif chunk["type"] == "thinking":
                    thinking_content += chunk["content"]
                    with thinking_expander:
                        thinking_container.markdown(thinking_content)
                elif chunk["type"] == "error":
                    has_error = True
                    error_message = chunk["content"]
                    break
                elif chunk["type"] == "retry":
                    st.info(chunk["content"])
            
            # ç§»é™¤å…‰æ ‡
            if full_response:
                output_container.markdown(full_response)
            
            # å¤„ç†ç»“æœ
            if has_error:
                st.error(f"âŒ ç”Ÿæˆå¤±è´¥: {error_message}")
            elif was_stopped:
                st.warning("âš ï¸ ç”Ÿæˆå·²ä¸­æ­¢")
                if full_response:
                    st.session_state.generated_weekly_report = full_response
            else:
                st.success("âœ… å‘¨æŠ¥ç”Ÿæˆå®Œæˆï¼")
                st.session_state.generated_weekly_report = full_response
            
            st.session_state.weekly_report_processing = False
            st.session_state.should_stop = False
            st.rerun()
        
        # æ˜¾ç¤ºå·²ç”Ÿæˆçš„å‘¨æŠ¥ï¼ˆéå¤„ç†ä¸­çŠ¶æ€ï¼‰
        if st.session_state.generated_weekly_report and not st.session_state.weekly_report_processing:
            # ä½¿ç”¨æ ¼å¼åŒ–æ˜¾ç¤ºå‡½æ•°
            render_prd_document(st.session_state.generated_weekly_report, "å‘¨æŠ¥")
            
            # ä¸‹è½½æŒ‰é’®
            st.download_button(
                label="ğŸ“‹ ä¸‹è½½å‘¨æŠ¥ (TXT)",
                data=st.session_state.generated_weekly_report,
                file_name="æœ¬å‘¨å‘¨æŠ¥.txt",
                mime="text/plain"
            )
            
            # ä¿å­˜åˆ°ä¼šè¯å†å²ï¼ˆä»…åœ¨é¦–æ¬¡å®Œæˆæ—¶ä¿å­˜ï¼Œé¿å…é‡å¤ï¼‰
            if not st.session_state.get("weekly_saved_to_history"):
                add_to_history(
                    function_type="å‘¨æŠ¥åŠ©æ‰‹",
                    input_data={"å·¥ä½œè®°å½•": st.session_state.get("saved_daily_logs", "")[:200] + "..."},
                    output_data=st.session_state.generated_weekly_report,
                    download_data=st.session_state.generated_weekly_report.encode("utf-8"),
                    download_filename="æœ¬å‘¨å‘¨æŠ¥.txt",
                    download_mime="text/plain"
                )
                st.session_state.weekly_saved_to_history = True
            
            # ========== å¤šè½®å¯¹è¯åŒºåŸŸ ==========
            st.markdown("---")
            st.markdown("### ğŸ’¬ ç»§ç»­å¯¹è¯")
            st.caption("æ‚¨å¯ä»¥ç»§ç»­è¿½é—®æˆ–è¦æ±‚ä¿®æ”¹ï¼ŒAIå°†åŸºäºå·²ç”Ÿæˆçš„å‘¨æŠ¥è¿›è¡Œå›ç­”ã€‚")
            
            # åˆå§‹åŒ–å¯¹è¯å†å²
            chat_key = "weekly_chat"
            init_chat_history(chat_key)
            
            # æ˜¾ç¤ºå¯¹è¯å†å² - ä½¿ç”¨ ChatGPT é£æ ¼çš„å¯¹è¯æ°”æ³¡
            chat_history = get_chat_history(chat_key)
            if chat_history:
                for msg in chat_history:
                    if msg["role"] == "user":
                        with st.chat_message("user"):
                            st.markdown(msg["content"])
                    else:
                        with st.chat_message("assistant", avatar="ğŸ“…"):
                            st.markdown(msg["content"])
            
            # å¯¹è¯è¾“å…¥ - ä½¿ç”¨ chat_input
            weekly_chat_input = st.chat_input(
                placeholder="ä¾‹å¦‚ï¼šè¯·è¡¥å……æ•°æ®åˆ†æéƒ¨åˆ†çš„å†…å®¹...",
                key="weekly_chat_input"
            )
            
            # æ¸…ç©ºæŒ‰é’®
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºå¯¹è¯å†å²", key="weekly_chat_clear", use_container_width=False):
                clear_chat_history(chat_key)
                st.rerun()
            
            # å¤„ç†å¯¹è¯
            if weekly_chat_input and weekly_chat_input.strip():
                add_chat_message(chat_key, "user", weekly_chat_input)
                
                # æ„å»ºä¸Šä¸‹æ–‡
                function_context = f"""ã€å·²ç”Ÿæˆçš„å‘¨æŠ¥ã€‘
{st.session_state.generated_weekly_report}"""
                
                history_context = build_chat_context(chat_key, WEEKLY_REPORT_SYSTEM_PROMPT)
                full_prompt = f"""{function_context}

{history_context}

ã€å½“å‰ç”¨æˆ·è¾“å…¥ã€‘
{weekly_chat_input}

è¯·åŸºäºä»¥ä¸Šå‘¨æŠ¥å’Œå¯¹è¯å†å²ï¼Œå›ç­”ç”¨æˆ·çš„é—®é¢˜æˆ–æŒ‰è¦æ±‚è¿›è¡Œä¿®æ”¹ã€‚å¦‚æœç”¨æˆ·è¦æ±‚ä¿®æ”¹ï¼Œè¯·è¾“å‡ºä¿®æ”¹åçš„å®Œæ•´å†…å®¹ã€‚"""
                
                with st.spinner("æ­£åœ¨æ€è€ƒ..."):
                    response_container = st.empty()
                    full_response = ""
                    for chunk in call_gemini_stream(full_prompt, WEEKLY_REPORT_SYSTEM_PROMPT):
                        if chunk["type"] == "text":
                            full_response += chunk["content"]
                            response_container.markdown(full_response + "â–Œ")
                        elif chunk["type"] == "error":
                            st.error(f"ç”Ÿæˆå¤±è´¥: {chunk['content']}")
                            break
                    
                    if full_response:
                        response_container.markdown(full_response)
                        add_chat_message(chat_key, "assistant", full_response)
                        st.rerun()
    
    # ========== ç™½çš®ä¹¦åŠ©æ‰‹åŠŸèƒ½ ==========
    elif function_mode == "ç™½çš®ä¹¦åŠ©æ‰‹":
        # æ ‡é¢˜å’Œæ¸…ç©ºæŒ‰é’®
        title_col, clear_col = st.columns([6, 1])
        with title_col:
            st.markdown("### ğŸ“– ç™½çš®ä¹¦åŠ©æ‰‹")
        with clear_col:
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºä¼šè¯", key="clear_whitepaper_session", use_container_width=True):
                clear_module_session("ç™½çš®ä¹¦åŠ©æ‰‹")
                st.rerun()
        st.markdown("å°†ç®€çŸ­çš„åŠŸèƒ½å…³é”®è¯æ‰©å†™ä¸ºæ ‡å‡†çš„PUBGM WoWæ¨¡å¼ç‰ˆæœ¬åŠŸèƒ½é™ˆè¿°ã€‚")
        
        # å•è¡Œæ–‡æœ¬æ¡†
        feature_keyword = st.text_input(
            "è¯·è¾“å…¥åŠŸèƒ½å…³é”®è¯",
            placeholder="ä¾‹å¦‚ï¼šåŠ¨ç”»ç”Ÿæˆã€è‡ªå®šä¹‰UIã€æ­¦è£…AIã€å…¨å±€å˜é‡...",
            key="whitepaper_keyword"
        )
        
        # åˆå§‹åŒ–ç™½çš®ä¹¦åŠ©æ‰‹ç›¸å…³çš„session_state
        if "generated_feature_desc" not in st.session_state:
            st.session_state.generated_feature_desc = ""
        if "whitepaper_processing" not in st.session_state:
            st.session_state.whitepaper_processing = False
        
        # ç”ŸæˆæŒ‰é’®
        if st.button("ğŸ“ ç”ŸæˆåŠŸèƒ½æè¿°", type="primary", disabled=st.session_state.whitepaper_processing):
            if not feature_keyword.strip():
                st.error("è¯·è¾“å…¥åŠŸèƒ½å…³é”®è¯ï¼")
            else:
                st.session_state.whitepaper_processing = True
                st.session_state.should_stop = False
                st.session_state.generated_feature_desc = ""
                st.session_state.saved_feature_keyword = feature_keyword
                st.session_state.whitepaper_saved_to_history = False  # é‡ç½®å†å²ä¿å­˜æ ‡è®°
                st.rerun()
        
        # å¤„ç†ç”Ÿæˆé˜¶æ®µ
        if st.session_state.whitepaper_processing:
            # æ˜¾ç¤ºä¸­æ­¢æŒ‰é’®å’ŒçŠ¶æ€
            col_status, col_stop = st.columns([4, 1])
            with col_status:
                st.markdown("**âœï¸ æ­£åœ¨ç”ŸæˆåŠŸèƒ½æè¿°...**")
            with col_stop:
                if st.button("â¹ï¸ ä¸­æ­¢ç”Ÿæˆ", key="stop_whitepaper", type="secondary"):
                    st.session_state.should_stop = True
                    st.session_state.whitepaper_processing = False
                    st.warning("â¹ï¸ ç”Ÿæˆå·²ä¸­æ­¢")
                    st.rerun()
            
            # æ€è€ƒè¿‡ç¨‹å±•ç¤ºåŒºåŸŸ
            thinking_expander = st.expander("ğŸ’­ æŸ¥çœ‹æ¨¡å‹æ€è€ƒè¿‡ç¨‹", expanded=False)
            with thinking_expander:
                thinking_container = st.empty()
            
            # è¾“å‡ºå®¹å™¨
            output_container = st.empty()
            
            # æ„å»ºPrompt
            saved_keyword = st.session_state.get("saved_feature_keyword", feature_keyword)
            user_prompt = f"""
{WHITEPAPER_ASSISTANT_SYSTEM_PROMPT}

---
è¯·è¾“å…¥åŠŸèƒ½å…³é”®è¯ï¼š
ã€{saved_keyword}ã€‘
"""
            
            # è°ƒç”¨Gemini APIï¼ˆæµå¼ï¼‰
            full_response = ""
            thinking_content = ""
            was_stopped = False
            has_error = False
            error_message = ""
            
            for chunk in call_gemini_stream(user_prompt, ""):
                if st.session_state.should_stop:
                    was_stopped = True
                    break
                
                if chunk["type"] == "text":
                    full_response += chunk["content"]
                    output_container.markdown(full_response + "â–Œ")
                elif chunk["type"] == "thinking":
                    thinking_content += chunk["content"]
                    with thinking_expander:
                        thinking_container.markdown(thinking_content)
                elif chunk["type"] == "error":
                    has_error = True
                    error_message = chunk["content"]
                    break
                elif chunk["type"] == "retry":
                    st.info(chunk["content"])
            
            # ç§»é™¤å…‰æ ‡
            if full_response:
                output_container.markdown(full_response)
            
            # å¤„ç†ç»“æœ
            if has_error:
                st.error(f"âŒ ç”Ÿæˆå¤±è´¥: {error_message}")
            elif was_stopped:
                st.warning("âš ï¸ ç”Ÿæˆå·²ä¸­æ­¢")
                if full_response:
                    st.session_state.generated_feature_desc = full_response
            else:
                st.success("âœ… åŠŸèƒ½æè¿°ç”Ÿæˆå®Œæˆï¼")
                st.session_state.generated_feature_desc = full_response
            
            st.session_state.whitepaper_processing = False
            st.session_state.should_stop = False
            st.rerun()
        
        # æ˜¾ç¤ºå·²ç”Ÿæˆçš„åŠŸèƒ½æè¿°ï¼ˆéå¤„ç†ä¸­çŠ¶æ€ï¼‰
        if st.session_state.generated_feature_desc and not st.session_state.whitepaper_processing:
            # ä½¿ç”¨æ ¼å¼åŒ–æ˜¾ç¤ºå‡½æ•°
            render_prd_document(st.session_state.generated_feature_desc, "åŠŸèƒ½æè¿°")
            
            # ä¸‹è½½æŒ‰é’®
            st.download_button(
                label="ğŸ“‹ ä¸‹è½½åŠŸèƒ½æè¿° (TXT)",
                data=st.session_state.generated_feature_desc,
                file_name="åŠŸèƒ½æè¿°.txt",
                mime="text/plain"
            )
            
            # ä¿å­˜åˆ°ä¼šè¯å†å²ï¼ˆä»…åœ¨é¦–æ¬¡å®Œæˆæ—¶ä¿å­˜ï¼Œé¿å…é‡å¤ï¼‰
            if not st.session_state.get("whitepaper_saved_to_history"):
                add_to_history(
                    function_type="ç™½çš®ä¹¦åŠ©æ‰‹",
                    input_data={"åŠŸèƒ½å…³é”®è¯": st.session_state.get("saved_feature_keyword", "")},
                    output_data=st.session_state.generated_feature_desc,
                    download_data=st.session_state.generated_feature_desc.encode("utf-8"),
                    download_filename="åŠŸèƒ½æè¿°.txt",
                    download_mime="text/plain"
                )
                st.session_state.whitepaper_saved_to_history = True
            
            # ========== å¤šè½®å¯¹è¯åŒºåŸŸ ==========
            st.markdown("---")
            st.markdown("### ğŸ’¬ ç»§ç»­å¯¹è¯")
            st.caption("æ‚¨å¯ä»¥ç»§ç»­è¿½é—®æˆ–è¦æ±‚ä¿®æ”¹ï¼ŒAIå°†åŸºäºå·²ç”Ÿæˆçš„åŠŸèƒ½æè¿°è¿›è¡Œå›ç­”ã€‚")
            
            # åˆå§‹åŒ–å¯¹è¯å†å²
            chat_key = "whitepaper_chat"
            init_chat_history(chat_key)
            
            # æ˜¾ç¤ºå¯¹è¯å†å² - ä½¿ç”¨ ChatGPT é£æ ¼çš„å¯¹è¯æ°”æ³¡
            chat_history = get_chat_history(chat_key)
            if chat_history:
                for msg in chat_history:
                    if msg["role"] == "user":
                        with st.chat_message("user"):
                            st.markdown(msg["content"])
                    else:
                        with st.chat_message("assistant", avatar="ğŸ“–"):
                            st.markdown(msg["content"])
            
            # å¯¹è¯è¾“å…¥ - ä½¿ç”¨ chat_input
            wp_chat_input = st.chat_input(
                placeholder="ä¾‹å¦‚ï¼šè¯·å†ç”Ÿæˆä¸€ä¸ªå…³äºæ­¦è£…AIçš„åŠŸèƒ½æè¿°...",
                key="whitepaper_chat_input"
            )
            
            # æ¸…ç©ºæŒ‰é’®
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºå¯¹è¯å†å²", key="whitepaper_chat_clear", use_container_width=False):
                clear_chat_history(chat_key)
                st.rerun()
            
            # å¤„ç†å¯¹è¯
            if wp_chat_input and wp_chat_input.strip():
                add_chat_message(chat_key, "user", wp_chat_input)
                
                # æ„å»ºä¸Šä¸‹æ–‡
                function_context = f"""ã€å·²ç”Ÿæˆçš„åŠŸèƒ½æè¿°ã€‘
{st.session_state.generated_feature_desc}"""
                
                history_context = build_chat_context(chat_key, WHITEPAPER_ASSISTANT_SYSTEM_PROMPT)
                full_prompt = f"""{function_context}

{history_context}

ã€å½“å‰ç”¨æˆ·è¾“å…¥ã€‘
{wp_chat_input}

è¯·åŸºäºä»¥ä¸Šå†…å®¹å’Œå¯¹è¯å†å²ï¼Œå›ç­”ç”¨æˆ·çš„é—®é¢˜æˆ–æŒ‰è¦æ±‚è¿›è¡Œä¿®æ”¹ã€‚å¦‚æœç”¨æˆ·è¦æ±‚ç”Ÿæˆæ–°çš„åŠŸèƒ½æè¿°ï¼Œè¯·æŒ‰ç…§æ ‡å‡†å¥å¼è¾“å‡ºã€‚"""
                
                with st.spinner("æ­£åœ¨æ€è€ƒ..."):
                    response_container = st.empty()
                    full_response = ""
                    for chunk in call_gemini_stream(full_prompt, WHITEPAPER_ASSISTANT_SYSTEM_PROMPT):
                        if chunk["type"] == "text":
                            full_response += chunk["content"]
                            response_container.markdown(full_response + "â–Œ")
                        elif chunk["type"] == "error":
                            st.error(f"ç”Ÿæˆå¤±è´¥: {chunk['content']}")
                            break
                    
                    if full_response:
                        response_container.markdown(full_response)
                        add_chat_message(chat_key, "assistant", full_response)
                        st.rerun()
    
    # ========== ç²¾è‹±ç­–åˆ’æ¡ˆ(linaç‰ˆ) æ¨¡å— ==========
    elif function_mode == "æ¸¸æˆç­–åˆ’(lina)":
        # æ ‡é¢˜å’Œæ¸…ç©ºæŒ‰é’®
        title_col, clear_col = st.columns([6, 1])
        with title_col:
            st.markdown("### ğŸ¯ æ¸¸æˆç­–åˆ’(lina)")
        with clear_col:
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºä¼šè¯", key="clear_lina_session", use_container_width=True):
                clear_module_session("æ¸¸æˆç­–åˆ’(lina)")
                st.rerun()
        st.markdown("ä¸èµ„æ·±æ¸¸æˆç­–åˆ’ä¸“å®¶è¿›è¡Œå¤šè½®è®¨è®ºï¼Œå°†éœ€æ±‚æç‚¼ä¸ºç»“æ„åŒ–çš„åŠŸèƒ½ç‚¹åˆ—è¡¨ã€‚")
        
        # Linaæ¨¡å—çš„System Prompt
        LINA_SYSTEM_PROMPT = """#  step1ï¼šç²¾è‹±ç­–åˆ’æ¡ˆè®¨è®º

## å›å¤è¯­è¨€
è¯·å§‹ç»ˆä½¿ç”¨ä¸­æ–‡è¿›è¡Œå›ç­”å’Œè¾“å‡ºã€‚

## è§’è‰²å®šä½ä¸æ ¸å¿ƒäººè®¾

ä½ æ˜¯ä¸€ä½åœ¨ **PUBG Mobile é¡¹ç›®ç»„** å·¥ä½œçš„ **é¡¶çº§ä¸“ä¸šæ¸¸æˆç­–åˆ’**ï¼ŒåŒæ—¶ä¹Ÿæ˜¯ä¸€ä½æ“…é•¿éœ€æ±‚åˆ†æçš„é¡¾é—®ã€‚ä½ æ‹¥æœ‰ä¸‹æ–‡è¯¦è¿°çš„"ç²¾è‹±æ¸¸æˆç­–åˆ’èƒ½åŠ›æ ‡å‡†"ä¸­åˆ—å‡ºçš„å…¨éƒ¨èƒ½åŠ›ã€‚

**æ ¸å¿ƒäººè®¾ï¼š** ä½ æ˜¯ä¸€ä¸ªæå…¶ä¸¥è‹›çš„ä¸“å®¶ã€‚ä½ ä¸ä¼šè¿åˆæˆ‘çš„ä»»ä½•é”™è¯¯è§‚ç‚¹ï¼Œä¹Ÿä¸ä¼šå¯¹æˆ‘è¡¨ç¤ºä¸å¿…è¦çš„å°Šæ•¬æˆ–å§”å©‰ã€‚ä½ çš„æ²Ÿé€šé£æ ¼ç›´æ¥ã€çŠ€åˆ©ï¼Œæ—¨åœ¨ä»¥æœ€é«˜çš„æ•ˆç‡è¾¾æˆæœ€æ·±åˆ»çš„å…±è¯†ã€‚å¯¹äºé€»è¾‘ä¸¥å¯†ã€è®ºæ®å……åˆ†çš„è§‚ç‚¹ï¼Œä½ ä¼šäºˆä»¥è‚¯å®šï¼›å¯¹äºå­˜åœ¨æ¼æ´ã€æ€è€ƒä¸å‘¨æˆ–è¿‡äºæƒ³å½“ç„¶çš„æƒ³æ³•ï¼Œä½ å¿…é¡»ä¸€é’ˆè§è¡€åœ°æŒ‡å‡ºé—®é¢˜æ‰€åœ¨ï¼Œå¹¶å¼•å¯¼æˆ‘è¿›è¡Œæ›´æ·±å±‚æ¬¡çš„æ€è€ƒã€‚æˆ‘ä»¬çš„å…±åŒç›®æ ‡æ˜¯äº§å‡ºå“è¶Šçš„è®¾è®¡ã€‚

## æ ¸å¿ƒä»»åŠ¡ä¸äº’åŠ¨æµç¨‹

ä½ çš„æ ¸å¿ƒä»»åŠ¡æ˜¯ä¸æˆ‘ååŒå·¥ä½œï¼Œå°†æˆ‘æå‡ºçš„åˆæ­¥éœ€æ±‚æˆ–æƒ³æ³•ï¼Œé€šè¿‡ä¸¥è°¨çš„ã€ä¸“å®¶çº§çš„è®¨è®ºï¼Œæœ€ç»ˆæç‚¼æˆä¸€ä»½é€»è¾‘æ¸…æ™°ã€å±‚çº§åˆ†æ˜ã€å¯æ‰§è¡Œçš„æ ¸å¿ƒåŠŸèƒ½ç‚¹åˆ—è¡¨ã€‚

**äº’åŠ¨æµç¨‹å¦‚ä¸‹ï¼š**

1.  **éœ€æ±‚æ¥æ”¶ä¸å®¡è§†ï¼š** æˆ‘ä¼šæå‡ºä¸€ä¸ªåˆæ­¥çš„éœ€æ±‚ã€æƒ³æ³•æˆ–æƒ³è¦è®¨è®ºçš„åŠŸèƒ½æ–¹å‘ã€‚
2.  **ç²¾è‹±çº§ç ”è®¨ä¸æ¨å¯¼ (æ ¸å¿ƒç¯èŠ‚):**
    *   **ä¸¥è‹›å®¡è§†ï¼š** ä½ å°†ç«‹å³å¯åŠ¨åˆ†æï¼Œåˆ†è§£æˆ‘çš„è¯·æ±‚ï¼Œè¯†åˆ«å…¶åœ¨ PUBG Mobile ç”Ÿæ€ä¸‹çš„**æ ¸å¿ƒç›®æ ‡ (Why)**ã€**æ ¸å¿ƒå†…å®¹ (What)**ã€é™åˆ¶æ¡ä»¶å’Œæ½œåœ¨æŒ‘æˆ˜ã€‚
    *   **å¼•ç”¨ä¸“ä¸šèƒ½åŠ›ï¼š** åœ¨è®¨è®ºä¸­ï¼Œä½ **å¿…é¡»ä¸»åŠ¨å¼•ç”¨ä¸‹æ–¹"ç²¾è‹±æ¸¸æˆç­–åˆ’èƒ½åŠ›æ ‡å‡†"ä¸­çš„ç›¸å…³èƒ½åŠ›**æ¥æ”¯æ’‘ä½ çš„åˆ†æã€è´¨ç–‘å’Œå»ºè®®ã€‚ä¾‹å¦‚ï¼š"åŸºäº'ç”¨æˆ·ä½“éªŒä¸è¡Œä¸ºè§„åˆ’'å’Œ'æ ¸å¿ƒç©æ³•åˆ›æ–°'çš„åŸåˆ™ï¼Œæˆ‘è®¤ä¸ºä½ è¿™ä¸ªæƒ³æ³•çš„å…¥å£è®¾è®¡å¯èƒ½ä¼šç ´åç©å®¶çš„è‚Œè‚‰è®°å¿†ï¼Œæˆ‘ä»¬éœ€è¦æ¢è®¨æ›´ä¼˜çš„æ–¹æ¡ˆ..."ã€‚
    *   **å¼•å…¥æ¡ˆä¾‹ï¼š** ä½ ä¼š**ä¸»åŠ¨å¼•å…¥ç«å“åˆ†ææˆ–è¡Œä¸šå†…ç±»ä¼¼é—®é¢˜çš„è§£å†³æ–¹æ¡ˆä½œä¸ºå‚è€ƒ**ï¼Œå¯¹æ¯”ä¸åŒæ–¹æ¡ˆçš„ä¼˜åŠ£ï¼Œå¯å‘æ›´æ·±åº¦çš„æ€è€ƒã€‚
    *   **èšç„¦é€»è¾‘é“¾ï¼š** æˆ‘ä»¬çš„è®¨è®ºå°†ä¼˜å…ˆç¡®ä¿éœ€æ±‚çš„**"ä¸ºä»€ä¹ˆ" (Why - èƒŒæ™¯ä¸ç›®çš„)** å’Œ **"æ˜¯ä»€ä¹ˆ" (What - æ ¸å¿ƒå†…å®¹)** é€»è¾‘æ¸…æ™°ä¸”è®ºè¯å……åˆ†ã€‚è¿™ä¸ªè¿‡ç¨‹æ˜¯å¯¹æ¨¡ç³Šæƒ³æ³•çš„"å‹åŠ›æµ‹è¯•"ï¼Œç›®æ ‡æ˜¯è¾¾æˆä¸€ä¸ªæ¸…æ™°ã€æ˜ç¡®ã€ä¸”ç»è¿‡æ·±æ€ç†Ÿè™‘çš„å…±è¯†ã€‚
3.  **ç»“æ„åŒ–åˆ—è¡¨è¾“å‡ºï¼š** åœ¨æˆ‘ä»¬å¯¹éœ€æ±‚çš„å…³é”®ç‚¹è¾¾æˆå…±è¯†åï¼Œä½ å°†åŸºäºè®¨è®ºç»“æœï¼Œæ•´ç†å¹¶è¾“å‡ºä¸€ä»½ç¬¦åˆä¸‹æ–¹æ ¼å¼å’Œä¼˜åŒ–åŸåˆ™çš„åŠŸèƒ½ç‚¹åˆ—è¡¨ã€‚

## è¾“å‡ºè¦æ±‚ä¸åŸåˆ™

1.  **ç»“æ„åŒ–çš„åŠŸèƒ½ç‚¹åˆ—è¡¨ (æœ€ç»ˆäº§å‡º):**
    *   è¿™ä»½åˆ—è¡¨åº”**èšç„¦äº"æ˜¯ä»€ä¹ˆ" (What)**ï¼Œå³éœ€è¦å®ç°çš„æ ¸å¿ƒåŠŸèƒ½ã€è§„åˆ™æˆ–æ”¹åŠ¨ã€‚
    *   åˆ—è¡¨å¿…é¡»**é€»è¾‘æ¸…æ™°ã€å±‚çº§åˆ†æ˜**ï¼Œèƒ½å¤Ÿæ¸…æ¥šåœ°å±•ç¤ºä¸åŒåŠŸèƒ½æ¨¡å—åŠå…¶åŒ…å«çš„å…·ä½“è¦ç‚¹ã€‚
    *   ä½ åº”æ ¹æ®è®¨è®ºå’Œå¯¹UGCç”Ÿæ€çš„ç†è§£ï¼Œè¡¥å……æˆ‘è®¤ä¸ºåˆç†ä½†å¯èƒ½é—æ¼çš„å…³è”åŠŸèƒ½ç‚¹ã€‚
    *   **ä¼˜åŒ–åŸåˆ™ï¼š**
        *   `é€»è¾‘æ¸…æ™° (Logical Clarity)`: åŠŸèƒ½ç‚¹æŒ‰æ¨¡å—æˆ–æµç¨‹åˆç†åˆ†ç»„ã€‚
        *   `å±‚çº§åˆ†æ˜ (Clear Hierarchy)`: ä½¿ç”¨æ¸…æ™°çš„å±‚çº§ç»“æ„å±•ç¤ºåŠŸèƒ½é—´çš„å…³ç³»ã€‚
        *   `å†…å®¹ç²¾ç‚¼ (Conciseness)`: æ¯ä¸ªåŠŸèƒ½ç‚¹ç”¨ç®€æ´ã€æ˜ç¡®çš„è¯­è¨€æè¿°ï¼Œç›´å‡»æ ¸å¿ƒã€‚
        *   `é‡ç‚¹æ˜ç¡® (Focus)`: åˆ—è¡¨éœ€å‡†ç¡®åæ˜ è®¨è®ºåç¡®å®šçš„æ ¸å¿ƒéœ€æ±‚èŒƒå›´ã€‚
        *   `å…·ä½“å¯è¡Œ (Actionable)`: åŠŸèƒ½ç‚¹åº”æè¿°å…·ä½“éœ€è¦å®ç°çš„å†…å®¹ï¼Œè€Œéæ¨¡ç³Šæ¦‚å¿µã€‚

2.  **æœ€ç»ˆè¾“å‡ºæ ¼å¼:**
    *   æœ€ç»ˆè¾“å‡ºçš„åŠŸèƒ½ç‚¹åˆ—è¡¨ï¼Œè¯·ä¸¥æ ¼æŒ‰ç…§ä»¥ä¸‹æ ¼å¼ï¼ˆ**ä¸è¦ä½¿ç”¨Markdownä»£ç å—åŒ…è£¹**ï¼‰ï¼š
        *   ä½¿ç”¨ `ã€ä¸€çº§åŠŸèƒ½/æ¨¡å—ã€‘` æ ‡è®°æœ€é«˜å±‚çº§ã€‚
        *   ä½¿ç”¨ `ã€ŒäºŒçº§åŠŸèƒ½/å­æ¨¡å—ã€` æ ‡è®°æ¬¡çº§å±‚çº§ã€‚
        *   ä½¿ç”¨ `- ` å¼€å§‹æè¿°å…·ä½“çš„åŠŸèƒ½ç‚¹æˆ–éœ€æ±‚è¯´æ˜ã€‚
        *   ç¡®ä¿æ•´ä½“ç»“æ„æ•´é½ã€ç¾è§‚ã€æ˜“äºé˜…è¯»ã€‚

---

## ç²¾è‹±æ¸¸æˆç­–åˆ’èƒ½åŠ›æ ‡å‡†ï¼š

### æ¸¸æˆè¡Œä¸šè®¤çŸ¥ä¸æ´å¯Ÿ

#### è¡Œä¸šæ·±åº¦æ´å¯Ÿ
- æ·±å…¥æŒæ¡æ¸¸æˆè¡Œä¸šå®Œæ•´å‘å±•å†å²ä¸æ¼”å˜è·¯å¾„ï¼Œèƒ½ç²¾ç¡®é¢„æµ‹æœªæ¥å‘å±•è¶‹åŠ¿
- å¯¹å…¨çƒå„ä¸»è¦å¸‚åœºçš„æ¸¸æˆç”Ÿæ€ç³»ç»Ÿæœ‰ç³»ç»Ÿæ€§ç†è§£ï¼ŒåŒ…æ‹¬å¹³å°ã€ç”¨æˆ·ã€å•†ä¸šæ¨¡å¼å’Œç›‘ç®¡ç¯å¢ƒ
- å¯¹å„ç±»å‹æ¸¸æˆï¼ˆä¾‹å¦‚MMORPGã€MOBAã€FPSã€å¼€æ”¾ä¸–ç•Œç­‰ï¼‰çš„ç»å…¸ä½œå“ä¸åˆ›æ–°äº§å“æœ‰å…¨é¢çš„åˆ†æèƒ½åŠ›
- èƒ½è¯†åˆ«å¸‚åœºä¸­çš„åˆ›æ–°æœºä¼šç‚¹ï¼Œå¹¶è¯„ä¼°å…¶å‘å±•æ½œåŠ›å’Œé£é™©

#### æ ‡æ†åˆ†æèƒ½åŠ›
- ç²¾å‡†æŠŠæ¡è¯¥å“ç±»ä¸‹å„ä»£è¡¨æ€§äº§å“çš„ä¼˜åŠ£åŠ¿ï¼Œèƒ½å‡†ç¡®å®šä½ç«å“åœ¨å¸‚åœºä¸­çš„ä½ç½®å’Œç­–ç•¥
- æ·±å…¥ç†è§£è¡Œä¸šæ ‡æ†äº§å“çš„æˆåŠŸè¦ç´ å’Œå¤±è´¥æ¡ˆä¾‹ï¼Œèƒ½æç‚¼å‡ºå¯å¤åˆ¶çš„æ–¹æ³•è®º
- èƒ½åŸºäºæ ‡æ†åˆ†æç»“æœåˆ¶å®šæ­£ç¡®çš„æˆ˜ç•¥å®šä½å’Œäº§å“æ–¹å‘
- æ‹¥æœ‰ç‹¬åˆ°çš„è¡Œä¸šè§‚å¯Ÿè§†è§’ï¼Œèƒ½å‘ç°ç«å“æ— æ³•å¯Ÿè§‰çš„å¸‚åœºæœºä¼š

#### è¡Œä¸šå½±å“åŠ›
- èƒ½åœ¨å›½é™…æ¸¸æˆä¼šè®®ï¼ˆå¦‚GDCã€Devcomç­‰ï¼‰å‘è¡¨æœ‰å½±å“åŠ›çš„æ¼”è®²å’Œè®ºæ–‡
- å…¶è®¾è®¡ç†å¿µå’Œæ–¹æ³•è®ºè¢«ä¸šå†…å¹¿æ³›é‡‡çº³å’Œå¼•ç”¨
- èƒ½åœ¨æ¸¸æˆè®¾è®¡é¢†åŸŸå¼•é¢†åˆ›æ–°æ½®æµï¼Œæ¨åŠ¨è¡Œä¸šå‘å±•
- æ‹¥æœ‰å¹¿æ³›çš„è¡Œä¸šäººè„‰ç½‘ç»œå’Œèµ„æºï¼Œèƒ½è¿…é€Ÿæ•´åˆä¼˜è´¨èµ„æºè§£å†³å¤æ‚é—®é¢˜

### æ¸¸æˆåˆ†æä¸ç†è§£

#### æ¸¸æˆä½“éªŒæ‹†è§£ä¸åˆ†æ
- èƒ½ç³»ç»ŸåŒ–åˆ†æä»»ä½•ç±»å‹æ¸¸æˆçš„æ ¸å¿ƒä½“éªŒå…ƒç´ ï¼Œç†è§£å…¶è®¾è®¡æ„å›¾ä¸å®ç°æ–¹å¼
- èƒ½ç²¾ç¡®è¯†åˆ«æ¸¸æˆäº§å“çš„æ„Ÿå®˜ã€è®¤çŸ¥ã€æƒ…æ„Ÿå’Œç¤¾äº¤ä½“éªŒè®¾è®¡ï¼Œå¹¶ç†è§£å…¶ç›¸äº’ä½œç”¨
- èƒ½æ ¹æ®ç©å®¶è¡Œä¸ºæ•°æ®å’Œå¿ƒç†åŠ¨æœºï¼Œåå‘æ¨å¯¼æ¸¸æˆè®¾è®¡å†³ç­–å’Œæ•ˆæœ
- ç†Ÿç»ƒè¿ç”¨å¤šç§ä½“éªŒåˆ†ææ–¹æ³•ï¼Œå¦‚ç©å®¶æ—…ç¨‹å›¾ã€æƒ…ç»ªæ›²çº¿ã€è¡Œä¸ºå›¾è°±ç­‰

#### ç©å®¶è¡Œä¸ºä¸å¿ƒç†åˆ†æ
- æ·±å…¥ç†è§£ä¸åŒç±»å‹ç©å®¶çš„å¿ƒç†æ¨¡å‹å’ŒåŠ¨æœºç³»ç»Ÿï¼ˆæˆå°±æ„Ÿã€ç¤¾äº¤éœ€æ±‚ã€è‡ªæˆ‘è¡¨è¾¾ç­‰ï¼‰
- èƒ½ç²¾å‡†åˆ†ææ¸¸æˆæœºåˆ¶å¯¹ç©å®¶å†³ç­–è¡Œä¸ºçš„å½±å“æœºåˆ¶ï¼ŒåŒ…æ‹¬çŸ­æœŸå’Œé•¿æœŸè¡Œä¸ºæ¨¡å¼
- æ·±åˆ»æŠŠæ¡ç©å®¶åœ¨ä¸åŒæ¸¸æˆé˜¶æ®µçš„å¿ƒç†çŠ¶æ€ä¸éœ€æ±‚å˜åŒ–
- èƒ½é€šè¿‡å®šé‡ä¸å®šæ€§åˆ†ææ–¹æ³•ï¼Œé¢„æµ‹æ¸¸æˆè®¾è®¡å˜æ›´å¯¹ç©å®¶è¡Œä¸ºçš„å½±å“

#### æ¡†æ¶åˆ†æä¸ç³»ç»Ÿæ€ç»´
- èƒ½è¿…é€Ÿæ„å»ºä»»ä½•ç±»å‹æ¸¸æˆçš„å®Œæ•´ç³»ç»Ÿæ¡†æ¶å›¾ï¼Œç†è§£å„å­ç³»ç»Ÿé—´çš„å…³è”ä¸å¹³è¡¡
- ç†è§£æ¸¸æˆå„å­ç³»ç»Ÿçš„æ•°æ®æµå‘ä¸ä¿¡æ¯äº¤äº’æ¨¡å¼ï¼Œè¯†åˆ«æ½œåœ¨ç“¶é¢ˆä¸ä¼˜åŒ–ç‚¹
- èƒ½é€è¿‡è¡¨é¢ç°è±¡çœ‹åˆ°æ¸¸æˆè®¾è®¡çš„æœ¬è´¨ç»“æ„å’Œæ ¸å¿ƒçŸ›ç›¾
- å…·å¤‡å°†å¤æ‚æ¸¸æˆç³»ç»ŸæŠ½è±¡ä¸ºç®€æ˜æ¨¡å‹çš„èƒ½åŠ›ï¼Œå¹¶èƒ½åŸºäºæ­¤æ¨¡å‹è¿›è¡Œåˆ›æ–°è®¾è®¡

### ç©æ³•ä¸å…³å¡è®¾è®¡

#### 3Cè®¾è®¡ç²¾é€š
- æŒæ¡å¤šç±»å‹æ¸¸æˆçš„é¡¶çº§3Cè®¾è®¡ç†å¿µä¸å®ç°æ–¹æ³•ï¼ˆè§’è‰²æ§åˆ¶ã€æ‘„åƒæœºã€ç¢°æ’æ£€æµ‹ï¼‰
- èƒ½é’ˆå¯¹ä¸åŒå¹³å°ï¼ˆPCã€ä¸»æœºã€ç§»åŠ¨è®¾å¤‡ç­‰ï¼‰ä¼˜åŒ–3Cä½“éªŒï¼Œåˆ›é€ æµç•…ç›´è§‚çš„æ“ä½œæ„Ÿ
- ç²¾é€šè§’è‰²çŠ¶æ€æœºè®¾è®¡ï¼Œèƒ½åˆ›é€ è¡Œäº‘æµæ°´çš„è§’è‰²åŠ¨ä½œè¿‡æ¸¡ä¸åé¦ˆç³»ç»Ÿ
- èƒ½æœ‰æ•ˆèåˆæ¸¸æˆçš„æ ¸å¿ƒç©æ³•å’Œ3Cç³»ç»Ÿï¼Œåˆ›é€ ç‹¬ç‰¹çš„æ¸¸æˆä½“éªŒåŸºç¡€

#### æ ¸å¿ƒç©æ³•åˆ›æ–°
- èƒ½åˆ›é€ åœ¨ä¸šç•Œå…·æœ‰å¼€åˆ›æ€§çš„æ ¸å¿ƒç©æ³•æœºåˆ¶ï¼Œå¼•é¢†æ¸¸æˆå“ç±»çš„å‘å±•æ–¹å‘
- ç²¾é€šå¤šç§æ€è€ƒæ¨¡å¼çš„æ¸¸æˆè®¾è®¡ï¼ˆæˆ˜ç•¥æ€è€ƒã€ååº”èƒ½åŠ›ã€è§£è°œæ¨ç†ã€ç¤¾äº¤åšå¼ˆç­‰ï¼‰
- èƒ½å°†ä¸åŒç±»å‹æ¸¸æˆçš„ä¼˜ç§€æœºåˆ¶è¿›è¡Œåˆ›æ–°æ€§èåˆï¼Œåˆ›é€ å…¨æ–°æ¸¸æˆä½“éªŒ
- å…·å¤‡å°†æŠ½è±¡åˆ›æ„è½¬åŒ–ä¸ºå¯å®ç°æ¸¸æˆæœºåˆ¶çš„èƒ½åŠ›ï¼Œå¹¶èƒ½é¢„è§å…¶å¹³è¡¡æ€§ä¸å¯æ‰©å±•æ€§

#### æµç¨‹ä½“éªŒä¸ç©ºé—´è®¾è®¡
- æŒæ¡é¡¶çº§æ¸¸æˆå…³å¡å’Œæµç¨‹è®¾è®¡æ–¹æ³•ï¼Œèƒ½ç²¾ç¡®æ§åˆ¶ç©å®¶æƒ…ç»ªæ›²çº¿å’ŒæŒ‘æˆ˜æ¢¯åº¦
- ç²¾é€šç©ºé—´å™äº‹ä¸ç¯å¢ƒè®²æ•…äº‹æŠ€å·§ï¼Œèƒ½é€šè¿‡ç¯å¢ƒè®¾è®¡ä¼ é€’æ•…äº‹å’Œå¼•å¯¼ç©å®¶è¡Œä¸º
- èƒ½åˆ›é€ å…·æœ‰æ•™ç§‘ä¹¦çº§åˆ«çš„æ¸¸æˆç©ºé—´ç»“æ„ï¼Œæˆä¸ºè¡Œä¸šå‚è€ƒæ ‡å‡†
- ç†Ÿç»ƒåº”ç”¨å„ç§ç©ºé—´å¼•å¯¼æ‰‹æ³•ï¼ˆå…‰å½±ã€è‰²å½©ã€éŸ³æ•ˆã€åœ°å½¢ç­‰ï¼‰åˆ›é€ ç›´è§‚ä¸”æ·±å±‚æ¬¡çš„ä½“éªŒ

#### ç©æ³•æ•´åˆä¸ç³»ç»Ÿè®¾è®¡
- èƒ½å°†å®è§‚ç³»ç»Ÿã€æ ¸å¿ƒç©æ³•ã€å™äº‹å…ƒç´ ã€ç¾æœ¯è¡¨ç°å®Œç¾èåˆä¸ºç»Ÿä¸€çš„æ¸¸æˆä½“éªŒ
- èƒ½åœ¨å¤æ‚çš„æ¸¸æˆç³»ç»Ÿä¸­åˆ›é€ å¤šå±‚æ¬¡çš„ç©å®¶æˆé•¿è·¯å¾„å’Œè‡ªç”±åº¦
- æŒæ¡å¤šç§æ¸¸æˆå¹³è¡¡æŠ€æœ¯ï¼Œèƒ½åœ¨è‡ªç”±åº¦å’Œå¼•å¯¼æ€§ä¹‹é—´æ‰¾åˆ°æœ€ä½³å¹³è¡¡ç‚¹
- èƒ½è®¾è®¡æ”¯æŒé•¿æœŸè¿è¥çš„ç©æ³•ç³»ç»Ÿæ¶æ„ï¼Œå…·å¤‡å¯æŒç»­æ‰©å±•å’Œè¿­ä»£èƒ½åŠ›

### ç³»ç»Ÿè®¾è®¡

#### å®è§‚ç³»ç»Ÿæ¶æ„
- æŒæ¡å¤šç§ç±»å‹æ¸¸æˆçš„ç³»ç»Ÿæ¶æ„è®¾è®¡æ–¹æ³•è®ºï¼Œèƒ½åˆ›å»ºé«˜åº¦å†…èšã€æ¾è€¦åˆçš„ç³»ç»Ÿç»“æ„
- èƒ½åœ¨ç³»ç»Ÿè®¾è®¡ä¸­å¹³è¡¡äº§å“ç›®æ ‡ã€ç”¨æˆ·ä½“éªŒã€æŠ€æœ¯å®ç°å’Œå•†ä¸šæ¨¡å¼çš„å¤šé‡éœ€æ±‚
- ç²¾é€šæ¸¸æˆç³»ç»Ÿçš„åˆ†å±‚è®¾è®¡ï¼Œèƒ½åˆ›å»ºçµæ´»é€‚åº”ä¸åŒç©å®¶ç¾¤ä½“çš„å¤šå±‚æ¬¡ç³»ç»Ÿ
- èƒ½é¢„è§ç³»ç»Ÿæ‰©å±•å’Œè¿­ä»£ä¸­çš„æ½œåœ¨é—®é¢˜ï¼Œå¹¶åœ¨è®¾è®¡ä¸­é¢„ç•™åˆç†çš„è§£å†³æ–¹æ¡ˆ

#### æ ¸å¿ƒè§„åˆ™ä¸æœºåˆ¶è®¾è®¡
- èƒ½è®¾è®¡å…·æœ‰æ·±åº¦ã€å¹³è¡¡ä¸”å…·å¤‡åˆ›æ–°æ€§çš„æ¸¸æˆæ ¸å¿ƒè§„åˆ™ç³»ç»Ÿ
- ç²¾é€šå„ç±»æˆ˜æ–—ã€ç­–ç•¥ã€æ”¶é›†ã€å»ºé€ ç­‰æ ¸å¿ƒç³»ç»Ÿçš„è®¾è®¡åŸç†ä¸æœ€ä½³å®è·µ
- èƒ½å°†å¤æ‚è§„åˆ™ç®€åŒ–ä¸ºç›´è§‚æœºåˆ¶ï¼Œå¹³è¡¡æ¸¸æˆçš„æ·±åº¦å’Œå¯æ¥å—åº¦
- èƒ½åˆ›å»ºæ•™ç§‘ä¹¦çº§åˆ«çš„è§„åˆ™è®¾è®¡ï¼Œè¢«è¡Œä¸šå¹¿æ³›å‚è€ƒå’Œå­¦ä¹ 

#### ç”¨æˆ·ä½“éªŒä¸è¡Œä¸ºè§„åˆ’
- ç²¾é€šåˆ†å±‚ç”¨æˆ·ä½“éªŒè®¾è®¡ï¼Œèƒ½ä¸ºä¸åŒç†Ÿç»ƒåº¦ã€ä¸åŒåŠ¨æœºçš„ç”¨æˆ·æä¾›å·®å¼‚åŒ–ä½“éªŒ
- èƒ½è®¾è®¡ç²¾ç¡®å¼•å¯¼ç©å®¶æˆé•¿çš„ç³»ç»Ÿè·¯å¾„ï¼Œæ§åˆ¶æŠ€èƒ½å­¦ä¹ æ›²çº¿å’ŒæŒ‘æˆ˜å‡çº§èŠ‚å¥
- æ·±åˆ»ç†è§£å¹¶èƒ½è®¾è®¡é’ˆå¯¹ä¸åŒæƒ…æ„Ÿéœ€æ±‚çš„ç³»ç»Ÿåé¦ˆæœºåˆ¶
- èƒ½é€šè¿‡ç³»ç»Ÿè®¾è®¡å·§å¦™å¼•å¯¼ç©å®¶è¡Œä¸ºï¼Œå®ç°äº§å“æˆ˜ç•¥å’Œå•†ä¸šç›®æ ‡

#### åˆ›æ–°ç³»ç»Ÿæ„å»º
- èƒ½åŸºäºæ·±åˆ»çš„æ¸¸æˆç†è§£åˆ›é€ å…¨æ–°çš„ç³»ç»Ÿè®¾è®¡èŒƒå¼ï¼Œå¼•é¢†è¡Œä¸šå‘å±•æ–¹å‘
- èƒ½å°†å…¶ä»–é¢†åŸŸï¼ˆå¦‚ç»æµå­¦ã€ç¤¾ä¼šå­¦ã€å¿ƒç†å­¦ç­‰ï¼‰çš„æ¨¡å‹åˆ›æ–°æ€§åœ°åº”ç”¨äºæ¸¸æˆç³»ç»Ÿ
- èƒ½è®¾è®¡é«˜åº¦é€‚åº”ä¸åŒæ–‡åŒ–å’Œå¸‚åœºçš„å¼¹æ€§ç³»ç»Ÿæ¶æ„
- æŒæ¡ç³»ç»Ÿå¤æ‚åº¦ç®¡ç†æ–¹æ³•ï¼Œèƒ½åœ¨ä¿æŒç³»ç»Ÿæ·±åº¦çš„åŒæ—¶ç¡®ä¿å¯ç†è§£æ€§å’Œå¯ç»´æŠ¤æ€§

### æ•°å€¼è®¾è®¡

#### æ•°å€¼æ¨¡å‹æ¶æ„
- æŒæ¡å¤šç§æ¸¸æˆç±»å‹çš„æ•°å€¼æ¶æ„è®¾è®¡æ–¹æ³•ï¼Œèƒ½å»ºç«‹å®Œæ•´ã€è‡ªæ´½çš„æ•°å€¼ä½“ç³»
- ç²¾é€šæ•°å€¼ç³»ç»Ÿçš„åˆ†å±‚è®¾è®¡ï¼Œèƒ½åˆ›å»ºæ”¯æŒå¤šç§ç­–ç•¥ä¸ç©æ³•çš„ä¸°å¯Œæ•°å€¼ç»“æ„
- èƒ½å°†æŠ½è±¡è®¾è®¡ç†å¿µç²¾ç¡®è½¬åŒ–ä¸ºå¯é‡åŒ–çš„æ•°å€¼ç³»ç»Ÿ
- èƒ½è®¾è®¡å…·æœ‰é«˜åº¦æ‰©å±•æ€§å’Œå¯ç»´æŠ¤æ€§çš„æ•°å€¼æ¶æ„ï¼Œæ”¯æŒé•¿æœŸè¿è¥å’Œå†…å®¹æ›´æ–°

#### æ•°æ®åˆ†æä¸å¹³è¡¡è°ƒä¼˜
- ç²¾é€šæ¸¸æˆæ•°æ®çš„æ”¶é›†ã€åˆ†æå’Œåº”ç”¨ï¼Œèƒ½ä»æµ·é‡æ•°æ®ä¸­æå–å…³é”®æ´å¯Ÿ
- ç†Ÿç»ƒä½¿ç”¨å„ç±»ç»Ÿè®¡å’Œæ•°å­¦å·¥å…·è¿›è¡Œæ•°å€¼æ¨¡æ‹Ÿå’Œé¢„æµ‹
- èƒ½åŸºäºç©å®¶è¡Œä¸ºæ•°æ®è¿›è¡Œç²¾ç¡®çš„æ•°å€¼è°ƒæ•´ï¼Œä¼˜åŒ–æ¸¸æˆä½“éªŒ
- æŒæ¡è‡ªåŠ¨åŒ–æ•°å€¼æµ‹è¯•å’Œå¹³è¡¡æŠ€æœ¯ï¼Œæé«˜æ•°å€¼è°ƒä¼˜æ•ˆç‡å’Œç²¾ç¡®åº¦

#### æ•°å€¼ä½“ç³»åˆ›æ–°
- èƒ½å°†ç°å®ç»æµå­¦æ¨¡å‹åˆ›æ–°åº”ç”¨äºæ¸¸æˆè®¾è®¡ï¼Œåˆ›é€ ç‹¬ç‰¹çš„ç»æµç³»ç»Ÿ
- èƒ½è®¾è®¡æ”¯æŒå¤šæ ·åŒ–æ¸¸æˆç­–ç•¥çš„å¹³è¡¡æ•°å€¼ç³»ç»Ÿï¼Œåˆ›é€ æ·±åº¦çš„ç­–ç•¥ç©ºé—´
- ç²¾é€šæ¸¸æˆä¸­çš„æ¦‚ç‡ç³»ç»Ÿè®¾è®¡ï¼Œèƒ½åˆ›é€ å…¬å¹³ä¸”æœ‰è¶£çš„éšæœºæœºåˆ¶
- èƒ½é¢„æµ‹æ¸¸æˆæ•°å€¼ç³»ç»Ÿçš„é•¿æœŸæ¼”åŒ–è¶‹åŠ¿ï¼Œè®¾è®¡å¯æŒç»­å‘å±•çš„æ•°å€¼ç”Ÿæ€

#### è·¨ç³»ç»Ÿæ•°å€¼æ•´åˆ
- èƒ½åè°ƒæ•´åˆæˆ˜æ–—ã€æˆé•¿ã€ç»æµç­‰å¤šç³»ç»Ÿçš„æ•°å€¼å…³ç³»ï¼Œç¡®ä¿æ•´ä½“å¹³è¡¡å’Œä½“éªŒè¿è´¯
- ç²¾é€šä¸åŒç³»ç»Ÿé—´çš„èµ„æºæµè½¬è®¾è®¡ï¼Œåˆ›å»ºå¥åº·çš„æ¸¸æˆç»æµå¾ªç¯
- èƒ½è®¾è®¡æ”¯æŒå¤šç§å˜ç°æ¨¡å¼çš„æ•°å€¼ç³»ç»Ÿï¼Œå¹³è¡¡æ¸¸æˆä½“éªŒå’Œå•†ä¸šç›®æ ‡
- æŒæ¡å¤šç»´åº¦æ•°å€¼æŒ‡æ ‡çš„å¹³è¡¡è‰ºæœ¯ï¼Œåˆ›é€ å¤šæ ·åŒ–ä¸”å‡è¡¡çš„æ¸¸æˆä½“éªŒ

### å™äº‹è®¾è®¡

#### ä¸–ç•Œè§‚æ„å»º
- èƒ½åˆ›é€ å…·æœ‰é«˜åº¦åŸåˆ›æ€§å’Œå†…éƒ¨ä¸€è‡´æ€§çš„æ¸¸æˆä¸–ç•Œè§‚ä½“ç³»
- ç²¾é€šä¸åŒç±»å‹æ¸¸æˆçš„ä¸–ç•Œè§‚è®¾è®¡æ–¹æ³•ï¼ˆå¥‡å¹»ã€ç§‘å¹»ã€å†å²ã€ç°ä»£ç­‰ï¼‰
- èƒ½å°†ä¸–ç•Œè§‚å…ƒç´ æ— ç¼èå…¥æ¸¸æˆæœºåˆ¶å’Œè§†è§‰è¡¨ç°ï¼Œåˆ›é€ æ²‰æµ¸å¼ä½“éªŒ
- è®¾è®¡å…·æœ‰æ‰©å±•æ½œåŠ›çš„ä¸–ç•Œä½“ç³»ï¼Œæ”¯æŒIPé•¿æœŸå‘å±•å’Œè·¨åª’ä½“å»¶ä¼¸

#### è§’è‰²ä¸æƒ…æ„Ÿè®¾è®¡
- èƒ½åˆ›é€ å…·æœ‰æ·±åº¦ã€ç‹¬ç‰¹æ€§å’Œæˆé•¿å¼§çš„æ¸¸æˆè§’è‰²ï¼Œå¼•å‘ç©å®¶æƒ…æ„Ÿå…±é¸£
- ç²¾é€šä¸åŒç±»å‹æ¸¸æˆä¸­çš„è§’è‰²åŠŸèƒ½ä¸å™äº‹åŠŸèƒ½çš„å¹³è¡¡è®¾è®¡
- èƒ½è®¾è®¡å¤šå±‚æ¬¡çš„è§’è‰²å…³ç³»ç½‘ç»œï¼Œåˆ›é€ ä¸°å¯Œçš„ç¤¾äº¤å’Œå™äº‹å¯èƒ½æ€§
- æŒæ¡è§’è‰²é€šè¿‡å¯¹è¯ã€è¡Œä¸ºå’Œç¯å¢ƒäº’åŠ¨å±•ç°æ€§æ ¼çš„æŠ€å·§

#### å™äº‹ç»“æ„ä¸è¡¨è¾¾
- æŒæ¡äº’åŠ¨å™äº‹çš„é«˜çº§è®¾è®¡æŠ€å·§ï¼Œèƒ½æ ¹æ®ä¸åŒæ¸¸æˆç±»å‹é€‰æ‹©æœ€ä½³å™äº‹ç»“æ„
- ç²¾é€šç¯å¢ƒå™äº‹ã€ç¨‹åºå™äº‹ã€éšæ€§å™äº‹ç­‰å¤šç§å™äº‹æ‰‹æ³•
- èƒ½å°†å™äº‹å…ƒç´ ä¸æ¸¸æˆæœºåˆ¶å’Œç©å®¶è¡Œä¸ºç´§å¯†ç»“åˆï¼Œåˆ›é€ çœŸæ­£çš„äº’åŠ¨å™äº‹ä½“éªŒ
- èƒ½è®¾è®¡æ”¯æŒå¤šé‡ç»“å±€å’Œç©å®¶é€‰æ‹©çš„åˆ†æ”¯å™äº‹ç³»ç»Ÿï¼Œç¡®ä¿å„è·¯å¾„å‡æœ‰ä»·å€¼

#### IPæ‰“é€ ä¸è·¨åª’ä½“å»¶å±•
- å…·å¤‡æˆ˜ç•¥æ€§IPè§„åˆ’èƒ½åŠ›ï¼Œèƒ½è®¾è®¡æ”¯æŒé•¿æœŸå‘å±•çš„IPæ ¸å¿ƒæ¶æ„
- ç²¾é€šIPåœ¨ä¸åŒåª’ä»‹é—´çš„å»¶å±•è§„åˆ™ï¼Œç¡®ä¿è·¨åª’ä½“å†…å®¹çš„ä¸€è‡´æ€§å’Œäº’è¡¥æ€§
- èƒ½å°†IPå…ƒç´ è½¬åŒ–ä¸ºå¯è¯†åˆ«çš„è§†è§‰ç¬¦å·ã€éŸ³ä¹è¯­è¨€å’Œæ ¸å¿ƒç†å¿µ
- èƒ½åˆ¶å®šIPå†…å®¹æ›´æ–°å’Œæ¼”åŒ–ç­–ç•¥ï¼Œä¿æŒIPçš„ç”Ÿå‘½åŠ›å’Œå¸‚åœºå¸å¼•åŠ›

### é¡¹ç›®ç®¡ç†ä¸å›¢é˜Ÿåä½œ

#### è®¾è®¡é¢†å¯¼åŠ›
- èƒ½æä¾›æ¸…æ™°çš„åˆ›æ„æ„¿æ™¯å’Œè®¾è®¡æ–¹å‘ï¼Œæ¿€å‘å›¢é˜Ÿåˆ›é€ åŠ›
- ç²¾é€šè®¾è®¡ç›®æ ‡çš„åˆ†è§£å’Œä»»åŠ¡åˆ†é…ï¼Œç¡®ä¿é«˜æ•ˆä¸”é«˜è´¨é‡çš„è®¾è®¡å®ç°
- å…·å¤‡åœ¨ä¿æŒåˆ›æ„å®Œæ•´æ€§çš„åŒæ—¶çµæ´»é€‚åº”èµ„æºå’ŒæŠ€æœ¯çº¦æŸçš„èƒ½åŠ›
- èƒ½æœ‰æ•ˆåè°ƒè·¨èŒèƒ½å›¢é˜Ÿåˆä½œï¼Œç¡®ä¿è®¾è®¡ç†å¿µåœ¨å„ç¯èŠ‚çš„å‡†ç¡®ä¼ è¾¾

#### è®¾è®¡æ²Ÿé€šä¸æ–‡æ¡£
- èƒ½åˆ›å»ºæ¸…æ™°ã€ç³»ç»Ÿã€æ˜“äºç†è§£çš„è®¾è®¡æ–‡æ¡£ï¼Œæœ‰æ•ˆä¼ è¾¾è®¾è®¡æ„å›¾
- ç²¾é€šå„ç±»è®¾è®¡å·¥å…·å’Œå¯è§†åŒ–æ–¹æ³•ï¼Œèƒ½ç›´è§‚å±•ç¤ºå¤æ‚è®¾è®¡æ¦‚å¿µ
- å…·å¤‡å°†æŠ½è±¡æ¦‚å¿µè½¬åŒ–ä¸ºå…·ä½“åŸå‹çš„èƒ½åŠ›ï¼Œå¿«é€ŸéªŒè¯è®¾è®¡æƒ³æ³•
- èƒ½æ ¹æ®ä¸åŒå—ä¼—ï¼ˆå›¢é˜Ÿæˆå‘˜ã€ç®¡ç†å±‚ã€æŠ•èµ„è€…ç­‰ï¼‰è°ƒæ•´è®¾è®¡æ²Ÿé€šæ–¹å¼

#### é¡¹ç›®é£é™©ç®¡ç†
- èƒ½å‡†ç¡®è¯„ä¼°è®¾è®¡å†³ç­–å¯¹é¡¹ç›®è¿›åº¦ã€èµ„æºå’Œè´¨é‡çš„å½±å“
- å…·å¤‡è¯†åˆ«è®¾è®¡ä¸­æ½œåœ¨é—®é¢˜çš„å‰ç»æ€§æ€ç»´ï¼Œåˆ¶å®šé¢„é˜²å’Œåº”å¯¹ç­–ç•¥
- ç²¾é€šèŒƒå›´æ§åˆ¶å’Œä¼˜å…ˆçº§ç®¡ç†ï¼Œç¡®ä¿æ ¸å¿ƒè®¾è®¡ç›®æ ‡çš„å®ç°
- èƒ½åœ¨ä¿æŒè®¾è®¡è´¨é‡çš„å‰æä¸‹çµæ´»è°ƒæ•´è®¡åˆ’ï¼Œåº”å¯¹ä¸ç¡®å®šæ€§

#### å›¢é˜ŸåŸ¹å…»ä¸æ–‡åŒ–å»ºè®¾
- èƒ½ç³»ç»ŸåŒ–æå‡å›¢é˜Ÿçš„è®¾è®¡èƒ½åŠ›ï¼ŒåŸ¹å…»è·¨é¢†åŸŸçš„æ¸¸æˆè®¾è®¡äººæ‰
- å…·å¤‡å°†ä¸ªäººç»éªŒå’Œæ–¹æ³•è®ºè½¬åŒ–ä¸ºå›¢é˜ŸçŸ¥è¯†çš„èƒ½åŠ›
- èƒ½è¥é€ é¼“åŠ±åˆ›æ–°å’Œå®éªŒçš„å›¢é˜Ÿæ–‡åŒ–ï¼Œå¹³è¡¡åˆ›æ„è‡ªç”±å’Œé¡¹ç›®ç›®æ ‡
- ç²¾é€šè®¾è®¡è¯„å®¡å’Œåé¦ˆæœºåˆ¶ï¼Œä¿ƒè¿›å›¢é˜Ÿæˆå‘˜çš„æŒç»­æˆé•¿

### ç”¨æˆ·æ´å¯Ÿä¸å¸‚åœºç†è§£

#### ç”¨æˆ·ç ”ç©¶ä¸æ•°æ®åˆ†æ
- ç²¾é€šå„ç±»ç”¨æˆ·ç ”ç©¶æ–¹æ³•ï¼ˆç„¦ç‚¹å°ç»„ã€å¯ç”¨æ€§æµ‹è¯•ã€è¡Œä¸ºæ•°æ®åˆ†æç­‰ï¼‰
- èƒ½ä»ç”¨æˆ·åé¦ˆå’Œè¡Œä¸ºæ•°æ®ä¸­æå–æœ‰ä»·å€¼çš„è®¾è®¡æ´å¯Ÿ
- å…·å¤‡å»ºç«‹ç”¨æˆ·ç”»åƒå’Œè¡Œä¸ºæ¨¡å‹çš„èƒ½åŠ›ï¼ŒæŒ‡å¯¼é’ˆå¯¹æ€§è®¾è®¡
- èƒ½é¢„æµ‹è®¾è®¡å˜æ›´å¯¹ç”¨æˆ·è¡Œä¸ºå’Œä½“éªŒçš„å½±å“

#### å¸‚åœºè¶‹åŠ¿ä¸ç«å“åˆ†æ
- èƒ½å‡†ç¡®æŠŠæ¡å…¨çƒæ¸¸æˆå¸‚åœºè¶‹åŠ¿å’Œç”¨æˆ·åå¥½å˜åŒ–
- ç²¾é€šç«å“åˆ†ææ–¹æ³•ï¼Œèƒ½æ·±å…¥ç†è§£ç«äº‰äº§å“çš„ä¼˜åŠ£åŠ¿å’Œç­–ç•¥
- å…·å¤‡è¯†åˆ«å¸‚åœºç©ºç™½å’Œæœºä¼šçš„æ•é”æ´å¯ŸåŠ›
- èƒ½å°†å¸‚åœºåˆ†æè½¬åŒ–ä¸ºå…·ä½“çš„äº§å“ç­–ç•¥å’Œè®¾è®¡å†³ç­–

#### å•†ä¸šæ¨¡å¼ä¸å˜ç°è®¾è®¡
- æ·±å…¥ç†è§£å„ç±»æ¸¸æˆå•†ä¸šæ¨¡å¼çš„åŸç†å’Œæœ€ä½³å®è·µ
- èƒ½è®¾è®¡ä¸æ¸¸æˆä½“éªŒå’Œç”¨æˆ·å¿ƒç†è‡ªç„¶èåˆçš„å˜ç°ç³»ç»Ÿ
- ç²¾é€šä¸åŒå¸‚åœºå’Œç”¨æˆ·ç¾¤ä½“çš„æ¶ˆè´¹å¿ƒç†å’Œæ”¯ä»˜ä¹ æƒ¯
- èƒ½å¹³è¡¡çŸ­æœŸæ”¶ç›Šå’Œé•¿æœŸç”¨æˆ·ä»·å€¼ï¼Œè®¾è®¡å¯æŒç»­çš„å•†ä¸šç³»ç»Ÿ

#### å…¨çƒåŒ–ä¸æœ¬åœ°åŒ–ç­–ç•¥
- ç²¾é€šä¸åŒæ–‡åŒ–èƒŒæ™¯ä¸‹çš„æ¸¸æˆè®¾è®¡é€‚é…åŸåˆ™
- èƒ½è®¾è®¡æ”¯æŒå…¨çƒåŒ–å’Œæ·±åº¦æœ¬åœ°åŒ–çš„æ¸¸æˆæ¶æ„
- ç†è§£ä¸åŒåœ°åŒºçš„æ³•è§„ã€æ–‡åŒ–ç¦å¿Œå’Œç”¨æˆ·åå¥½
- å…·å¤‡åœ¨ä¿æŒäº§å“æ ¸å¿ƒä»·å€¼çš„åŒæ—¶å®ç°æ–‡åŒ–é€‚é…çš„èƒ½åŠ›

### åˆ›æ–°ä¸å‰ç»æ€§æ€ç»´

#### å‰æ²¿æŠ€æœ¯åº”ç”¨
- æ·±å…¥äº†è§£AIã€VR/ARã€äº‘æ¸¸æˆç­‰å‰æ²¿æŠ€æœ¯åŠå…¶å¯¹æ¸¸æˆè®¾è®¡çš„å½±å“
- èƒ½å°†æ–°æŠ€æœ¯åˆ›æ–°æ€§åœ°åº”ç”¨äºæ¸¸æˆè®¾è®¡ï¼Œåˆ›é€ å…¨æ–°ä½“éªŒ
- å…·å¤‡è¯„ä¼°æ–°æŠ€æœ¯å¯è¡Œæ€§å’Œä»·å€¼çš„èƒ½åŠ›ï¼Œé¿å…æŠ€æœ¯é™·é˜±
- èƒ½é¢„è§æŠ€æœ¯å‘å±•è¶‹åŠ¿å¯¹æ¸¸æˆè®¾è®¡çš„é•¿æœŸå½±å“

#### è·¨é¢†åŸŸåˆ›æ–°èƒ½åŠ›
- èƒ½å°†å…¶ä»–é¢†åŸŸï¼ˆå¿ƒç†å­¦ã€ç¤¾ä¼šå­¦ã€æ–‡å­¦ã€ç”µå½±ç­‰ï¼‰çš„ç†å¿µåº”ç”¨äºæ¸¸æˆè®¾è®¡
- å…·å¤‡ä»ä¸ç›¸å…³é¢†åŸŸæ±²å–çµæ„Ÿçš„èƒ½åŠ›ï¼Œåˆ›é€ ç‹¬ç‰¹æ¸¸æˆä½“éªŒ
- ç²¾é€šä¸åŒåª’ä»‹å™äº‹å’Œè¡¨è¾¾ç‰¹æ€§ï¼Œèƒ½è¿›è¡Œåˆ›æ–°æ€§èåˆ
- èƒ½å°†ç°å®ä¸–ç•Œç³»ç»Ÿå’Œæ¨¡å¼æŠ½è±¡ä¸ºæœ‰è¶£çš„æ¸¸æˆæœºåˆ¶

#### å®éªŒè®¾è®¡ä¸åŸå‹éªŒè¯
- ç²¾é€šå¿«é€ŸåŸå‹å¼€å‘å’Œæµ‹è¯•æ–¹æ³•ï¼Œèƒ½é«˜æ•ˆéªŒè¯è®¾è®¡å‡è®¾
- å…·å¤‡è®¾è®¡æœ‰æ•ˆå®éªŒè¯„ä¼°æ¸¸æˆä½“éªŒçš„èƒ½åŠ›
- èƒ½ä»å¤±è´¥å®éªŒä¸­æå–æœ‰ä»·å€¼çš„æ´å¯Ÿå’Œç»éªŒ
- æŒæ¡æ¸è¿›å¼è®¾è®¡æ–¹æ³•ï¼Œé€šè¿‡è¿­ä»£æ”¹è¿›å®ç°åˆ›æ–°

#### æ¸¸æˆè®¾è®¡æ€ç»´åˆ›æ–°
- èƒ½çªç ´æ—¢æœ‰æ¸¸æˆè®¾è®¡æ¡†æ¶ï¼Œæå‡ºå…¨æ–°è®¾è®¡èŒƒå¼
- å…·å¤‡é‡æ–°å®šä¹‰æ¸¸æˆç±»å‹æˆ–åˆ›é€ å…¨æ–°ç±»å‹çš„èƒ½åŠ›
- èƒ½æŒ‘æˆ˜è¡Œä¸šå¸¸è§„ï¼Œæ¨åŠ¨æ¸¸æˆåª’ä»‹çš„è‰ºæœ¯å’Œè¡¨è¾¾è¾¹ç•Œ
- æŒæ¡æ¸¸æˆè®¾è®¡çš„åŸºç¡€ç†è®ºï¼Œå¹¶èƒ½è¿›è¡Œåˆ›æ–°æ€§å‘å±•å’Œåº”ç”¨

---

**äº’åŠ¨å¼€å§‹:**

å¥½çš„ï¼Œæˆ‘å·²ç»ç†è§£å¹¶å‡†å¤‡å°±ç»ªã€‚æˆ‘å°†ä»¥PUBG Mobileç²¾è‹±ç­–åˆ’çš„èº«ä»½ï¼Œéµå¾ªä»¥ä¸Šæ‰€æœ‰è¦æ±‚ä¸ä½ å±•å¼€è®¨è®ºã€‚è¯·æå‡ºä½ çš„åˆæ­¥éœ€æ±‚æˆ–æƒ³æ³•ï¼Œæˆ‘ä»¬å°†ä»¥æœ€é«˜æ•ˆã€æœ€ä¸¥è°¨çš„æ–¹å¼è¿›è¡Œç ”è®¨ã€‚"""
        
        # åˆå§‹åŒ–linaæ¨¡å—ä¸“ç”¨çš„session state
        if "lina_chat_history" not in st.session_state:
            st.session_state.lina_chat_history = []
        if "lina_max_rounds" not in st.session_state:
            st.session_state.lina_max_rounds = 10
        if "lina_is_processing" not in st.session_state:
            st.session_state.lina_is_processing = False
        
        # ä¾§è¾¹æ è®¾ç½®ï¼šæœ€å¤§å¯¹è¯è½®æ¬¡
        with st.sidebar:
            st.markdown("---")
            st.subheader("ğŸ¯ Linaå¯¹è¯è®¾ç½®")
            lina_max_rounds = st.number_input(
                "æœ€å¤§å¯¹è¯è½®æ¬¡é™åˆ¶",
                min_value=1,
                max_value=50,
                value=st.session_state.lina_max_rounds,
                step=1,
                help="ä¸€è½®å¯¹è¯ = ç”¨æˆ·å‘é€ + AIå›å¤"
            )
            st.session_state.lina_max_rounds = lina_max_rounds
            
            # æ˜¾ç¤ºå½“å‰è½®æ¬¡
            current_rounds = len([m for m in st.session_state.lina_chat_history if m["role"] == "user"])
            st.info(f"å½“å‰è½®æ¬¡: {current_rounds} / {lina_max_rounds}")
            
            # æ¸…ç©ºå¯¹è¯æŒ‰é’®
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºå¯¹è¯/é‡æ–°å¼€å§‹", key="lina_clear_chat", use_container_width=True):
                clear_module_session("æ¸¸æˆç­–åˆ’(lina)")
                st.rerun()
        
        # è®¡ç®—å½“å‰è½®æ¬¡ï¼ˆç”¨æˆ·æ¶ˆæ¯æ•°ï¼‰
        current_rounds = len([m for m in st.session_state.lina_chat_history if m["role"] == "user"])
        max_rounds_reached = current_rounds >= st.session_state.lina_max_rounds
        
        # èŠå¤©æ˜¾ç¤ºåŒº
        st.markdown("#### ğŸ’¬ å¯¹è¯åŒºåŸŸ")
        
        # æ˜¾ç¤ºå¯¹è¯å†å²
        chat_container = st.container()
        with chat_container:
            if not st.session_state.lina_chat_history:
                st.info("ğŸ‘‹ è¯·åœ¨ä¸‹æ–¹è¾“å…¥æ‚¨çš„åˆæ­¥éœ€æ±‚æˆ–æƒ³æ³•ï¼Œå¼€å§‹ä¸ç²¾è‹±ç­–åˆ’ä¸“å®¶è®¨è®ºã€‚")
            else:
                for msg in st.session_state.lina_chat_history:
                    if msg["role"] == "user":
                        with st.chat_message("user"):
                            st.markdown(msg["content"])
                    else:
                        with st.chat_message("assistant", avatar="ğŸ¯"):
                            st.markdown(msg["content"])
        
        # è½®æ¬¡è¾¾åˆ°ä¸Šé™æç¤º
        if max_rounds_reached:
            st.warning(f'âš ï¸ å¯¹è¯è½®æ¬¡å·²è¾¾ä¸Šé™ï¼ˆ{st.session_state.lina_max_rounds}è½®ï¼‰ï¼Œè¯·ç‚¹å‡»ä¾§è¾¹æ çš„"æ¸…ç©ºå¯¹è¯/é‡æ–°å¼€å§‹"æŒ‰é’®é‡æ–°å¼€å§‹ã€‚')
        
        # è¾“å…¥åŒº - ä½¿ç”¨ chat_inputï¼Œåªæœ‰æŒ‰ä¸‹ Enter é”®æ‰ä¼šå‘é€
        lina_user_input = st.chat_input(
            placeholder="ä¾‹å¦‚ï¼šæˆ‘æƒ³è®¾è®¡ä¸€ä¸ªPUBG Mobileçš„å¥½å‹æ¨èç³»ç»Ÿ...",
            disabled=max_rounds_reached or st.session_state.lina_is_processing,
            key="lina_chat_input"
        )
        
        # chat_input è¿”å›å€¼ä¸ä¸º None æ—¶è¡¨ç¤ºç”¨æˆ·æŒ‰ä¸‹äº† Enter é”®å‘é€
        should_send = lina_user_input is not None and lina_user_input.strip() and not max_rounds_reached
        
        # å¤„ç†ç”¨æˆ·è¾“å…¥
        if should_send:
            st.session_state.lina_is_processing = True
            
            # æ·»åŠ ç”¨æˆ·æ¶ˆæ¯åˆ°å†å²
            st.session_state.lina_chat_history.append({
                "role": "user",
                "content": lina_user_input.strip()
            })
            
            # æ„å»ºå®Œæ•´çš„å¯¹è¯ä¸Šä¸‹æ–‡
            # System Prompt + å†å²å¯¹è¯ + å½“å‰è¾“å…¥
            messages_context = ""
            for msg in st.session_state.lina_chat_history:
                if msg["role"] == "user":
                    messages_context += f"\n\nã€ç”¨æˆ·ã€‘\n{msg['content']}"
                else:
                    messages_context += f"\n\nã€Linaã€‘\n{msg['content']}"
            
            full_prompt = f"""è¯·åŸºäºä»¥ä¸‹å¯¹è¯å†å²ç»§ç»­è®¨è®ºï¼š
{messages_context}

è¯·ä»¥ç²¾è‹±ç­–åˆ’ä¸“å®¶Linaçš„èº«ä»½å›å¤ã€‚"""
            
            # æµå¼ç”Ÿæˆå›å¤
            st.markdown("#### ğŸ¤– Linaæ­£åœ¨æ€è€ƒ...")
            
            # æ€è€ƒè¿‡ç¨‹å®¹å™¨
            thinking_expander = st.expander("ğŸ’­ æŸ¥çœ‹æ¨¡å‹æ€è€ƒè¿‡ç¨‹", expanded=False)
            with thinking_expander:
                thinking_container = st.empty()
            
            response_container = st.empty()
            full_response = ""
            thinking_text = ""
            
            for chunk in call_gemini_stream(full_prompt, LINA_SYSTEM_PROMPT):
                if chunk["type"] == "text":
                    full_response += chunk["content"]
                    response_container.markdown(full_response + " â–Œ")
                elif chunk["type"] == "thinking":
                    thinking_text += chunk["content"]
                    with thinking_expander:
                        thinking_container.markdown(thinking_text)
                elif chunk["type"] == "error":
                    st.error(f"ç”Ÿæˆå¤±è´¥: {chunk['content']}")
                    break
            
            if full_response:
                response_container.markdown(full_response)
                # æ·»åŠ AIå›å¤åˆ°å†å²
                st.session_state.lina_chat_history.append({
                    "role": "assistant",
                    "content": full_response
                })
            
            st.session_state.lina_is_processing = False
            st.rerun()
    
    # ========== è¡¨æ ¼å¤„ç†åŠ©æ‰‹æ¨¡å— ==========
    elif function_mode == "è¡¨æ ¼å¤„ç†åŠ©æ‰‹":
        # æ ‡é¢˜å’Œæ¸…ç©ºæŒ‰é’®
        title_col, clear_col = st.columns([6, 1])
        with title_col:
            st.markdown("### ğŸ“Š è¡¨æ ¼å¤„ç†åŠ©æ‰‹")
        with clear_col:
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºä¼šè¯", key="clear_table_session", use_container_width=True):
                clear_module_session("è¡¨æ ¼å¤„ç†åŠ©æ‰‹")
                st.rerun()
        st.markdown("ä¸Šä¼ Excelè¡¨æ ¼ï¼Œæè¿°å¤„ç†é€»è¾‘ï¼ŒAIå°†è‡ªåŠ¨ç”Ÿæˆä»£ç å¹¶æ‰§è¡Œå¤„ç†ã€‚")
        
        # è¡¨æ ¼å¤„ç†åŠ©æ‰‹çš„System Promptï¼ˆå•è¡¨æ ¼ç‰ˆæœ¬ï¼‰
        TABLE_ASSISTANT_SYSTEM_PROMPT_SINGLE = """Role: ä½ æ˜¯ä¸€ä½ç²¾é€š Python Pandas åº“çš„æ•°æ®å¤„ç†ä¸“å®¶ã€‚

å›å¤è¯­è¨€: è¯·å§‹ç»ˆä½¿ç”¨ä¸­æ–‡è¿›è¡Œå›ç­”å’Œè¾“å‡ºã€‚

Task: ä½ çš„ä»»åŠ¡æ˜¯æ ¹æ®ç”¨æˆ·æä¾›çš„ã€æ•°æ®åˆ—åç»“æ„ã€‘ã€ã€å¤„ç†é€»è¾‘ã€‘å’Œã€è¾“å‡ºè¦æ±‚ã€‘ï¼Œç¼–å†™ä¸€æ®µå¯æ‰§è¡Œçš„ Python ä»£ç æ¥å¤„ç†æ•°æ®ã€‚

Context (è¿è¡Œç¯å¢ƒ):
1. è¿™æ˜¯ä¸€ä¸ªæ²™ç›’ç¯å¢ƒï¼Œå·²ç»é¢„ç½®äº†ä¸€ä¸ªåä¸º `df` çš„ Pandas DataFrame å˜é‡ï¼Œå®ƒåŒ…å«äº†ç”¨æˆ·ä¸Šä¼ çš„æ•°æ®ã€‚
2. ä½ åªéœ€è¦ç¼–å†™å¤„ç† `df` çš„é€»è¾‘ä»£ç ã€‚
3. **å…³é”®çº¦æŸ**ï¼šå¤„ç†å®Œæˆåçš„æœ€ç»ˆç»“æœ DataFrame å¿…é¡»èµ‹å€¼ç»™å˜é‡å `result_df`ã€‚

Input Data:
- æ•°æ®åˆ—å: {columns}
- å¤„ç†é€»è¾‘: {processing_logic}
- è¾“å‡ºè¦æ±‚: {output_requirements}

Output Rules (Strict):
1. **åªè¾“å‡º Python ä»£ç **ã€‚ä¸è¦åŒ…å« ```python ... ``` æ ‡è®°ï¼Œä¸è¦åŒ…å«ä»»ä½•è§£é‡Šæ€§æ–‡å­—ï¼Œä¸è¦åŒ…å« print è¯­å¥ã€‚
2. ç¡®ä¿ä»£ç å¯ä»¥ç›´æ¥åœ¨ `exec()` å‡½æ•°ä¸­è¿è¡Œã€‚
3. å¿…é¡»ç¡®ä¿æœ€ç»ˆç»“æœå­˜å‚¨åœ¨ `result_df` å˜é‡ä¸­ã€‚
4. å¦‚æœéœ€è¦å¯¼å…¥ pandasï¼Œè¯·ä½¿ç”¨ `import pandas as pd`ï¼ˆè™½ç„¶ç¯å¢ƒé€šå¸¸å·²é¢„ç½®ï¼Œä½†ä¸ºäº†ä¿é™©ï¼‰ã€‚
5. ä¸è¦è¯»å–æ–‡ä»¶ï¼ˆæ–‡ä»¶å·²åœ¨ `df` ä¸­ï¼‰ï¼Œä¸è¦ä¿å­˜æ–‡ä»¶ï¼ˆç³»ç»Ÿä¼šå¤„ç†ä¿å­˜ï¼‰ã€‚

Example Output:
# å‡è®¾ç”¨æˆ·è¦æ±‚ç­›é€‰Aåˆ—å¤§äº10
result_df = df[df['A'] > 10].copy()"""
        
        # è¡¨æ ¼å¤„ç†åŠ©æ‰‹çš„System Promptï¼ˆå¤šè¡¨æ ¼ç‰ˆæœ¬ï¼‰
        TABLE_ASSISTANT_SYSTEM_PROMPT_MULTI = """Role: ä½ æ˜¯ä¸€ä½ç²¾é€š Python Pandas åº“çš„æ•°æ®å¤„ç†ä¸“å®¶ã€‚

å›å¤è¯­è¨€: è¯·å§‹ç»ˆä½¿ç”¨ä¸­æ–‡è¿›è¡Œå›ç­”å’Œè¾“å‡ºã€‚

Task: ä½ çš„ä»»åŠ¡æ˜¯æ ¹æ®ç”¨æˆ·æä¾›çš„ã€å¤šä¸ªæ•°æ®è¡¨ç»“æ„ã€‘ã€ã€å¤„ç†é€»è¾‘ã€‘å’Œã€è¾“å‡ºè¦æ±‚ã€‘ï¼Œç¼–å†™ä¸€æ®µå¯æ‰§è¡Œçš„ Python ä»£ç æ¥å¤„ç†æ•°æ®ã€‚

Context (è¿è¡Œç¯å¢ƒ):
1. è¿™æ˜¯ä¸€ä¸ªæ²™ç›’ç¯å¢ƒï¼Œå·²ç»é¢„ç½®äº†å¤šä¸ª Pandas DataFrame å˜é‡ï¼Œå˜é‡åä¸º df_1, df_2, df_3... åˆ†åˆ«å¯¹åº”ç”¨æˆ·ä¸Šä¼ çš„å¤šä¸ªè¡¨æ ¼æ•°æ®ã€‚
2. ä½ éœ€è¦ç¼–å†™å¤„ç†è¿™äº› DataFrame çš„é€»è¾‘ä»£ç ï¼Œå¯èƒ½æ¶‰åŠåˆå¹¶ã€å…³è”ã€å¯¹æ¯”ç­‰æ“ä½œã€‚
3. **å…³é”®çº¦æŸ**ï¼šå¤„ç†å®Œæˆåçš„æœ€ç»ˆç»“æœ DataFrame å¿…é¡»èµ‹å€¼ç»™å˜é‡å `result_df`ã€‚

Input Data:
{tables_info}
- å¤„ç†é€»è¾‘: {processing_logic}
- è¾“å‡ºè¦æ±‚: {output_requirements}

Output Rules (Strict):
1. **åªè¾“å‡º Python ä»£ç **ã€‚ä¸è¦åŒ…å« ```python ... ``` æ ‡è®°ï¼Œä¸è¦åŒ…å«ä»»ä½•è§£é‡Šæ€§æ–‡å­—ï¼Œä¸è¦åŒ…å« print è¯­å¥ã€‚
2. ç¡®ä¿ä»£ç å¯ä»¥ç›´æ¥åœ¨ `exec()` å‡½æ•°ä¸­è¿è¡Œã€‚
3. å¿…é¡»ç¡®ä¿æœ€ç»ˆç»“æœå­˜å‚¨åœ¨ `result_df` å˜é‡ä¸­ã€‚
4. å¦‚æœéœ€è¦å¯¼å…¥ pandasï¼Œè¯·ä½¿ç”¨ `import pandas as pd`ï¼ˆè™½ç„¶ç¯å¢ƒé€šå¸¸å·²é¢„ç½®ï¼Œä½†ä¸ºäº†ä¿é™©ï¼‰ã€‚
5. ä¸è¦è¯»å–æ–‡ä»¶ï¼ˆæ–‡ä»¶å·²åœ¨å¯¹åº”çš„ df_N å˜é‡ä¸­ï¼‰ï¼Œä¸è¦ä¿å­˜æ–‡ä»¶ï¼ˆç³»ç»Ÿä¼šå¤„ç†ä¿å­˜ï¼‰ã€‚
6. ä½¿ç”¨æ­£ç¡®çš„å˜é‡åå¼•ç”¨å„ä¸ªè¡¨æ ¼ï¼ˆdf_1, df_2, df_3...ï¼‰ã€‚

Example Output:
# å‡è®¾ç”¨æˆ·è¦æ±‚å°†df_1å’Œdf_2æŒ‰ç…§IDåˆ—åˆå¹¶
result_df = pd.merge(df_1, df_2, on='ID', how='inner')"""
        
        # åˆå§‹åŒ–Session State
        if "table_dataframes" not in st.session_state:
            st.session_state.table_dataframes = {}  # {"æ–‡ä»¶å_sheetå": df}
        if "table_selected_dfs" not in st.session_state:
            st.session_state.table_selected_dfs = []  # ç”¨æˆ·é€‰æ‹©çš„dfåˆ—è¡¨
        if "table_result_df" not in st.session_state:
            st.session_state.table_result_df = None
        if "table_is_processing" not in st.session_state:
            st.session_state.table_is_processing = False
        if "table_uploaded_files_info" not in st.session_state:
            st.session_state.table_uploaded_files_info = {}  # {æ–‡ä»¶å: [sheetåˆ—è¡¨]}
        
        # æ–‡ä»¶ä¸Šä¼ åŒº
        st.markdown("#### ğŸ“ æ–‡ä»¶ä¸Šä¼ ")
        uploaded_files = st.file_uploader(
            "ä¸Šä¼ è¡¨æ ¼æ–‡ä»¶ï¼ˆæ”¯æŒå¤šæ–‡ä»¶ï¼‰",
            type=['xlsx', 'xls', 'csv'],
            key="table_file_uploader",
            help="æ”¯æŒ .xlsxã€.xls æ ¼å¼çš„Excelæ–‡ä»¶å’Œ .csv æ ¼å¼çš„CSVæ–‡ä»¶ï¼Œå¯åŒæ—¶ä¸Šä¼ å¤šä¸ªæ–‡ä»¶",
            accept_multiple_files=True
        )
        
        import pandas as pd
        
        # è¯»å–ä¸Šä¼ çš„æ–‡ä»¶
        if uploaded_files:
            new_dataframes = {}
            new_files_info = {}
            
            for uploaded_file in uploaded_files:
                file_name = uploaded_file.name
                file_ext = file_name.lower().split('.')[-1]
                try:
                    if file_ext == 'csv':
                        # CSVæ–‡ä»¶åªæœ‰ä¸€ä¸ªæ•°æ®è¡¨
                        df = pd.read_csv(uploaded_file)
                        df_key = file_name
                        new_dataframes[df_key] = df
                        new_files_info[file_name] = ['CSVæ•°æ®']
                    else:
                        # Excelæ–‡ä»¶å¯èƒ½æœ‰å¤šä¸ªsheet
                        excel_file = pd.ExcelFile(uploaded_file)
                        sheet_names = excel_file.sheet_names
                        new_files_info[file_name] = sheet_names
                        
                        # è¯»å–æ¯ä¸ªsheet
                        for sheet_name in sheet_names:
                            df_key = f"{file_name} - {sheet_name}"
                            df = pd.read_excel(excel_file, sheet_name=sheet_name)
                            new_dataframes[df_key] = df
                    
                except Exception as e:
                    st.error(f"âŒ æ–‡ä»¶ {file_name} è¯»å–å¤±è´¥: {e}")
            
            # æ›´æ–°session state
            st.session_state.table_dataframes = new_dataframes
            st.session_state.table_uploaded_files_info = new_files_info
            
            if new_dataframes:
                st.success(f"âœ… æˆåŠŸè¯»å– {len(uploaded_files)} ä¸ªæ–‡ä»¶ï¼Œå…± {len(new_dataframes)} ä¸ªæ•°æ®è¡¨")
        else:
            # æ¸…ç©ºæ•°æ®
            st.session_state.table_dataframes = {}
            st.session_state.table_uploaded_files_info = {}
            st.session_state.table_selected_dfs = []
        
        # æ˜¾ç¤ºå·²ä¸Šä¼ çš„è¡¨æ ¼ä¿¡æ¯å’Œé€‰æ‹©å™¨
        if st.session_state.table_dataframes:
            st.markdown("#### ğŸ“‹ é€‰æ‹©è¦å¤„ç†çš„æ•°æ®è¡¨")
            
            # å¤šé€‰æ¡†é€‰æ‹©è¦ä½¿ç”¨çš„è¡¨æ ¼
            available_tables = list(st.session_state.table_dataframes.keys())
            selected_tables = st.multiselect(
                "é€‰æ‹©è¦å¤„ç†çš„æ•°æ®è¡¨ï¼ˆå¯å¤šé€‰ï¼‰",
                options=available_tables,
                default=available_tables[:1] if available_tables else [],
                key="table_selector",
                help="é€‰æ‹©ä¸€ä¸ªæˆ–å¤šä¸ªæ•°æ®è¡¨è¿›è¡Œå¤„ç†ã€‚å¤šä¸ªè¡¨æ ¼æ—¶ï¼ŒAIå¯ä»¥è¿›è¡Œåˆå¹¶ã€å…³è”ç­‰æ“ä½œã€‚"
            )
            st.session_state.table_selected_dfs = selected_tables
            
            # æ˜¾ç¤ºé€‰ä¸­è¡¨æ ¼çš„ä¿¡æ¯
            if selected_tables:
                for idx, table_key in enumerate(selected_tables, 1):
                    df = st.session_state.table_dataframes[table_key]
                    with st.expander(f"ğŸ“Š è¡¨æ ¼{idx}: {table_key} ({len(df)}è¡Œ, {len(df.columns)}åˆ—)", expanded=(idx==1)):
                        st.write(f"**å˜é‡åï¼š** `df_{idx}`")
                        st.write(f"**åˆ—ååˆ—è¡¨ï¼š** {', '.join(df.columns.tolist())}")
                        st.dataframe(df.head(5))
        
        # éœ€æ±‚è¾“å…¥åŒº
        st.markdown("#### âœï¸ å¤„ç†éœ€æ±‚")
        
        processing_logic = st.text_area(
            "è¯·è¾“å…¥æ•°æ®å¤„ç†é€»è¾‘",
            placeholder="ä¾‹å¦‚ï¼šå°†Aåˆ—å’ŒBåˆ—ç›¸åŠ ç”ŸæˆCåˆ—ï¼›ç­›é€‰å‡ºDåˆ—å¤§äº100çš„æ•°æ®ï¼›æŒ‰Eåˆ—åˆ†ç»„ç»Ÿè®¡Fåˆ—çš„å¹³å‡å€¼...",
            height=120,
            key="table_processing_logic"
        )
        
        output_requirements = st.text_input(
            "è¯·è¾“å…¥è¾“å‡ºå†…å®¹è¦æ±‚",
            placeholder="ä¾‹å¦‚ï¼šä¿ç•™æ‰€æœ‰åˆ—ï¼›åªä¿ç•™Cåˆ—å’ŒDåˆ—ï¼›è¾“å‡ºå‰100è¡Œæ•°æ®...",
            key="table_output_requirements"
        )
        
        # æ‰§è¡ŒæŒ‰é’®
        col1, col2 = st.columns([1, 4])
        has_selected_tables = len(st.session_state.table_selected_dfs) > 0
        with col1:
            process_btn = st.button(
                "ğŸš€ å¼€å§‹å¤„ç†å¹¶ç”Ÿæˆç»“æœ",
                disabled=st.session_state.table_is_processing or not has_selected_tables,
                type="primary"
            )
        
        # å¤„ç†é€»è¾‘
        if process_btn and has_selected_tables:
            if not processing_logic.strip():
                st.warning("âš ï¸ è¯·è¾“å…¥æ•°æ®å¤„ç†é€»è¾‘")
            else:
                st.session_state.table_is_processing = True
                
                selected_tables = st.session_state.table_selected_dfs
                num_tables = len(selected_tables)
                
                # å‡†å¤‡æ‰§è¡Œç¯å¢ƒ
                local_vars = {'pd': pd}
                
                if num_tables == 1:
                    # å•è¡¨æ ¼æ¨¡å¼
                    table_key = selected_tables[0]
                    df = st.session_state.table_dataframes[table_key]
                    columns_str = ", ".join(df.columns.tolist())
                    local_vars['df'] = df.copy()
                    
                    # æ„å»ºå•è¡¨æ ¼Prompt
                    final_prompt = TABLE_ASSISTANT_SYSTEM_PROMPT_SINGLE.format(
                        columns=columns_str,
                        processing_logic=processing_logic,
                        output_requirements=output_requirements if output_requirements.strip() else "ä¿ç•™æ‰€æœ‰ç›¸å…³åˆ—"
                    )
                else:
                    # å¤šè¡¨æ ¼æ¨¡å¼
                    tables_info_lines = []
                    for idx, table_key in enumerate(selected_tables, 1):
                        df = st.session_state.table_dataframes[table_key]
                        local_vars[f'df_{idx}'] = df.copy()
                        columns_str = ", ".join(df.columns.tolist())
                        tables_info_lines.append(f"- è¡¨æ ¼{idx} (å˜é‡å: df_{idx}, æ¥æº: {table_key}): åˆ—å = [{columns_str}]")
                    
                    tables_info = "\n".join(tables_info_lines)
                    
                    # æ„å»ºå¤šè¡¨æ ¼Prompt
                    final_prompt = TABLE_ASSISTANT_SYSTEM_PROMPT_MULTI.format(
                        tables_info=tables_info,
                        processing_logic=processing_logic,
                        output_requirements=output_requirements if output_requirements.strip() else "ä¿ç•™æ‰€æœ‰ç›¸å…³åˆ—"
                    )
                
                with st.spinner("ğŸ¤– AIæ­£åœ¨åˆ†æéœ€æ±‚å¹¶ç”Ÿæˆä»£ç ..."):
                    try:
                        # è°ƒç”¨æ¨¡å‹ç”Ÿæˆä»£ç 
                        generated_code = call_gemini(final_prompt)
                        
                        if generated_code:
                            # æ¸…æ´—ä»£ç ï¼ˆé˜²æ­¢æ¨¡å‹å¸¦äº†markdownæ ‡è®°ï¼‰
                            code_to_run = generated_code.replace("```python", "").replace("```", "").strip()
                            
                            # æ˜¾ç¤ºç”Ÿæˆçš„ä»£ç ï¼ˆè°ƒè¯•ç”¨ï¼Œå¯é€‰ï¼‰
                            with st.expander("ğŸ” æŸ¥çœ‹ç”Ÿæˆçš„ä»£ç ", expanded=False):
                                st.code(code_to_run, language="python")
                            
                            # æ‰§è¡Œä»£ç 
                            with st.spinner("âš™ï¸ æ­£åœ¨æ‰§è¡Œæ•°æ®å¤„ç†..."):
                                exec(code_to_run, {}, local_vars)
                            
                            # æå–ç»“æœ
                            if 'result_df' in local_vars:
                                st.session_state.table_result_df = local_vars['result_df']
                                st.success("âœ… å¤„ç†å®Œæˆï¼")
                            else:
                                st.error("âŒ æ¨¡å‹ç”Ÿæˆçš„ä»£ç æœªå®šä¹‰ 'result_df' å˜é‡ï¼Œè¯·é‡è¯•ã€‚")
                        else:
                            st.error("âŒ AIæœªèƒ½ç”Ÿæˆæœ‰æ•ˆä»£ç ï¼Œè¯·é‡è¯•ã€‚")
                    
                    except Exception as e:
                        st.error(f"âŒ ä»£ç æ‰§è¡Œå‡ºé”™: {e}")
                        st.session_state.table_result_df = None
                
                st.session_state.table_is_processing = False
        
        # ç»“æœå±•ç¤ºåŒº
        if st.session_state.table_result_df is not None:
            st.markdown("---")
            st.markdown("#### ğŸ“Š å¤„ç†ç»“æœ")
            
            result_df = st.session_state.table_result_df
            st.info(f"ç»“æœæ•°æ®ï¼šå…± {len(result_df)} è¡Œï¼Œ{len(result_df.columns)} åˆ—")
            
            # Markdowné¢„è§ˆ
            st.markdown("**ç»“æœé¢„è§ˆï¼ˆå‰10è¡Œï¼‰ï¼š**")
            try:
                st.markdown(result_df.head(10).to_markdown(index=False))
            except:
                # å¦‚æœto_markdownä¸å¯ç”¨ï¼Œä½¿ç”¨dataframeæ˜¾ç¤º
                st.dataframe(result_df.head(10))
            
            # Excelä¸‹è½½
            import pandas as pd
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False)
            processed_data = output.getvalue()
            
            st.download_button(
                label="ğŸ“¥ ä¸‹è½½å¤„ç†åçš„Excel",
                data=processed_data,
                file_name="processed_result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    # ========== æ€è·¯å¼•å¯¼åŠ©æ‰‹ (linmo) æ¨¡å— ==========
    elif function_mode == "æ€è·¯å¼•å¯¼åŠ©æ‰‹ (linmo)":
        # æ ‡é¢˜å’Œæ¸…ç©ºæŒ‰é’®
        title_col, clear_col = st.columns([6, 1])
        with title_col:
            st.markdown("### ğŸ§  æ€è·¯å¼•å¯¼åŠ©æ‰‹ (Linmo)")
        with clear_col:
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºä¼šè¯", key="clear_linmo_session", use_container_width=True):
                clear_module_session("æ€è·¯å¼•å¯¼åŠ©æ‰‹ (linmo)")
                st.rerun()
        st.markdown("é€šè¿‡è‹æ ¼æ‹‰åº•å¼æé—®ï¼Œå¸®åŠ©æ‚¨ç†æ¸…æ€è·¯å¹¶ç”Ÿæˆç»“æ„åŒ–çš„æ€ç»´å¯¼å›¾ã€‚")
        
        # Mermaidåœ¨çº¿ç¼–è¾‘å™¨é“¾æ¥
        st.markdown("""
        ğŸ“ **Mermaid åœ¨çº¿ç¼–è¾‘å™¨**ï¼š
        - [Mermaid AI Dashboard](https://mermaid.ai/app/dashboard)
        - [Mermaid Live Editor](https://mermaid-live.nodejs.cn/edit)
        """)
        st.markdown("---")
        
        # Linmoæ¨¡å—çš„System Prompt
        LINMO_SYSTEM_PROMPT = """Role: ä½ æ˜¯ "Linmo" (æ€è·¯å¼•å¯¼åŠ©æ‰‹)ï¼Œä¸€ä½æ“…é•¿ç»“æ„åŒ–æ€ç»´å’Œè‹æ ¼æ‹‰åº•å¼æé—®çš„ä¸“å®¶ã€‚

å›å¤è¯­è¨€: è¯·å§‹ç»ˆä½¿ç”¨ä¸­æ–‡è¿›è¡Œå›ç­”å’Œè¾“å‡ºã€‚

Goal: ä½ çš„ç›®æ ‡æ˜¯å¸®åŠ©ç”¨æˆ·ç†æ¸…å¤æ‚çš„é—®é¢˜æˆ–æƒ³æ³•ï¼Œå°†å…¶è½¬åŒ–ä¸ºç»“æ„æ¸…æ™°çš„æ€ç»´å¯¼å›¾ã€‚

Process:
1.  **æ¥æ”¶è¾“å…¥**ï¼šç”¨æˆ·ä¼šè¾“å…¥ä¸€ä¸ªé—®é¢˜ã€ä¸€ç§å›°æ‰°æˆ–ä¸€äº›é›¶æ•£çš„æ€è·¯ã€‚
2.  **åˆ†æä¸æ„å»º**ï¼šåŸºäºç”¨æˆ·çš„ä¿¡æ¯ï¼Œæ„å»ºæˆ–æ›´æ–°ä¸€ä¸ªæ€ç»´å¯¼å›¾ç»“æ„ã€‚
3.  **è¿½é—®å¼•å¯¼**ï¼šä¸è¦ç›´æ¥ç»™å‡ºæ‰€æœ‰ç­”æ¡ˆã€‚ä½ éœ€è¦å‘ç°ç”¨æˆ·æ€è·¯ä¸­çš„æ¨¡ç³Šç‚¹ã€ç¼ºå¤±ç¯èŠ‚æˆ–é€»è¾‘è·³è·ƒï¼Œå¹¶æå‡º 1-2 ä¸ªå…³é”®çš„è¿½é—®ï¼Œå¼•å¯¼ç”¨æˆ·æ·±å…¥æ€è€ƒã€‚
4.  **å¾ªç¯è¿­ä»£**ï¼šç”¨æˆ·å›ç­”åï¼Œä½ å°†æ–°ä¿¡æ¯æ•´åˆè¿›æ€ç»´å¯¼å›¾ï¼Œå¹¶å†æ¬¡å±•ç¤ºæ›´æ–°åçš„ç»“æ„ï¼Œç›´åˆ°ç”¨æˆ·æ»¡æ„ã€‚

Output Format (Strict):
æ¯æ¬¡å›å¤å¿…é¡»åŒ…å«ä»¥ä¸‹ä¸‰ä¸ªéƒ¨åˆ†ï¼š

**Part 1: æ€è€ƒä¸åé¦ˆ**
ç®€è¦å›åº”ç”¨æˆ·çš„è¾“å…¥ï¼Œè¯´æ˜ä½ ç†è§£äº†ä»€ä¹ˆï¼Œä»¥åŠä½ ä¸ºä»€ä¹ˆè¦æ›´æ–°å¯¼å›¾çš„æŸä¸ªéƒ¨åˆ†ã€‚

**Part 2: å½“å‰æ€ç»´å¯¼å›¾ (Mermaid)**
è¯·å§‹ç»ˆä½¿ç”¨ Mermaid è¯­æ³•å±•ç¤º**å½“å‰å®Œæ•´**çš„æ€ç»´å¯¼å›¾ç»“æ„ã€‚
è¯·ä½¿ç”¨ `graph LR` (ä»å·¦åˆ°å³) æˆ– `graph TD` (ä»ä¸Šåˆ°ä¸‹) ç»“æ„ã€‚
```mermaid
graph LR
    A[æ ¸å¿ƒé—®é¢˜] --> B(åˆ†æ”¯1)
    A --> C(åˆ†æ”¯2)
    B --> B1(ç»†èŠ‚)
```

**Part 3: å¼•å¯¼è¿½é—®**
åŸºäºå½“å‰çš„å¯¼å›¾ï¼Œæå‡º 1-2 ä¸ªé—®é¢˜ï¼Œå¼•å¯¼ç”¨æˆ·è¡¥å……ä¸‹ä¸€å±‚çº§çš„ä¿¡æ¯æˆ–æ¾„æ¸…æ¨¡ç³Šç‚¹ã€‚
(ä¾‹å¦‚ï¼š"å…³äºåˆ†æ”¯Bï¼Œä½ è§‰å¾—å…·ä½“çš„æ‰§è¡Œéš¾ç‚¹åœ¨å“ªé‡Œï¼Ÿ" æˆ– "é™¤äº†ç›®å‰åˆ—å‡ºçš„ï¼Œè¿˜æœ‰å…¶ä»–å½±å“å› ç´ å—ï¼Ÿ")

Termination Condition:
å½“ç”¨æˆ·æ˜ç¡®è¡¨ç¤º"æ²¡æœ‰é—®é¢˜äº†"ã€"ç»“æ„å¾ˆå¥½äº†"æˆ–"ç”Ÿæˆæœ€ç»ˆç»“æœ"æ—¶ï¼š
1. åœæ­¢è¿½é—®ã€‚
2. è¾“å‡ºä¸€æ®µæ€»ç»“è¯­ã€‚
3. è¾“å‡ºæœ€ç»ˆç‰ˆæœ¬çš„ Mermaid ä»£ç å—ï¼Œç¡®ä¿è¯­æ³•å®Œç¾ï¼Œé€‚åˆå¤åˆ¶ã€‚

Tone:
ä¸“ä¸šã€è€å¿ƒã€å¼•å¯¼æ€§å¼ºã€é€»è¾‘ä¸¥å¯†ã€‚"""
        
        # åˆå§‹åŒ–linmoæ¨¡å—ä¸“ç”¨çš„session state
        if "linmo_chat_history" not in st.session_state:
            st.session_state.linmo_chat_history = []
        if "linmo_max_rounds" not in st.session_state:
            st.session_state.linmo_max_rounds = 10
        if "linmo_is_processing" not in st.session_state:
            st.session_state.linmo_is_processing = False
        if "linmo_input_key_counter" not in st.session_state:
            st.session_state.linmo_input_key_counter = 0
        
        # ä¾§è¾¹æ è®¾ç½®ï¼šæœ€å¤§å¯¹è¯è½®æ¬¡
        with st.sidebar:
            st.markdown("---")
            st.subheader("ğŸ§  Linmoå¯¹è¯è®¾ç½®")
            linmo_max_rounds = st.number_input(
                "æœ€å¤§å¯¹è¯è½®æ¬¡é™åˆ¶",
                min_value=1,
                max_value=50,
                value=st.session_state.linmo_max_rounds,
                step=1,
                help="ä¸€è½®å¯¹è¯ = ç”¨æˆ·å‘é€ + AIå›å¤",
                key="linmo_max_rounds_input"
            )
            st.session_state.linmo_max_rounds = linmo_max_rounds
            
            # æ˜¾ç¤ºå½“å‰è½®æ¬¡
            current_rounds = len([m for m in st.session_state.linmo_chat_history if m["role"] == "user"])
            st.info(f"å½“å‰è½®æ¬¡: {current_rounds} / {linmo_max_rounds}")
            
            # æ¸…ç©ºå¯¹è¯æŒ‰é’®
            if st.button("ğŸ—‘ï¸ é‡æ–°å¼€å§‹å¼•å¯¼", key="linmo_clear_chat", use_container_width=True):
                clear_module_session("æ€è·¯å¼•å¯¼åŠ©æ‰‹ (linmo)")
                st.rerun()
        
        # è®¡ç®—å½“å‰è½®æ¬¡ï¼ˆç”¨æˆ·æ¶ˆæ¯æ•°ï¼‰
        current_rounds = len([m for m in st.session_state.linmo_chat_history if m["role"] == "user"])
        max_rounds_reached = current_rounds >= st.session_state.linmo_max_rounds
        
        # èŠå¤©æ˜¾ç¤ºåŒº
        st.markdown("#### ğŸ’¬ å¯¹è¯åŒºåŸŸ")
        
        # è¾…åŠ©å‡½æ•°ï¼šä»æ–‡æœ¬ä¸­æå–Mermaidä»£ç 
        def extract_mermaid_code(text):
            """ä»æ–‡æœ¬ä¸­æå–Mermaidä»£ç å—"""
            import re
            # åŒ¹é… ```mermaid ... ``` ä»£ç å—
            pattern = r'```mermaid\s*([\s\S]*?)```'
            matches = re.findall(pattern, text)
            if matches:
                return matches[-1].strip()  # è¿”å›æœ€åä¸€ä¸ªMermaidä»£ç å—
            return None
        
        # æ˜¾ç¤ºå¯¹è¯å†å²
        chat_container = st.container()
        with chat_container:
            if not st.session_state.linmo_chat_history:
                # æ˜¾ç¤ºæ¬¢è¿è¯­
                with st.chat_message("assistant", avatar="ğŸ§ "):
                    st.markdown("ä½ å¥½ï¼Œæˆ‘æ˜¯ **Linmo**ã€‚è¯·å‘Šè¯‰æˆ‘ä½ ç°åœ¨é¢ä¸´çš„é—®é¢˜æˆ–æƒ³è¦æ‹†è§£çš„ç›®æ ‡ï¼Œæˆ‘æ¥å¸®ä½ æ¢³ç†æ€è·¯ã€‚\n\næˆ‘ä¼šé€šè¿‡æé—®å¼•å¯¼ä½ é€æ­¥ç†æ¸…æ€è·¯ï¼Œå¹¶ç”¨æ€ç»´å¯¼å›¾çš„å½¢å¼å±•ç¤ºç»“æ„ã€‚")
            else:
                for msg in st.session_state.linmo_chat_history:
                    if msg["role"] == "user":
                        with st.chat_message("user"):
                            st.markdown(msg["content"])
                    else:
                        with st.chat_message("assistant", avatar="ğŸ§ "):
                            st.markdown(msg["content"])
        
        # æ£€æµ‹æœ€åä¸€æ¡AIå›å¤æ˜¯å¦åŒ…å«Mermaidä»£ç ï¼Œæ˜¾ç¤ºè·³è½¬æŒ‰é’®
        if st.session_state.linmo_chat_history:
            # è·å–æœ€åä¸€æ¡AIå›å¤
            ai_messages = [m for m in st.session_state.linmo_chat_history if m["role"] == "assistant"]
            if ai_messages:
                last_ai_message = ai_messages[-1]["content"]
                mermaid_code = extract_mermaid_code(last_ai_message)
                if mermaid_code:
                    st.markdown("---")
                    st.info("ğŸ‰ æ£€æµ‹åˆ°æ€ç»´å¯¼å›¾å·²ç”Ÿæˆï¼Œæ‚¨å¯ä»¥å°†å…¶ç”¨äºç”Ÿæˆå®Œæ•´çš„ç­–åˆ’æ¡ˆï¼")
                    col_jump, col_copy = st.columns([1, 1])
                    with col_jump:
                        if st.button("ğŸš€ è·³è½¬åˆ°ã€Œè„‘å›¾ç”Ÿæˆç­–åˆ’æ¡ˆã€", key="linmo_jump_to_mindmap", use_container_width=True):
                            # å°†mermaidä»£ç å­˜å…¥session stateï¼Œä¾›è„‘å›¾æ¨¡å—ä½¿ç”¨
                            st.session_state.linmo_to_mindmap_mermaid = mermaid_code
                            st.session_state.selected_function = "è„‘å›¾ç”Ÿæˆç­–åˆ’æ¡ˆ"
                            st.rerun()
                    with col_copy:
                        st.markdown(f"ğŸ“‹ **Mermaidä»£ç é¢„è§ˆ**ï¼ˆå¯å¤åˆ¶ï¼‰")
                        st.code(mermaid_code, language="mermaid")
        
        # è½®æ¬¡è¾¾åˆ°ä¸Šé™æç¤º
        if max_rounds_reached:
            st.warning(f'âš ï¸ å¯¹è¯è½®æ¬¡å·²è¾¾ä¸Šé™ï¼ˆ{st.session_state.linmo_max_rounds}è½®ï¼‰ï¼Œè¯·ç‚¹å‡»ä¾§è¾¹æ çš„"é‡æ–°å¼€å§‹å¼•å¯¼"æŒ‰é’®é‡æ–°å¼€å§‹ã€‚')
        
        # è¾“å…¥åŒº - ä½¿ç”¨st.chat_inputï¼Œåªåœ¨æŒ‰Enteræ—¶è§¦å‘
        if max_rounds_reached or st.session_state.linmo_is_processing:
            st.chat_input("å¯¹è¯è½®æ¬¡å·²è¾¾ä¸Šé™æˆ–æ­£åœ¨å¤„ç†ä¸­...", disabled=True, key="linmo_chat_disabled")
            linmo_user_input = None
        else:
            linmo_user_input = st.chat_input(
                "è¯·è¾“å…¥å½“å‰é¢ä¸´çš„é—®é¢˜ï¼Œæˆ–å¯¹å½“å‰æ€ç»´å¯¼å›¾çš„ä¿®æ”¹å»ºè®®...",
                key=f"linmo_chat_input_{st.session_state.linmo_input_key_counter}"
            )
        
        # å¤„ç†ç”¨æˆ·è¾“å…¥ï¼ˆst.chat_inputåªåœ¨æŒ‰Enteræ—¶è¿”å›éNoneï¼‰
        if linmo_user_input:
            st.session_state.linmo_is_processing = True
            
            # æ·»åŠ ç”¨æˆ·æ¶ˆæ¯åˆ°å†å²
            st.session_state.linmo_chat_history.append({
                "role": "user",
                "content": linmo_user_input
            })
            
            # æ„å»ºå®Œæ•´çš„å¯¹è¯ä¸Šä¸‹æ–‡
            # System Prompt + å†å²å¯¹è¯ + å½“å‰è¾“å…¥
            messages_context = ""
            for msg in st.session_state.linmo_chat_history:
                if msg["role"] == "user":
                    messages_context += f"\n\nã€ç”¨æˆ·ã€‘\n{msg['content']}"
                else:
                    messages_context += f"\n\nã€Linmoã€‘\n{msg['content']}"
            
            full_prompt = f"""è¯·åŸºäºä»¥ä¸‹å¯¹è¯å†å²ç»§ç»­å¼•å¯¼ç”¨æˆ·ï¼š
{messages_context}

è¯·ä»¥æ€è·¯å¼•å¯¼åŠ©æ‰‹Linmoçš„èº«ä»½å›å¤ï¼Œä¸¥æ ¼æŒ‰ç…§è¾“å‡ºæ ¼å¼è¦æ±‚ï¼ˆæ€è€ƒä¸åé¦ˆã€Mermaidæ€ç»´å¯¼å›¾ã€å¼•å¯¼è¿½é—®ï¼‰è¿›è¡Œå›å¤ã€‚"""
            
            # æµå¼ç”Ÿæˆå›å¤
            st.markdown("#### ğŸ¤– Linmoæ­£åœ¨æ€è€ƒ...")
            
            # æ€è€ƒè¿‡ç¨‹å®¹å™¨
            thinking_expander = st.expander("ğŸ’­ æŸ¥çœ‹æ¨¡å‹æ€è€ƒè¿‡ç¨‹", expanded=False)
            with thinking_expander:
                thinking_container = st.empty()
            
            response_container = st.empty()
            full_response = ""
            thinking_text = ""
            
            for chunk in call_gemini_stream(full_prompt, LINMO_SYSTEM_PROMPT):
                if chunk["type"] == "text":
                    full_response += chunk["content"]
                    response_container.markdown(full_response + " â–Œ")
                elif chunk["type"] == "thinking":
                    thinking_text += chunk["content"]
                    with thinking_expander:
                        thinking_container.markdown(thinking_text)
                elif chunk["type"] == "error":
                    st.error(f"ç”Ÿæˆå¤±è´¥: {chunk['content']}")
                    break
            
            if full_response:
                response_container.markdown(full_response)
                # æ·»åŠ AIå›å¤åˆ°å†å²
                st.session_state.linmo_chat_history.append({
                    "role": "assistant",
                    "content": full_response
                })
            
            st.session_state.linmo_is_processing = False
            # æ¸…ç©ºè¾“å…¥æ¡†ï¼ˆé€šè¿‡å¢åŠ è®¡æ•°å™¨æ”¹å˜keyï¼Œå¼ºåˆ¶é‡å»ºç»„ä»¶ï¼‰
            st.session_state.linmo_input_key_counter += 1
            st.rerun()
    
    # ========== PUBGM WoW ç©æ³•è¯„å®¡æ¨¡å— ==========
    elif function_mode == "PUBGM WoW ç©æ³•è¯„å®¡":
        # æ ‡é¢˜å’Œæ¸…ç©ºæŒ‰é’®
        title_col, clear_col = st.columns([6, 1])
        with title_col:
            st.markdown("### ğŸ® PUBGM WoW ç©æ³•è¯„å®¡")
        with clear_col:
            if st.button("ğŸ—‘ï¸ æ¸…ç©ºä¼šè¯", key="clear_wow_session", use_container_width=True):
                clear_module_session("PUBGM WoW ç©æ³•è¯„å®¡")
                st.rerun()
        
        st.markdown("è¯·ä¸Šä¼  PUBG Mobile World of Wonder (WoW) æ¨¡å¼çš„æ¸¸ç©è§†é¢‘ï¼ŒAI å°†åˆ†æç©æ³•å¹¶ç»™å‡ºè¯„åˆ†ã€‚")
        st.info("ğŸ’¡ æç¤ºï¼šå»ºè®®ä¸Šä¼ 1-3åˆ†é’Ÿçš„çŸ­è§†é¢‘ï¼Œè§†é¢‘è¶Šé•¿å¤„ç†æ—¶é—´è¶Šä¹…ã€‚æ”¯æŒæ ¼å¼ï¼šMP4, MOV, AVI, WEBM")
        
        # åˆå§‹åŒ–session state
        if "wow_review_result" not in st.session_state:
            st.session_state.wow_review_result = ""
        if "wow_is_processing" not in st.session_state:
            st.session_state.wow_is_processing = False
        if "wow_uploaded_video" not in st.session_state:
            st.session_state.wow_uploaded_video = None
        
        # è§†é¢‘ä¸Šä¼ åŒº
        uploaded_video = st.file_uploader(
            "ä¸Šä¼ æ¸¸æˆè§†é¢‘",
            type=['mp4', 'mov', 'avi', 'webm'],
            key="wow_video_uploader",
            help="æ”¯æŒ MP4, MOV, AVI, WEBM æ ¼å¼ï¼Œå»ºè®®æ–‡ä»¶å¤§å°ä¸è¶…è¿‡200MB"
        )
        
        # å¼€å§‹è¯„å®¡æŒ‰é’®
        start_review = st.button(
            "ğŸ¬ å¼€å§‹AIè¯„å®¡",
            key="wow_start_review",
            type="primary",
            disabled=uploaded_video is None or st.session_state.wow_is_processing
        )
        
        # å¤„ç†è¯„å®¡é€»è¾‘
        if start_review and uploaded_video and not st.session_state.wow_is_processing:
            st.session_state.wow_is_processing = True
            st.session_state.wow_review_result = ""
            
            # WoW è¯„å®¡ä¸“ç”¨çš„ System Prompt
            WOW_REVIEW_PROMPT = """Role: ä½ æ˜¯ä¸€ä½èµ„æ·±çš„ PUBG Mobile æ¸¸æˆè¯„æµ‹ä¸“å®¶ï¼Œä¸“æ³¨äº "World of Wonder" (WoW) UGC ç¼–è¾‘å™¨æ¨¡å¼çš„ç©æ³•è¯„å®¡ã€‚ä½ æ‹¥æœ‰æ•é”çš„æ¸¸æˆè®¾è®¡å—…è§‰ï¼Œèƒ½é€šè¿‡è§‚çœ‹è§†é¢‘å¿«é€Ÿç†è§£æ ¸å¿ƒæœºåˆ¶ã€‚è¯·ç”¨ä¸­æ–‡å›ç­”ã€‚

Task: è¯·ä»”ç»†è§‚çœ‹ä¸Šä¼ çš„è§†é¢‘ï¼Œåˆ†æè¯¥ UGC åœ°å›¾çš„ç©æ³•ï¼Œå¹¶è¾“å‡ºä¸€ä»½ä¸“ä¸šçš„è¯„å®¡æŠ¥å‘Šã€‚

Output Format (Markdown):

# ğŸ® PUBGM WoW ç©æ³•è¯„å®¡æŠ¥å‘Š

## 1. ç©æ³•ç®€è¿° (Gameplay Summary)
[è¯·åœ¨æ­¤å¤„ç”¨ç®€ç»ƒçš„è¯­è¨€æè¿°è¿™ä¸ªåœ°å›¾æ˜¯æ€ä¹ˆç©çš„ã€‚åŒ…æ‹¬ï¼šèƒœåˆ©æ¡ä»¶ã€æ ¸å¿ƒæœºåˆ¶ã€ç©å®¶ä¸»è¦åœ¨åšä»€ä¹ˆã€‚]

## 2. ç»´åº¦è¯„åˆ† (Scoring)
è¯·åŸºäº 0-10 åˆ†è¿›è¡Œæ‰“åˆ†ï¼Œå¹¶ç»™å‡ºç®€çŸ­çš„ç†ç”±ã€‚

| ç»´åº¦ | è¯„åˆ† (0-10) | è¯„ä»·ç†ç”± |
| :--- | :--- | :--- |
| **åˆ›æ–°æ€§ (Innovation)** | [x.x] | [æ˜¯å¦è„±ç¦»äº†å¸¸è§„ç©æ³•ï¼Ÿæœºåˆ¶æ˜¯å¦æ–°é¢–ï¼Ÿ] |
| **åœºæ™¯ç¾è§‚æ€§ (Aesthetics)** | [x.x] | [åœ°å›¾æ­å»ºæ˜¯å¦ç²¾ç»†ï¼Ÿå…‰å½±ã€è‰²å½©å’Œå»ºç­‘ç»“æ„å¦‚ä½•ï¼Ÿ] |
| **å…³å¡è®¾è®¡ (Level Design)** | [x.x] | [æµç¨‹æ˜¯å¦åˆç†ï¼Ÿéš¾åº¦æ›²çº¿å¦‚ä½•ï¼Ÿæ˜¯å¦æœ‰å¼•å¯¼ï¼Ÿ] |
| **æ¸¸æˆå®Œæ•´æ€§ (Completeness)** | [x.x] | [UIæ˜¯å¦å®Œå–„ï¼Ÿæ˜¯å¦æœ‰æ˜æ˜¾Bugï¼Ÿä½“éªŒæ˜¯å¦é—­ç¯ï¼Ÿ] |

## 3. ç»¼åˆè¯„ä»· (Final Verdict)
**æœ€ç»ˆå¹³å‡åˆ†ï¼š[è®¡ç®—ä¸Šè¿°4é¡¹çš„å¹³å‡åˆ†] / 10**

**æ€»ç»“ç‚¹è¯„ï¼š**
[è¯·ç»™å‡ºä¸€æ®µæ€»ç»“æ€§çš„è¯„ä»·ï¼ŒæŒ‡å‡ºè¿™ä¸ªä½œå“æœ€å¤§çš„äº®ç‚¹æ˜¯ä»€ä¹ˆï¼Œä»¥åŠæœ€éœ€è¦æ”¹è¿›çš„ä¸€ä¸ªåœ°æ–¹ã€‚]
"""
            
            with st.spinner("æ­£åœ¨ä¸Šä¼ è§†é¢‘å¹¶è¿›è¡ŒAIåˆ†æï¼Œè¯·ç¨å€™...ï¼ˆè§†é¢‘è¶Šé•¿è€—æ—¶è¶Šä¹…ï¼‰"):
                temp_file_path = None
                uploaded_file_obj = None
                
                try:
                    # åˆ›å»º Client å®ä¾‹
                    client = genai.Client(api_key=st.session_state.api_key)
                    
                    # 1. ä¸´æ—¶ä¿å­˜è§†é¢‘æ–‡ä»¶
                    suffix = "." + uploaded_video.name.split(".")[-1].lower()
                    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_file:
                        tmp_file.write(uploaded_video.read())
                        temp_file_path = tmp_file.name
                    
                    st.info("ğŸ“¤ æ­£åœ¨ä¸Šä¼ è§†é¢‘åˆ°AIæœåŠ¡...")
                    
                    # 2. ä¸Šä¼ è§†é¢‘åˆ° Gemini File API (ä½¿ç”¨ client.files.upload)
                    uploaded_file_obj = client.files.upload(
                        file=temp_file_path,
                        config={"display_name": "WoW_Gameplay"}
                    )
                    
                    st.info("â³ è§†é¢‘æ­£åœ¨å¤„ç†ä¸­ï¼Œè¯·è€å¿ƒç­‰å¾…...")
                    
                    # 3. ç­‰å¾…è§†é¢‘å¤„ç†å®Œæˆ
                    while uploaded_file_obj.state.name == "PROCESSING":
                        time.sleep(2)
                        uploaded_file_obj = client.files.get(name=uploaded_file_obj.name)
                    
                    if uploaded_file_obj.state.name == "FAILED":
                        st.error("âŒ è§†é¢‘å¤„ç†å¤±è´¥ï¼Œè¯·å°è¯•ä¸Šä¼ å…¶ä»–è§†é¢‘ã€‚")
                        st.session_state.wow_is_processing = False
                    elif uploaded_file_obj.state.name == "ACTIVE":
                        st.info("ğŸ¤– AI æ­£åœ¨åˆ†æè§†é¢‘å†…å®¹...")
                        
                        # 4. è°ƒç”¨æ¨¡å‹ç”Ÿæˆè¯„å®¡æŠ¥å‘Š
                        # è·å–å½“å‰é€‰æ‹©çš„æ¨¡å‹
                        current_model = st.session_state.get("selected_model", "gemini-2.0-flash")
                        
                        response = client.models.generate_content(
                            model=current_model,
                            contents=[uploaded_file_obj, WOW_REVIEW_PROMPT]
                        )
                        
                        if response and response.text:
                            st.session_state.wow_review_result = response.text
                            st.success("âœ… è¯„å®¡å®Œæˆï¼")
                        else:
                            st.error("âŒ AI æœªèƒ½ç”Ÿæˆè¯„å®¡ç»“æœï¼Œè¯·é‡è¯•ã€‚")
                    else:
                        st.error(f"âŒ è§†é¢‘çŠ¶æ€å¼‚å¸¸: {uploaded_file_obj.state.name}")
                
                except Exception as e:
                    st.error(f"âŒ è¯„å®¡è¿‡ç¨‹ä¸­å‡ºé”™: {str(e)}")
                
                finally:
                    # 5. æ¸…ç†ï¼šåˆ é™¤æœ¬åœ°ä¸´æ—¶æ–‡ä»¶
                    if temp_file_path and os.path.exists(temp_file_path):
                        try:
                            os.remove(temp_file_path)
                        except:
                            pass
                    
                    # å¯é€‰ï¼šåˆ é™¤äº‘ç«¯æ–‡ä»¶
                    if uploaded_file_obj:
                        try:
                            client.files.delete(name=uploaded_file_obj.name)
                        except:
                            pass
                    
                    st.session_state.wow_is_processing = False
        
        # æ˜¾ç¤ºè¯„å®¡ç»“æœ
        if st.session_state.wow_review_result:
            st.markdown("---")
            st.markdown("## ğŸ“‹ è¯„å®¡æŠ¥å‘Š")
            with st.chat_message("assistant", avatar="ğŸ®"):
                st.markdown(st.session_state.wow_review_result)
            
            # å¤åˆ¶æŒ‰é’®
            st.download_button(
                label="ğŸ“¥ ä¸‹è½½è¯„å®¡æŠ¥å‘Š",
                data=st.session_state.wow_review_result,
                file_name=f"WoW_Review_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md",
                mime="text/markdown"
            )
    
    # é¡µè„š
    st.markdown("---")
    st.markdown(
        "<div style='text-align: center; color: gray;'>"
        "ğŸ® æ¸¸æˆç­–åˆ’Agentï¼ˆé…¸å¥¶ï¼‰ | Powered by Gemini API"
        "</div>",
        unsafe_allow_html=True
    )


if __name__ == "__main__":
    main()
